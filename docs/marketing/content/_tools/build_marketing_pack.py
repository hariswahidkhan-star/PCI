#!/usr/bin/env python3
"""Build the marketing team's working pack: a posting tracker and a copy pack.

Everything built so far serves a reviewer or a developer. This serves the person who has to
actually post 347 things and needs to know what, where, when, with which link, and what they
are not allowed to change.

The order is not invented. _BRIEF.md section 5 sets one real dependency — publish on the PCI
site first, let it index, then republish elsewhere with the canonical pointing home —
because a republication that goes out before its origin is indexed hands the ranking to a
platform with more authority than yours. Phases follow that dependency; the week numbers are
a suggested cadence a person can change, and no calendar dates are invented.
"""
import importlib.util
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
_b = importlib.util.spec_from_file_location("bb", HERE / "build_bundle.py")
bb = importlib.util.module_from_spec(_b); _b.loader.exec_module(bb)

NAVY, GOLD, CRIMSON = "1D3C92", "B8923E", "C13329"
WHITE, SUNKEN, LINE = "FFFFFF", "F1F3F9", "DDE2EE"
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def phase_of(r):
    """Which wave a piece belongs to, from the publish-then-republish dependency."""
    p = r["platform"].lower()
    if r["path"].parts[-2] == "flagship":
        return 1, "Launch — flagship assets"
    if p.startswith("own site"):
        return 2, "Own site — publish and let it index"
    if any(k in p for k in ("medium", "dev community", "hashnode", "blogger", "wordpress")):
        return 4, "Republish with canonical (only after its origin has indexed)"
    if any(k in p for k in ("linkedin article", "substack", "vocal")):
        return 3, "Off-site originals"
    return 5, "Social amplification"


def clean(md, limit=None, keep_urls=True):
    """Markdown to plain paste-ready text, with the links kept.

    An earlier cut dropped the URL and kept only the anchor text, on the reasoning that the
    link the piece carries is already in its own column. That was wrong: it left a poster
    copying a cell with a sentence whose link had silently vanished, and only 58 of 347 rows
    still had a URL anywhere in the copy. An inline link is placed where it is because that
    sentence raises the question the target answers, so the position matters as much as the
    address. Markdown links now render as "anchor text (url)" — the sentence still reads, and
    where the link belongs is visible. Bare URLs pass through untouched, which is what social
    posts and first comments want.
    """
    t = re.sub(r"^#{1,6}\s*", "", md, flags=re.M)
    if keep_urls:
        t = re.sub(r"\[([^\]]*)\]\((https?://[^)]*)\)", r"\1 (\2)", t)
    else:
        t = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", t)
    t = re.sub(r"[*_`]", "", t)
    t = re.sub(r"^\s*>\s?", "", t, flags=re.M)
    t = re.sub(r"\n{3,}", "\n\n", t).strip()
    if limit and len(t) > limit:
        t = t[:limit].rsplit(" ", 1)[0] + " …"
    return t


def head(ws, row, cols, fill=NAVY, color=WHITE, size=10):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=color, size=size, name="Calibri")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def build():
    files = bb.load()
    for r in files:
        m = bb.FM.match(r["path"].read_text(encoding="utf-8"))
        fm = m.group(1) if m else ""
        r["cta"] = bb.g(fm, "cta_link")
        r["when"] = bb.g(fm, "when_to_post")
        r["phase"], r["phase_name"] = phase_of(r)

    files.sort(key=lambda r: (r["phase"], r["group"], r["path"].name))

    wb = Workbook()

    # ---------------- Read me ----------------
    ws = wb.active
    ws.title = "Read me"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104
    rows = [
        ("h", "PCI content posting pack"),
        ("p", f"{len(files)} pieces to publish across {len({r['group'] for r in files})} surfaces. "
              "Everything you need to post is on the Posting schedule tab."),
        ("s", "How to use the Posting schedule"),
        ("p", "One row per piece, in the order to post. Filter by Phase or Platform to give a "
              "person a week's work. Fill in Owner, Posted on and Status as you go — those three "
              "columns are yours and nothing else in the pack depends on them."),
        ("p", "The Copy column holds the exact text to paste for anything short. For long "
              "articles it says which file to open in the copy pack, because a 2,000-word article "
              "does not belong in a spreadsheet cell."),
        ("p", "The Link to use column is the one URL that piece should carry. It has been checked "
              "against the pages that actually exist. Do not swap it for a different page."),
        ("s", "The one sequencing rule that matters"),
        ("p", "Publish on the PCI site first and let it index. Only then republish on Medium, "
              "DEV, Hashnode, Blogger or WordPress with the canonical pointing home. A "
              "republication that goes out before its origin is indexed hands the ranking to a "
              "platform with far more authority than ours, and we end up below our own article."),
        ("p", "That is why the phases are ordered the way they are. Weeks are a suggested cadence "
              "and yours to change; the order within them is not."),
        ("s", "Four things that must not be changed, on any platform"),
        ("p", "1.  15,613 machine calculation checks — only ever in a sentence that also says it "
              "covers PFL-AI and PML-AI. PCL-AI has no equivalent suite. Split from that scope the "
              "number is simply false."),
        ("p", "2.  40/40/20 describes the Body of Knowledge, never the examination. No exam "
              "weighting exists to publish, because the blueprint is still an open decision."),
        ("p", "3.  No accreditation, recognition, endorsement, affiliation or partnership may be "
              "claimed or implied. PCI holds none and says so."),
        ("p", "4.  No pass rates, student numbers, salary figures or worked-example counts. If a "
              "number cannot be pointed at, the sentence goes."),
        ("s", "Links — why there are so few"),
        ("p", "No piece links to all five domains, and none carries more than one link to any "
              "single other domain. Five commonly-owned sites each linking to the other four on "
              "every page is the pattern search engines treat as a link network, and the cost is "
              "all five losing standing together. Adding links will not help and can cost the lot."),
        ("s", "Tabs"),
        ("p", "Posting schedule — the working list.    By platform — how much work sits where.    "
              "Links — every link in the run, by destination."),
    ]
    rr = 2
    for kind, text in rows:
        c = ws.cell(row=rr, column=2, value=text)
        if kind == "h":
            c.font = Font(bold=True, size=20, color=NAVY, name="Calibri")
            ws.row_dimensions[rr].height = 30
        elif kind == "s":
            c.font = Font(bold=True, size=12, color=CRIMSON, name="Calibri")
            ws.row_dimensions[rr].height = 26
        else:
            c.font = Font(size=10.5, name="Calibri")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[rr].height = max(16, 15 * (len(text) // 100 + 1))
        rr += 1 if kind != "s" else 1
        if kind == "s":
            rr += 0
    ws.sheet_view.showGridLines = False

    # ---------------- Posting schedule ----------------
    ws = wb.create_sheet("Posting schedule")
    cols = ["#", "Phase", "Week", "Platform", "Type", "Title", "Primary keyword",
            "Link to use", "Hashtags", "Length", "Copy — paste this", "Timing notes",
            "Owner", "Posted on", "Status"]
    head(ws, 1, cols)
    widths = [5, 34, 7, 26, 15, 46, 26, 46, 34, 9, 62, 40, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    per_week = 12
    for n, r in enumerate(files, 1):
        week = (n - 1) // per_week + 1
        # The call to action the piece was written around, where it declares one. Falling
        # straight to links[0] picked whichever estate link happened to come first in the
        # prose, which on several pieces is a supporting reference rather than the ask.
        link = ""
        if r["cta"]:
            m = re.search(r"https?://\S+", r["cta"])
            if m:
                link = m.group(0).rstrip(".,;)")
        if not link and r["links"]:
            l = r["links"][0]
            link = f"https://{l['domain']}{l['path']}"
        # Put the copy in the cell wherever it fits. An Excel cell holds 32,767 characters,
        # so the only pieces that genuinely need the copy pack are full-length articles; a
        # 700-word carousel script is perfectly pasteable and sending someone to a second
        # document for it just costs them a page-flip per post. The cap is set well below the
        # cell limit so a long cell stays readable rather than merely legal.
        txt = clean(r["body"])
        copy = txt if len(txt) <= 7000 else \
            f"Full article — see entry {n} in the copy pack ({r['words']:,} words)"
        vals = [n, r["phase_name"], week, r["platform"][:110], r["type"], r["title"],
                r["kw"], link, "" if r["hashtags"].lower().startswith(("n/a", "none"))
                else r["hashtags"][:220], f"{r['words']:,} w", copy,
                clean(r["when"], 600, keep_urls=False), "", "", ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=n + 1, column=i, value=v)
            c.border = BORDER
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(wrap_text=i in (2, 4, 6, 7, 8, 9, 11, 12),
                                    vertical="top", horizontal="center" if i in (1, 3, 10) else "left")
            if i == 1:
                c.font = Font(size=10, bold=True, color=GOLD, name="Calibri")
            if i in (13, 14, 15):
                c.fill = PatternFill("solid", fgColor="FFFDF5")
        ws.row_dimensions[n + 1].height = 58
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:O{len(files)+1}"

    # ---------------- By platform ----------------
    ws = wb.create_sheet("By platform")
    head(ws, 1, ["Platform", "Phase", "Pieces", "Words", "Pieces carrying a link"])
    for i, w in enumerate([44, 34, 10, 12, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    agg = {}
    for r in files:
        k = (r["group"], r["phase_name"])
        a = agg.setdefault(k, [0, 0, 0])
        a[0] += 1; a[1] += r["words"]; a[2] += 1 if r["links"] else 0
    for n, ((grp, ph), a) in enumerate(sorted(agg.items(), key=lambda x: (-x[1][0], x[0])), 2):
        for i, v in enumerate([grp, ph, a[0], a[1], a[2]], 1):
            c = ws.cell(row=n, column=i, value=v)
            c.border = BORDER; c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(wrap_text=i in (1, 2), vertical="center")
    ws.freeze_panes = "A2"

    # ---------------- Links ----------------
    ws = wb.create_sheet("Links")
    head(ws, 1, ["Destination domain", "Page", "Anchor text used", "In which piece", "Platform"])
    for i, w in enumerate([28, 40, 52, 46, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    n = 2
    for r in files:
        for l in r["links"]:
            for i, v in enumerate([l["domain"], l["path"] or "/", l["anchor"] or "(bare URL)",
                                   r["title"], r["platform"][:80]], 1):
                c = ws.cell(row=n, column=i, value=v)
                c.border = BORDER; c.font = Font(size=10, name="Calibri")
                c.alignment = Alignment(wrap_text=i in (3, 4, 5), vertical="top")
            n += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{n-1}"

    dest = bb.ROOT.parent / "PCI-content-posting-pack.xlsx"
    wb.save(dest)
    return dest, len(files), n - 2


if __name__ == "__main__":
    d, n, links = build()
    print(f"{d}  ({d.stat().st_size/1024:.0f} KB)")
    print(f"{n} pieces scheduled | {links} link rows")
