#!/usr/bin/env python3
"""Comprehensive 10-lens audit run in-process (replaces the judge agents).

Lenses: formulas/ranges, DV integrity, content cross-refs, protection,
article-bank quality, KPI wiring, print/export, guides coherence, example
rows, navigation. Prints findings only — read-only on the workbook.
"""
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "PCI_AI_Growth_OS_V9.xlsx"
wb = openpyxl.load_workbook(SRC)
names = set(wb.sheetnames)
F = []


def f(sev, where, msg):
    F.append((sev, where, msg))


# ---------------------------------------------------------------- L1 formulas
sheet_ref = re.compile(r"'([^']+)'!")
bad_refs = set()
formula_cells = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                formula_cells.append((ws.title, c.coordinate, v))
                for s in sheet_ref.findall(v):
                    if s not in names:
                        bad_refs.add((ws.title, c.coordinate, s))
                if "#REF" in v or "#NAME" in v:
                    f("BLOCKER", f"{ws.title}!{c.coordinate}", "broken reference in formula")
for t in list(bad_refs)[:10]:
    f("BLOCKER", f"{t[0]}!{t[1]}", f"references missing sheet {t[2]!r}")
print(f"L1 formulas scanned: {len(formula_cells)}; bad sheet refs: {len(bad_refs)}")

# layout endpoints that must not regress
LAYOUT = {"'DAILY ENTRY'": 1006, "'LinkedIn Outreach'": 1203, "'Content Calendar'": 403,
          "'Community & PR'": 403, "'Partnership Pipeline'": 403, "'Job Postings'": 203,
          "'Link Building'": 403, "'Content Scheduler'": 103}
end_re = re.compile(r"('(?:[^']+)')!\$?[A-Z]{1,2}\$?(\d+):\$?[A-Z]{1,2}\$?(\d+)")
odd = Counter()
for sheet, coord, v in formula_cells:
    for m in end_re.finditer(v):
        s, lo, hi = m.group(1), int(m.group(2)), int(m.group(3))
        if s in LAYOUT and hi != LAYOUT[s] and hi > 50:
            odd[(s, hi)] += 1
for (s, hi), n in odd.most_common(8):
    f("MAJOR", "formulas", f"{n} formulas end {s} at row {hi}, layout says {LAYOUT[s]}")

# ------------------------------------------------- L6 Dashboard KPI wiring
wp = wb["Weekly Pulse"]
pulse_labels = {r: wp.cell(r, 1).value for r in range(6, 25) if wp.cell(r, 1).value}
db = wb["Dashboard"]
EXPECT = {"REVENUE THIS WEEK": "Revenue recorded (USD)", "MEETINGS THIS WEEK": "Meetings booked",
          "MINUTES THIS WEEK": "Minutes logged", "BACKLINKS THIS WEEK": "Backlinks gone live",
          "CONTENT THIS WEEK": "Content published"}
for r in range(3, 30):
    lab = db.cell(r, 6).value
    if isinstance(lab, str) and lab in EXPECT:
        formula = str(db.cell(r + 1, 6).value or "")
        m = re.search(r"'Weekly Pulse'!\$B\$(\d+)", formula)
        if not m:
            f("MAJOR", f"Dashboard!F{r+1}", f"tile {lab} does not read Weekly Pulse")
            continue
        got = pulse_labels.get(int(m.group(1)))
        if got != EXPECT[lab]:
            f("BLOCKER", f"Dashboard!F{r+1}",
              f"tile {lab} reads Weekly Pulse row {m.group(1)} = {got!r}, expected {EXPECT[lab]!r}")
print("L6 dashboard tiles checked against pulse labels")

# ------------------------------------------------------------------- L2 DVs
ls = wb["Lists"]
col_extent = {}
for c in range(1, 26):
    n = 0
    for r in range(4, 400):
        if ls.cell(r, c).value not in (None, ""):
            n = r
    col_extent[c] = n
dv_total = 0
for ws in wb.worksheets:
    for dv in ws.data_validations.dataValidation:
        dv_total += 1
        f1 = str(dv.formula1 or "")
        if not dv.showErrorMessage:
            f("MINOR", f"{ws.title}", f"DV {dv.sqref} has no error message")
        m = re.match(r"Lists!\$([A-Z])\$(\d+):\$([A-Z])\$(\d+)", f1)
        if m:
            col = openpyxl.utils.column_index_from_string(m.group(1))
            hi = int(m.group(4))
            ext = col_extent.get(col, 0)
            if ext and hi != ext and not (col == 18):     # R = roster selector (intentional)
                f("MAJOR", f"{ws.title}!{dv.sqref}",
                  f"DV covers Lists {m.group(1)}4:{hi} but column holds values to row {ext}")
print(f"L2 data validations checked: {dv_total}")

# ------------------------------------ L3 the first grid row must be EMPTY
# The logs used to ship a worked example on their first row while every report
# range started below it, so typing over the example lost the work silently.
# The examples now live on TEAM GUIDE and the grids start empty.
GRID_FIRST = ["DAILY ENTRY", "LinkedIn Outreach", "Content Calendar",
              "Community & PR", "Partnership Pipeline", "Job Postings",
              "Link Building", "Content Scheduler", "Experiments", "UTM Builder"]
lov = {c: [ls.cell(r, c).value for r in range(4, col_extent.get(c, 3) + 1)] for c in range(1, 26)}
for sheet in GRID_FIRST:
    ws = wb[sheet]
    dirty = [ws.cell(4, c).coordinate for c in range(1, ws.max_column + 1)
             if ws.cell(4, c).value is not None
             and not str(ws.cell(4, c).value).startswith("=")]
    if dirty:
        f("BLOCKER", sheet, f"first grid row is not empty: {dirty[:5]}")
print("L3 first grid rows checked (must ship empty)")

# ------------------------------------------------------- L4 protection
EXAMPLE_UNLOCK = {}          # no example rows survive: every formula is locked
unlocked_formula = []
for ws in wb.worksheets:
    if not ws.protection.sheet:
        f("BLOCKER", ws.title, "sheet not protected")
    elif not ws.protection.password:
        f("BLOCKER", ws.title, "sheet protected without password")
    keep = EXAMPLE_UNLOCK.get(ws.title, set())
    for row in ws.iter_rows():
        for c in row:
            if (isinstance(c.value, str) and c.value.startswith("=")
                    and c.protection.locked is False and c.row not in keep):
                unlocked_formula.append(f"{ws.title}!{c.coordinate}")
if unlocked_formula:
    f("MAJOR", "protection", f"{len(unlocked_formula)} formula cells unlocked, e.g. "
      + ", ".join(unlocked_formula[:6]))
if not (wb.security and wb.security.workbookPassword):
    f("BLOCKER", "workbook", "structure not password-locked")
print(f"L4 protection checked; unlocked formulas: {len(unlocked_formula)}")

# secrets scan (the protection password must never be text in a cell)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            _pw = os.environ.get("PCI_XLSX_PASSWORD")
            if _pw and isinstance(c.value, str) and _pw in c.value:
                f("BLOCKER", f"{ws.title}!{c.coordinate}", "protection password appears as text")

# -------------------------------------------------- L5 Article Bank quality
ab = wb["Article Bank"]
titles, ids, defects = [], [], Counter()
r = 4
while ab.cell(r, 1).value:
    tid = ab.cell(r, 1).value
    t = str(ab.cell(r, 2).value or "")
    pk = str(ab.cell(r, 10).value or "")
    pr = str(ab.cell(r, 13).value or "")
    ids.append(tid)
    titles.append(t)
    if "  " in t: defects["double space in title"] += 1
    if "{" in t or "}" in t: defects["template brace left in title"] += 1
    if re.search(r"\b(\w+) \1\b", t.lower()): defects["repeated word in title"] += 1
    if not pk or not pr: defects["missing keyword/prompt"] += 1
    if t and t[:1].islower() and not t.startswith(("nPlan", "iOS", "eLearning")):
        defects["title starts lowercase"] += 1
    if pr and t and t not in pr: defects["prompt does not embed its title"] += 1
    r += 1
norm = defaultdict(list)
for tid, t in zip(ids, titles):
    k = re.sub(r"[^a-z0-9]+", " ", t.lower()).strip()
    norm[k].append(tid)
dupes = {k: v for k, v in norm.items() if len(v) > 1}
print(f"L5 article bank: {len(ids)} rows, unique ids {len(set(ids))}, exact-dupe titles {len(dupes)}")
for k, v in list(defects.items()):
    sev = "MAJOR" if v > 50 else "MINOR"
    f(sev, "Article Bank", f"{v} rows: {k}")
if dupes:
    f("MAJOR", "Article Bank", f"{len(dupes)} duplicate normalised titles, e.g. {list(dupes)[:3]}")

# ------------------------------------------------- L7 print/export hygiene
for ws in wb.worksheets:
    if not ws.page_margins or abs((ws.page_margins.left or 0) - 0.4) > 0.01:
        f("MINOR", ws.title, "non-standard left margin")
    if not str(ws.oddFooter.left.text or ""):
        f("MINOR", ws.title, "no footer")
    if ws.print_area:
        pa = str(ws.print_area)
        m = re.search(r"\$?([A-Z]{1,2})\$?(\d+):\$?([A-Z]{1,2})\$?(\d+)", pa)
        if m:
            last_col = openpyxl.utils.column_index_from_string(m.group(3))
            used = ws.max_column
            if used > last_col and ws.title not in ("Article Bank", "Weekly Review",
                                                    "Content Scheduler", "Dashboard"):
                f("MINOR", ws.title, f"print area ends col {m.group(3)} but content to "
                  f"{openpyxl.utils.get_column_letter(used)}")
print("L7 print/export checked")

# ------------------------------------------------ L8 navigation / cross-refs
mp = wb["MAP"]
linked = set()
for r in range(4, 80):
    c = mp.cell(r, 1)
    if c.hyperlink is not None:
        loc = str(c.hyperlink.location or "")
        tgt = loc.split("!")[0].strip("'")
        linked.add(tgt)
        if tgt not in names:
            f("BLOCKER", f"MAP!A{r}", f"link to missing sheet {tgt!r}")
missing_from_map = [n for n in wb.sheetnames if n not in linked and n not in ("MAP",)]
if missing_from_map:
    f("MINOR", "MAP", f"sheets not listed: {missing_from_map}")
print(f"L8 MAP links: {len(linked)}; unlisted sheets: {len(missing_from_map)}")

# pointer sanity in guidance text
PTR = [("Platform Setup", 16, "P"), ("Platform Setup", 17, "Q")]
ps = wb["Platform Setup"]
if ps.cell(3, 16).value and "rank" not in str(ps.cell(3, 16).value).lower():
    f("MINOR", "Platform Setup!P3", "column P is referenced as value rank in guides")

# ------------------------------------------------------------- summary
print("\n================ FINDINGS ================")
if not F:
    print("NONE")
for sev in ("BLOCKER", "MAJOR", "MINOR", "COSMETIC"):
    for s, w, m in F:
        if s == sev:
            print(f"[{s}] {w}: {m}")
print(f"\nTOTAL: {len(F)}  (blockers {sum(1 for x in F if x[0]=='BLOCKER')})")
