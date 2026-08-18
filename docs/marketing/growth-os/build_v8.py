#!/usr/bin/env python3
"""V8 layer on top of the V7 build: team-first settings, Accounts Register,
TEAM GUIDE and GROWTH PLAYBOOK sheets, and the premium formatting pass that
matches the workbook's own design system (Arial, navy 1F3864, band 2E5C9A).

Run build_v7.py first; this loads its output and saves PCI_AI_Growth_OS_V8.xlsx.
"""
import copy as copymod
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "PCI_AI_Growth_OS_V7.xlsx"
OUT = "PCI_AI_Growth_OS_V9.xlsx"

NAVY = "1F3864"; BAND = "2E5C9A"; ACCENT = "C00000"
YELLOW = PatternFill("solid", fgColor="FFF2CC")
NAVYFILL = PatternFill("solid", fgColor=NAVY)
BANDFILL = PatternFill("solid", fgColor=BAND)
ZEBRA = PatternFill("solid", fgColor="F2F6FB")
GREYNOTE = Font(name="Arial", size=8, italic=True, color="666666")
TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
SUB = Font(name="Arial", size=9, italic=True, color="444444")
H2 = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=9)
BODYB = Font(name="Arial", size=9, bold=True)
THIN = Side(style="thin", color="D0D7E5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.load_workbook(SRC)


def title_row(ws, ncols, text, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1); c.value = text; c.font = TITLE; c.fill = NAVYFILL
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(2, 1); s.value = subtitle; s.font = SUB; s.alignment = WRAP
        ws.row_dimensions[2].height = 24


def header_band(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i); c.value = h; c.font = H2; c.fill = BANDFILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BOX
    ws.row_dimensions[row].height = 22
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def section(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1); c.value = text; c.font = H2; c.fill = BANDFILL
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


# ------------------------------------------------ 1. Team-first START HERE
sh = wb["START HERE"]
sh["A17"] = "2.  TEAM SETTINGS  —  the manager fills these once (yellow cells)"
sh["A18"] = "Team / department name"
sh["B18"].fill = YELLOW
sh["C18"] = ""
sh["A19"] = "Manager / reviewer"
sh.column_dimensions["A"].width = 44
sh["A23"] = ("3.  DAILY TARGETS  —  per person, per working day. The Dashboard multiplies "
             "by the number of people on the roster.")
sh["A24"] = "Activity"
b24 = sh["B24"]; b24.value = "Target per day (each person)"
# restyle the V7-added sections (93+) into the house style
for r in (93, 103):
    c = sh.cell(r, 1); c.font = Font(name="Arial", size=11, bold=True, color=NAVY)
for r in list(range(94, 102)) + list(range(104, 114)):
    sh.cell(r, 1).font = BODY
    sh.cell(r, 1).alignment = WRAP
sh.cell(19, 3).font = GREYNOTE

# §7 becomes the brand/property model: every log row now carries a
# For (brand) column, so one file serves the whole estate.
sh["A68"] = "7.  BRANDS & PROPERTIES  —  every logged row says WHO it is for"
sh["A69"] = "Primary brand (default focus)"
sh["B69"] = "PCI AI - Institute (umbrella)"
sh["A70"] = ("Each log tab has a For (brand) column: the institute, one certification "
             "(PCL-AI / PFL-AI / PML-AI), PCI World or Certuvo. Tag every row — the "
             "Objective Performance sheet splits results by brand as well as by "
             "campaign, so one file serves the whole estate without mixing the numbers.")
sh["A70"].font = GREYNOTE; sh["A70"].alignment = WRAP

# ---- Web domains: where every link, UTM and press release must land
from canonical_lov import DOMAINS as _DOMAINS
_dr = 115
sec = sh.cell(_dr, 1)
sec.value = "11.  OUR WEB DOMAINS  —  use these in links, UTMs, bios and releases"
sec.font = Font(name="Arial", size=11, bold=True, color=NAVY)
_dr += 1
for _c, _h in ((1, "Domain"), (2, "Brand / property it serves"), (3, "What it is for (manager edits if wrong)")):
    hc = sh.cell(_dr, _c); hc.value = _h; hc.font = H2; hc.fill = BANDFILL
    hc.alignment = Alignment(wrap_text=True, vertical="center"); hc.border = BOX
sh.row_dimensions[_dr].height = 18
_dr += 1
for dom, brand, use in _DOMAINS:
    sh.cell(_dr, 1).value = dom
    sh.cell(_dr, 1).font = Font(name="Arial", size=9, bold=True, color=NAVY)
    sh.cell(_dr, 2).value = brand
    sh.cell(_dr, 2).font = BODY
    sh.cell(_dr, 3).value = use
    sh.cell(_dr, 3).font = BODY
    for _c in (1, 2, 3):
        sh.cell(_dr, _c).border = BOX
        sh.cell(_dr, _c).alignment = WRAP
    for _c in (2, 3):
        sh.cell(_dr, _c).fill = YELLOW
        sh.cell(_dr, _c).protection = Protection(locked=False)
    sh.row_dimensions[_dr].height = 22
    _dr += 1
note = sh.cell(_dr, 1)
note.value = ("Every UTM Builder link and every press release must land on one of these "
              "domains. A link anywhere else is off-estate — flag it in the Monday review.")
note.font = GREYNOTE
sh.merge_cells(start_row=_dr, start_column=1, end_row=_dr, end_column=3)
sh.column_dimensions["C"].width = 46

# ------------------------------------------------ 2. Accounts Register sheet
ps = wb["Platform Setup"]
NP = 0
while ps.cell(4 + NP, 2).value:
    NP += 1
ar = wb.create_sheet("Accounts Register")
title_row(ar, 10, "ACCOUNTS REGISTER  —  every marketing account, one page",
          "This page builds itself from Platform Setup — edit there, never here. The only "
          "cells you type on this page are the yellow ones.")
ar["A3"] = ("SECURITY RULE — WHY THERE IS NO PASSWORD COLUMN.  Passwords never go in this "
            "file. A shared workbook is copied, synced and version-historied — one paste "
            "here would put every account credential in front of the whole team, forever. "
            "Passwords live in the team password manager (e.g. Bitwarden / 1Password shared "
            "collection); this register tells you exactly which vault entry to open. The "
            "manager grants vault access on day 1 (Day-1 Checklist, START HERE).")
ar.merge_cells("A3:J3")
ar["A3"].font = Font(name="Arial", size=9, bold=True, color=ACCENT)
ar["A3"].alignment = WRAP
ar.row_dimensions[3].height = 46
ar["A4"] = "Team password vault (name / where it lives):"
ar["A4"].font = BODYB
ar.merge_cells("A4:C4")
ar.merge_cells("D4:G4")
ar["D4"].fill = YELLOW
header_band(ar, 6, ["#", "Platform", "Area", "Priority", "Profile / page URL",
                    "Login email or username", "Vault entry name", "2FA on?",
                    "Owner", "Setup status"],
            [5, 40, 14, 10, 34, 26, 24, 8, 16, 13])
for i in range(NP):
    r, s = 7 + i, 4 + i
    def pt(col):
        ref = f"'Platform Setup'!${col}${s}"
        return f'=IF({ref}="","",{ref})'
    vals = [pt("A"), pt("B"), pt("C"), pt("D"), pt("F"),
            pt("G"), pt("H"), pt("I"), pt("M"), pt("J")]
    for c, v in enumerate(vals, 1):
        cell = ar.cell(r, c); cell.value = v; cell.font = BODY; cell.border = BOX
        if i % 2:
            cell.fill = ZEBRA
ar.freeze_panes = "A7"
ar.sheet_view.showGridLines = False
ar.sheet_properties.tabColor = "FF1F3864"
for row in ar["D4:G4"]:
    for cell in row:
        cell.protection = Protection(locked=False)
ar.protection.sheet = True

# ------------------------------------------------ 3. TEAM GUIDE sheet
tg = wb.create_sheet("TEAM GUIDE", 1)
title_row(tg, 5, "TEAM GUIDE  —  how this system works, end to end",
          "For every team member, new or old. Fifteen minutes here saves the manager "
          "explaining the same thing ten times. The deeper references are How-To Guides, "
          "LinkedIn Playbook and GROWTH PLAYBOOK.")
tg.sheet_view.showGridLines = False
tg.sheet_properties.tabColor = "FF548235"
tg.column_dimensions["A"].width = 30
for col, w in (("B", 44), ("C", 44), ("D", 26), ("E", 24)):
    tg.column_dimensions[col].width = w

r = 4
section(tg, r, 5, "THE WHOLE SYSTEM IN ONE SENTENCE"); r += 1
tg.cell(r, 1).value = ("You do real marketing work → you log it the same day with evidence → "
                       "the Dashboard, scorecards and funnel build themselves → the manager "
                       "reviews Monday morning → next week gets better. Nothing here is "
                       "busywork: every cell you fill feeds a number a decision is made on.")
tg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
tg.cell(r, 1).font = BODY; tg.cell(r, 1).alignment = WRAP
tg.row_dimensions[r].height = 42
r += 2

section(tg, r, 5, "WHERE DO I LOG WHAT  —  the only mapping you must memorise"); r += 1
header_band(tg, r, ["You did this…", "Log it here", "Also here (if it applies)",
                    "It shows up in…", "Same-day rule"])
r += 1
MAP = [
 ("Researched leads in Sales Navigator", "DAILY ENTRY (count + minutes)", "LinkedIn Outreach — one row per lead", "Team Scorecard, Dashboard §4", "Log before you finish for the day"),
 ("Sent a connection request / message / follow-up", "LinkedIn Outreach — tick the columns on that lead's row", "DAILY ENTRY (count for the day)", "Funnel, Employee Score", "Paste the exact message text"),
 ("Got a reply / booked a meeting", "LinkedIn Outreach — Outcome column", "—", "Funnel stages 5–6, Dashboard §3", "Update the same day; add Meeting date"),
 ("Published a post / article / video", "Content Calendar — one row per item", "DAILY ENTRY (count)", "Dashboard §5, Publishing Plan", "Paste the published URL"),
 ("Answered on Quora / Reddit / a community", "Community & PR — one row per answer", "DAILY ENTRY (count)", "Dashboard §5", "Paste the thread URL"),
 ("Contacted a partner / university / podcast", "Partnership Pipeline — one row per organisation", "Community & PR for individual touches", "Partnership Pipeline stage, Dashboard §8", "Score ICP fit honestly"),
 ("Ran an A/B test", "Experiments — one row per test", "—", "Playbook improvements", "Conclude before starting the next"),
 ("Built a tracked link", "UTM Builder", "—", "Google Analytics campaigns", "Copy only the generated link"),
 ("Set up / improved an account", "Platform Setup — status + profile %", "DAILY ENTRY ('Account created / profile completed')", "Accounts Register, Dashboard §2", "Record the vault entry name"),
 ("Reviewed analytics", "DAILY ENTRY (Analytics review)", "SEO Clusters for search data", "Platform Progress", "One row per session, primary tool"),
 ("Sent an email campaign / WhatsApp / Telegram / SMS broadcast", "DAILY ENTRY — 'Email campaign sent' or 'WhatsApp / Telegram / SMS sent', How many = messages delivered", "—", "Weekly Pulse (email / WhatsApp-SMS rows)", "Log the same day with the campaign report link"),
 ("Posted a job opening", "Job Postings — one row per platform per position", "DAILY ENTRY ('Job post published')", "Weekly Pulse", "Close filled roles the same day"),
 ("Chased or earned a backlink / reclaimed a mention", "Link Building — one row per link prospect, Status moved as it progresses", "DAILY ENTRY ('Link building / outreach (off-page)')", "Weekly Pulse ('Backlinks gone live')", "Set Date live + Dofollow/Nofollow when it lands; never buy a link"),
 ("Need something to write?", "Article Bank — 5,000+ briefs with keywords and an AI prompt per row; filter by Cluster/Difficulty, claim with Owner + Status", "Content Calendar (one row when you publish)", "Keyword Plan, SEO Clusters", "Easy difficulty first; gold flagship rows before anything else"),
 ("Set up or changed a posting schedule", "Content Scheduler — one row per platform × cadence × window", "Content Calendar (each published post still gets its row)", "Dashboard (active schedules + coverage)", "Coverage under 100% = the schedule is slipping — fix it before Friday"),
 ("EVERY row, whatever it is", "Set BOTH tag columns: Objective (which campaign — honorary outreach, a certification's sales, authority building…) and For (brand) (who it is for — the institute, PCL-AI / PFL-AI / PML-AI, PCI World or Certuvo)", "—", "Objective Performance — results per campaign AND per brand", "An untagged row lands in a red '(no objective set)' / '(no brand set)' line the manager reviews Monday"),
]
for row_ in MAP:
    for c, v in enumerate(row_, 1):
        cell = tg.cell(r, c); cell.value = v; cell.font = BODY
        cell.alignment = WRAP; cell.border = BOX
        if (r % 2) == 0: cell.fill = ZEBRA
    # height fits the longest wrapped cell (cols A~28, B/C~48, D~28, E~26 chars/line)
    _need = max(len(str(row_[0])) / 26, len(str(row_[1])) / 44, len(str(row_[2])) / 44,
                len(str(row_[3])) / 24, len(str(row_[4])) / 22)
    tg.row_dimensions[r].height = max(28, int(_need + 1) * 11 + 6)
    r += 1
r += 1

section(tg, r, 5, "YOUR FIRST DAY  (the manager ticks these on START HERE §10)"); r += 1
FIRST = [
 "1)  Read START HERE top to bottom — especially the Golden Rules and the sharing rules (§9).",
 "2)  Check your name is on the roster exactly as you will select it in dropdowns.",
 "3)  Get password-vault access from the manager. Open Accounts Register to see which vault entry each platform uses.",
 "4)  Optimise your own LinkedIn profile using LinkedIn Playbook step A. Do not send anything yet.",
 "5)  Read the Message Bank. You may only ever send rows marked Approved — personalised, never pasted blind.",
 "6)  Log your first 5 researched leads in LinkedIn Outreach and show the manager BEFORE sending any request.",
 "7)  After the manager reviews your first 3 sent messages, you are live. Follow the daily targets from START HERE §3.",
]
for t in FIRST:
    tg.cell(r, 1).value = t; tg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    tg.cell(r, 1).font = BODY; tg.cell(r, 1).alignment = WRAP
    r += 1
r += 1

section(tg, r, 5, "YOUR EVERY DAY  —  the 20-minute logging habit"); r += 1
DAY = [
 "Morning (5 min): open Dashboard →  'Follow-ups due today' and your red cells. Clear those first — a due follow-up beats a new lead.",
 "During the day: work in your platforms. Keep LinkedIn Outreach open; a lead takes 30 seconds to log while the profile is in front of you.",
 "Before you stop (15 min): fill DAILY ENTRY — one row per activity type, with counts, minutes and one evidence link. If it is not logged today, it did not happen.",
 "Never: sort a shared sheet, paste whole rows, type over grey or locked cells, send an unapproved message, or contact anyone marked Declined.",
]
for t in DAY:
    tg.cell(r, 1).value = "•  " + t
    tg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    tg.cell(r, 1).font = BODY; tg.cell(r, 1).alignment = WRAP
    tg.row_dimensions[r].height = 26
    r += 1
r += 1

section(tg, r, 5, "HOW YOUR SCORE WORKS  (Employee Score tab — quality beats volume)"); r += 1
SCORE = [
 ("Lead quality (15)", "Are the leads you research genuinely in-profile (8+ years, right roles)?"),
 ("Acceptance rate (20)", "Do your connection requests get accepted? Benchmark: 30%+ is good. Personalisation drives this."),
 ("Reply rate (20)", "Do your messages get replies? Fluent copying of the Message Bank with real personalisation drives this."),
 ("Positive replies (10)", "Interested / info requested / meeting booked."),
 ("Meetings & applications (20)", "The outcomes that matter."),
 ("Downstream results (10)", "Conversions attributed to your leads."),
 ("Compliance (5)", "No over-length messages, no rule breaches. Breaches zero this and flag the manager."),
]
for name, desc in SCORE:
    tg.cell(r, 1).value = name; tg.cell(r, 1).font = BODYB; tg.cell(r, 1).border = BOX
    tg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    tg.cell(r, 2).value = desc; tg.cell(r, 2).font = BODY; tg.cell(r, 2).border = BOX
    tg.cell(r, 2).alignment = WRAP
    r += 1
tg.cell(r, 1).value = ("No grade appears until you have ~20 connections of history — a blank "
                       "grade means 'not enough data yet', not 'failing'.")
tg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
tg.cell(r, 1).font = GREYNOTE
r += 2

section(tg, r, 5, "FREQUENTLY ASKED"); r += 1
FAQ = [
 ("A cell went red when I typed a name/platform.", "Red = that exact text is not on the official list, so it would be invisible to every report. Use the dropdown, or fix the spelling."),
 ("I can't type in a cell.", "It is protected on purpose — it is a formula or a manager-owned cell. Nothing is broken. Yellow cells are yours."),
 ("The lead I found shows DUPLICATE.", "Someone already logged them. Ctrl+F the URL, read the earlier row, and do not re-contact — especially if they declined."),
 ("Which platform value do I pick for X?", "Lists tab, column 'Platform logging rule' — every ambiguous case has a written rule (Shorts → YouTube, newsletter issue → LinkedIn Newsletter, etc.)."),
 ("I ran out of empty rows.", "Tell the manager — they archive a dated copy monthly and clear logged rows. Never delete rows yourself."),
 ("I made a typo in a logged row.", "Edit the cell to correct it. Never delete rows; disputes are resolved from version history."),
 ("Where do I see how the team is doing?", "Dashboard (whole team), Team Scorecard (per person), Weekly Review (trend). You never type on those pages."),
 ("Someone asked me to promise them a certification.", "Never. Read the Golden Rules — no outcome promises, ever. Report the conversation to the manager."),
]
for q, a in FAQ:
    tg.cell(r, 1).value = q; tg.cell(r, 1).font = BODYB
    tg.cell(r, 1).alignment = WRAP; tg.cell(r, 1).border = BOX
    tg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    tg.cell(r, 2).value = a; tg.cell(r, 2).font = BODY
    tg.cell(r, 2).alignment = WRAP; tg.cell(r, 2).border = BOX
    tg.row_dimensions[r].height = 28
    r += 1
tg.protection.sheet = True

# ------------------------------------------------ 4. GROWTH PLAYBOOK sheet
gp = wb.create_sheet("GROWTH PLAYBOOK", 2)
title_row(gp, 6, "GROWTH PLAYBOOK  —  the techniques, in working detail",
          "Each technique: why it works, the exact steps, cadence, the number that proves "
          "it is working, and the mistake that gets accounts banned or rankings burned. "
          "Every step logs into the sheet named in the last column.")
gp.sheet_view.showGridLines = False
gp.sheet_properties.tabColor = "FF548235"
widths = [26, 46, 52, 14, 22, 20]
for i, w in enumerate(widths, 1):
    gp.column_dimensions[get_column_letter(i)].width = w

PLAYS = [
 ("1. SEO content clusters (pillar + spokes)",
  "Google ranks depth, not scattered posts. One pillar page per money topic "
  "('project controls certification', 'earned value management') with 6-10 supporting "
  "articles all linking back to it makes PCI the topical authority.",
  "1) Pick the cluster in SEO Clusters — pillar phrase + 8 supporting questions people search.  "
  "2) Write the pillar (2,500+ words, answers the whole topic, cites standards).  "
  "3) One supporting article per week answering ONE question, linking UP to the pillar with exact-phrase anchor.  "
  "4) Interlink siblings.  5) Add FAQ blocks with the literal question as a heading.  "
  "6) Track impressions/clicks/position monthly from Search Console into SEO Clusters.",
  "1 pillar / month; 1-2 spokes / week",
  "Search Console: position for the pillar phrase; clicks/month",
  "SEO Clusters"),
 ("2. E-E-A-T authority signals",
  "Google's quality systems reward demonstrated Experience, Expertise, Authority, Trust — "
  "critical for an education brand ranked among established bodies.",
  "1) Every article carries a named author with credentials and a bio page.  "
  "2) Author bios link to their LinkedIn and publications.  "
  "3) Cite primary sources (standards, published data) — never invent numbers.  "
  "4) Keep an About/Governance page path from every article.  "
  "5) Get authors quoted elsewhere (see technique 4) and link those mentions back to the bio.",
  "Continuous — a rule, not a task",
  "Branded-search volume; author-name searches",
  "Content Calendar"),
 ("3. Answer-engine optimisation (AEO) — being the cited answer",
  "AI answers (Google AI Overviews, ChatGPT, Perplexity) are a fast-growing discovery "
  "surface, and the traffic they send arrives with the question already framed — treat "
  "published conversion multiples as vendor marketing until you measure your own in GA4. "
  "The pipelines differ: ChatGPT retrieves from Bing's index, Perplexity crawls itself, "
  "AI Overviews uses Google — each needs its own work. The pattern that repeats across "
  "cited pages is a 2-3 sentence direct answer in the opening.",
  "1) robots.txt: explicitly allow GPTBot, OAI-SearchBot, PerplexityBot, ClaudeBot, "
  "Google-Extended (check the CDN/WAF is not blocking them).  "
  "2) Verify the domain in Bing Webmaster Tools + enable IndexNow — that is the ChatGPT pipeline.  "
  "3) Restructure the top 30 pages: question-form H2s with a 2-3 sentence answer capsule "
  "directly under each, then depth; add FAQ blocks and comparison tables.  "
  "4) Put one cited statistic in every major section — and publish original data so PCI "
  "becomes the source others cite.  "
  "5) Off-page is most of the battle: get PCI named in the pages the engines already cite "
  "(certification listicles, Reddit/Quora threads, trade-press roundups) — pitch inclusion "
  "into existing ranking articles before writing new ones.  "
  "6) Keep entity facts identical across the site, Wikidata, Crunchbase, LinkedIn, GBP.  "
  "7) Monthly: run the 20-prompt audit; log which domains get cited; target them.",
  "5 pages restructured / month + monthly audit",
  "Brand mention rate across 20 prompts; AI-referral sessions in GA4",
  "SEO Clusters / Experiments"),
 ("4. Digital PR & backlink earning (post-HARO stack)",
  "Links and mentions from real publications remain the strongest ranking and AI-citation "
  "input. HARO is dead (shut Dec 2024); the working 2026 stack is journalist-request "
  "services + contributed articles + podcast guesting + one quarterly data study.",
  "1) Build a one-page expert kit per spokesperson (bio, credentials, 5 angles, headshot).  "
  "2) Subscribe: Source of Sources + Help a B2B Writer (free), Qwoted / Featured, "
  "ResponseSource for UK press; monitor #journorequest on X and Bluesky twice daily.  "
  "3) Answer only on-expertise queries within 2 hours: 80-120 words, data-led, quote-ready.  "
  "4) Publish one original data study per quarter (salary survey, AI-adoption survey) — "
  "the single best link magnet — and pitch it to the trade press on the PR & Target "
  "Directory tab.  5) One contributed article and one podcast pitch per month from the "
  "same tab.  6) Every placement goes on an 'As featured in' page (E-E-A-T + AI-citation asset).  "
  "7) Never pay for links; never mass-pitch off-topic — accounts get deprioritised.",
  "3-5 responses / week; monthly article + podcast; quarterly study",
  "Linked placements / month; referring domains; AI-answer brand mentions",
  "Community & PR"),
 ("5. LinkedIn organic engine (2026 mechanics)",
  "The decision-makers are here, but the algorithm has changed: personal profiles reach "
  "several times what company pages do on identical content (published estimates cluster "
  "around 5-10x; measure your own), document/carousel posts lead "
  "engagement, dwell time (15+ seconds of reading) is the top ranking signal, and "
  "external links in the post body still suppress reach.",
  "1) Put 70% of effort behind 2-3 named staff profiles; the company page is the "
  "archive/credibility layer that reshares people.  "
  "2) Weekly per profile: one document/carousel (exam tips, salary data), one short "
  "native video (under 90s), one text story — written for dwell: strong first two lines, "
  "line-broken body, a chart screenshot.  "
  "3) Comment meaningfully on 15 target-audience posts daily BEFORE posting.  "
  "4) Links go in the first comment or profile — never the post body.  "
  "5) Reply to every comment within 60 minutes of posting (first-hour velocity gates reach).  "
  "6) Reshares from colleagues staggered, not simultaneous — identical simultaneous "
  "reposts suppress all copies.  "
  "7) Outreach only through the approved message system; never automate — engagement pods "
  "and AI-comment tools are detected and reach-penalised.",
  "Daily comments; 3 posts / week / profile; fortnightly newsletter",
  "Member reach per post; acceptance %; reply %; newsletter subscribers",
  "Content Calendar + LinkedIn Outreach"),
 ("6. Community-led growth (answer first, never spam)",
  "Quora, Reddit, Planning Planet, PMI Community, Stack Exchange rank in Google for years "
  "and reach practitioners directly. One good answer works forever; one spammy answer "
  "burns the account and the brand.",
  "1) Pick 2 communities per person, build real profiles (disclose PCI affiliation).  "
  "2) Answer the question COMPLETELY first; PCI gets one contextual mention only where "
  "genuinely relevant — max 1 link per 5 answers on Reddit, none on Stack Exchange.  "
  "3) Target questions that already rank in Google (search the question first).  "
  "4) 3-5 quality answers/week per person beats 20 thin ones.  "
  "5) Log every answer in Community & PR with the thread URL.  "
  "6) Read each community's self-promotion rules before the first post — bans are permanent.",
  "3-5 answers / person / week",
  "Answer views, upvotes, referral sessions in GA4",
  "Community & PR"),
 ("7. Syndication waterfall (one asset, many surfaces — canonicals verified)",
  "Every original article can reach several extra audiences without duplicate-content "
  "damage — IF each platform's canonical reality is respected. Verified 2026: Medium, "
  "DEV and Hashnode support canonicals; Substack and LinkedIn Articles do NOT.",
  "1) Publish the original on the website first; allow 2-10 days and verify it is indexed before syndicating.  "
  "2) Medium: use the Import tool — it sets the canonical to the original automatically.  "
  "3) LinkedIn Articles: 300-word excerpt + 'read the full analysis' link — never the full "
  "piece (no canonical support).  "
  "4) Substack: native newsletter framing + link only — never paste the full post (no "
  "canonical; the copy can outrank your site).  "
  "5) Slides to SlideShare; a 60-second summary to Shorts; an audio take to the podcast.  "
  "6) Each copy = its own Content Calendar row, status Repurposed, linking the original; "
  "plan the waterfall in the 'Repurpose to' column at writing time.  "
  "7) If a syndicated copy outranks the original, strengthen internal links to the "
  "original or request removal.",
  "Top 2 posts / month",
  "Referral sessions per surface; original URL keeps the ranking",
  "Content Calendar / Publishing Plan"),
 ("8. Webinar → content engine",
  "One monthly webinar produces the month's best content and the warmest leads: "
  "registrants are self-qualified by topic.",
  "1) Monthly topic from the best-performing cluster.  2) Deliver on Zoom Webinars; "
  "list on Eventbrite/Luma/Meetup + LinkedIn Event.  3) UTM every registration link.  "
  "4) Replay to YouTube; cut 3-5 Shorts; write the recap article (technique 7 waterfall).  "
  "5) Registrants enter the ESP nurture sequence; attendees get the certification next-step "
  "email within 24h.  6) Every speaker invitation logged in Partnership Pipeline.",
  "Monthly",
  "Registrants, attendance %, replay views, applications started",
  "Partnership Pipeline + Content Calendar"),
 ("9. Email nurture (the owned list)",
  "The only audience no algorithm can take away, and the highest-converting surface for a "
  "considered purchase like certification.",
  "1) One lead magnet per certification (syllabus PDF, formula sheet, salary guide) behind "
  "a form.  2) Welcome sequence: 5 emails over 2 weeks — value, value, story, proof, ask.  "
  "3) Fortnightly newsletter after that: one insight, one resource, one next step.  "
  "4) Every send is one Email Marketing (ESP) row in DAILY ENTRY.  "
  "5) Consent only — record lawful basis at capture; unsubscribe honoured same day.",
  "Fortnightly sends; sequences always-on",
  "List growth/month, open %, click %, applications from email",
  "DAILY ENTRY + UTM Builder"),
 ("10. Digital badge loop (Credly)",
  "Every certified professional who shares a verifiable badge markets to their whole "
  "network with third-party credibility — near-zero cost, compounds with every cohort.",
  "1) Issue every credential as a Credly badge automatically on conferral.  "
  "2) Prompt shares to LinkedIn in the conferral email (one click).  "
  "3) Repost great badge-shares from the company page (with permission).  "
  "4) Track share-rate; a shared badge is a logged Engagement.",
  "Automatic per conferral; monthly share-rate review",
  "Badge share rate; referral traffic from credly.com",
  "Platform Setup + DAILY ENTRY"),
 ("11. Directories, citations & reviews",
  "Consistent entity citations (name-address-site) across credible directories underpin "
  "the knowledge panel and AI-engine trust; course directories add referral leads directly.",
  "1) Complete Google Business Profile + Bing Places with identical details.  "
  "2) List courses on findcourses/coursetakers-type aggregators (UK/US/Gulf).  "
  "3) CPD directories where factually eligible — never claim accreditation that does not exist.  "
  "4) Invite genuinely satisfied certified professionals to Trustpilot — never incentivised, "
  "never gated.  5) Keep Crunchbase/Wikidata facts current.",
  "One-time setup + quarterly verification",
  "Citation consistency; referral sessions from directories",
  "Platform Setup + QA & Compliance"),
 ("12. YouTube: Shorts discover, long-form converts (2026)",
  "The Dec 2025 algorithm shift cut long-form home-feed exposure sharply; Shorts are the "
  "discovery front door while long-form lives on search intent. Both feed Google's video "
  "results and AI-engine video citations.",
  "1) Long-form targets typed searches with title-match ('PCL-AI exam: what to expect', "
  "'Primavera P6 vs AI scheduling tools') — check autocomplete first.  "
  "2) Answer in the first 30 seconds, then depth; chapters + accurate transcripts.  "
  "3) Cut 3-5 Shorts from every long-form video: one exam question in 45s, salary stat "
  "of the week, a sharp take.  "
  "4) Cadence beats polish: one long-form / week + 3 Shorts / week beats one cinematic "
  "video a month.  "
  "5) End screens, pinned comments and description UTMs route to the matching cluster "
  "page or Certuvo.  6) Never buy subscribers; never delete old videos (age helps search).",
  "1 long-form / week + 3 Shorts / week",
  "Search-sourced views %; Shorts-to-subscriber rate; site sessions",
  "Content Calendar"),
 ("13. Measurement discipline",
  "Growth work without attribution becomes opinion. Everything above survives only if the "
  "numbers are trustworthy.",
  "1) Every external link in outreach, email, webinars and directories carries a UTM from "
  "the UTM Builder — no exceptions.  2) Weekly: Search Console queries → update SEO "
  "Clusters positions.  3) Weekly: GA4 acquisition by channel → note in Weekly Review.  "
  "4) Monthly: Channel Costs updated; cost-per-meeting reviewed on Dashboard §8.  "
  "5) Kill or fix any channel with zero results after 8 honest weeks — the Experiments tab "
  "decides, not opinion.",
  "Weekly reviews; monthly costs",
  "Cost per meeting; per-channel sessions and conversions",
  "UTM Builder + Weekly Review"),
 ("14. Entity authority — how AI engines rate an institution (researched)",
  "AI engines cite entities they can verify, not pages that merely rank: 77% of AI-cited "
  "URLs sit OUTSIDE the organic top 10; branded mentions correlate 3x more with AI "
  "visibility than backlinks (0.664 vs 0.218); matched profiles across LinkedIn / "
  "Crunchbase / org databases give roughly a 3x citation lift. The scoreboard is the "
  "Google Knowledge Panel.",
  "1) Entity graph first: entity home page + EducationalOrganization JSON-LD + sameAs to "
  "every profile + a complete Wikidata record (with ISNI / Crunchbase IDs).  "
  "2) Publish all three credentials to the Credential Engine Registry (free) and email "
  "CareerOneStop's Certification Finder — the machine-readable, government-consumed "
  "sources of 'this credential is real', both syndicated by API.  "
  "3) Deepen third-party reviews (Trustpilot / Coursecheck / G2) with volume and recency "
  "— review profiles roughly 3x ChatGPT presence; on-site star markup for yourself is "
  "banned.  4) Sustain Reddit and community answers — Reddit is the single most-cited "
  "domain across AI engines.  5) Earn branded editorial mentions over link-building.  "
  "6) Feed Bing (Webmaster Tools + IndexNow) — Bing's index is ChatGPT Search's "
  "retrieval layer.  7) Keep the site server-rendered, redirect-clean and open to "
  "CCBot / OAI-SearchBot / Claude-SearchBot / PerplexityBot; verify Common Crawl "
  "inclusion (CCBot runs no JavaScript).  8) Ship Course List + "
  "EducationalOccupationalCredential markup (the old Course Info rich result is retired "
  "— skip it).  9) One canonical name, address and description on every profile in this "
  "workbook — inconsistency fragments the entity.  10) Gulf: /en/ and /ar/ paths with "
  "ar-SA / ar-AE hreflang and genuinely written Arabic credential pages (~80% of Saudi "
  "commercial queries are Arabic-first).",
  "Entity graph once + quarterly audit; reviews and mentions continuous",
  "Knowledge Panel live and claimed; brand-mention rate in the 20-prompt AI audit",
  "Platform Setup + PR & Target Directory"),
  ("15. What NOT to do — dead and dangerous tactics (verified 2026)",
  "Half of 'SEO tips' lists still circulate tactics that now range from useless to "
  "penalty-inducing. This row exists so nobody on the team wastes a week on them.",
  "DEAD: the original HARO/Connectively (shut Dec 2024; brand since relaunched under Featured.com — use the technique-4 stack). Article "
  "directories and bulk PDF-submission lists (classed as link spam). Web 2.0 blog "
  "networks. llms.txt files (evidence-negative; Google confirmed no support).  "
  "DANGEROUS: paid links and link schemes; engagement pods and AI-comment tools on "
  "LinkedIn (detected, reach-penalised); identical answers pasted across Reddit threads "
  "(shadowban); undisclosed brand advocacy in communities (FTC fake-endorsement rule); "
  "incentivised or gated reviews; AI-generated thin content at scale (exactly what "
  "2025-26 core updates demote); automating LinkedIn outreach (account bans).",
  "Standing rule — review quarterly",
  "Zero penalties, zero banned accounts",
  "QA & Compliance"),
 ("16. Free publishing & news pipeline (researched)",
  "An institution earns authority by being findable in educational libraries, citable "
  "with DOIs, and corroborated by indexed news — all of which exists free. AI engines "
  "read indexed third-party references; a release's links are nofollow by Google "
  "policy, so the value is corroboration, never link equity.",
  "1) Every PCI framework/whitepaper gets a DOI on Zenodo under named authors (cite the DOI in articles).  "
  "2) One CC-licensed exam-prep primer per quarter to OER Commons + MERLOT.  "
  "3) One practical article per quarter to TrainingZone/HRZone (UK L&D angle), on top of the "
  "eLearning Industry / Training Industry cadence.  "
  "4) Guest posts from the PR & Target Directory 'Guest publications (PM)' rows — one per month across the team.  "
  "5) News: publish on the own-site newsroom first (NewsArticle schema), then PRLog (free, always) "
  "and openPR (free, 1 per 30 days); paid EIN Presswire only for launches that justify $149.  "
  "6) The Conversation only via university-affiliated authors (adjunct/honorary roles qualify) — "
  "route it through honorary fellows in academia.  "
  "7) Institutional endgame: the ISO/IEC 17024 accreditation row on PR & Target Directory is "
  "leadership's — marketing's job is making every existing proof visible.",
  "Monthly guest post; quarterly primer + DOI; releases as news happens",
  "Indexed mentions on non-PCI domains (search the brand monthly); DOI citations",
  "Content Calendar + Community & PR"),
 ("17. Original research & data studies — the link magnet",
  "Nothing earns links and press like data nobody else has. An annual 'State of "
  "Project Controls & AI' survey makes PCI the cited source instead of the citer.",
  "1) 10-question survey (salary, AI adoption, skills gaps) to the community, email list and LinkedIn.  "
  "2) Partners push it to reach 300+ responses (give them early access as the incentive).  "
  "3) Publish the report with named findings, charts and a Zenodo DOI.  "
  "4) Pitch individual statistics to trade press; cite the stats in every journalist-request answer.  "
  "5) Refresh annually — the '2027 edition' earns links all over again.",
  "Annual flagship + quarterly one-question pulses",
  "Backlinks + citations to the report (check monthly)",
  "Content Calendar + Community & PR"),
 ("18. Free tools & calculators",
  "Tool SERPs convert and rank: an EVM calculator, float calculator or exam-readiness "
  "quiz earns links, emails and daily return visits.",
  "1) Build one tool per quarter. Pick it from the Keyword Plan: filter Intent to "
  "Informational and read the 'Asset to build' column — anything that names a "
  "calculator, quiz or checker is a tool brief (start: the Certuvo question bank, a P1 keyword).  "
  "2) Each tool = its own landing page with schema, a results-share CTA and an email-capture on results.  "
  "3) Log the launch in Content Calendar; submit to relevant tool directories.  "
  "4) Internal-link every related article to its tool.",
  "One tool per quarter",
  "Tool sessions + emails captured per month",
  "Content Calendar"),
 ("19. Arabic & Gulf localisation",
  "The Gulf audience searches and scrolls in Arabic too. Localised pages and captions "
  "compound the KSA/UAE advantage no global competitor bothers to build.",
  "1) Translate the Gulf landing pages (Dubai / KSA / Qatar keyword rows) with a native speaker — never machine-only.  "
  "2) hreflang tags between English and Arabic versions.  "
  "3) Arabic captions on Instagram, TikTok and Snapchat posts during Gulf campaigns.  "
  "4) Arabic WhatsApp reminder templates (consented contacts).  "
  "5) Post in Gulf working hours (Sun-Thu, GST).",
  "Localise one page cluster per month during Gulf pushes",
  "Gulf organic sessions + Arabic-page conversions",
  "Content Calendar + SEO Clusters"),
 ("20. Internal linking architecture",
  "Links you own are free authority: pillars rank because spokes vote for them. "
  "Orphan pages rank for nothing.",
  "1) Every new spoke links UP to its pillar (exact-phrase anchor) and across to 2 siblings.  "
  "2) Money pages (certification landing pages) get a link from every top-traffic article.  "
  "3) Quarterly crawl (Search Console + a crawler) to find and fix orphan pages.  "
  "4) Breadcrumbs with BreadcrumbList schema site-wide.",
  "Rule on every publish + quarterly crawl",
  "Zero orphans; money-page positions rising",
  "SEO Clusters"),
 ("21. Employee advocacy & social-proof flywheel",
  "Company pages reach roughly 2-5% of followers; personal profiles reach several times "
  "that (see technique 5 — measure your own before quoting a multiple). Every "
  "certified candidate is a broadcast moment waiting to be asked.",
  "1) Every Company Page post is reshared by each team member with their OWN one-line take within 24h.  "
  "2) Every new certificate triggers the ask: Credly share to LinkedIn + a review (Trustpilot/Coursecheck) + a testimonial quote.  "
  "3) Testimonial bank grows monthly; rotate fresh quotes onto landing pages.  "
  "4) Leaders post 2x/week in their own voice — the institute's humans outrank its logo.",
  "Reshare within 24h, always; asks in the completion flow",
  "Employee-driven reach share; new testimonials per month",
  "Content Calendar + QA & Compliance"),
 ("22. Programmatic geo × role landing pages (with guardrail)",
  "The Keyword Plan's geo and role rows multiply: 'planning engineer certification' x "
  "'Dubai / KSA / UK / India' — but ONLY with genuinely local substance, or it is the "
  "exact thin-content spam the 2025-26 core updates demote.",
  "1) One template with real local content per page: salary data, market context, local employers, currency, testimonials from that market.  "
  "2) Build only pages backed by a Keyword Plan row — no keyword, no page.  "
  "3) Ship in batches of 10; watch indexation in Search Console for 30 days before the next batch.  "
  "4) Any page Google won't index or that reads templated: rewrite or noindex — never leave thin pages live.",
  "One 10-page batch per quarter, indexation-gated",
  "Geo-page impressions + conversions; 100% indexation rate",
  "SEO Clusters + Content Calendar"),
 ("23. Off-page SEO — the link engine (guardrailed)",
  "Referring domains are the strongest off-page ranking signal a new domain can "
  "earn, and for AI engines every credible link is one more corroborating source. "
  "This technique turns the linkable assets (research report, tools, guides) into "
  "a systematic weekly link programme instead of hoping.",
  "1) Quarterly competitor backlink gap: list who links to AACE, Project Control Academy, EVMi and "
  "projectcontrolsinstitute.com but not to us (free: their public mention SERPs + Ahrefs/Semrush free "
  "webmaster tiers on our own domain) — every gap is a prospect row in Link Building.  "
  "2) Monthly unlinked-mention reclaim: search the brand names in quotes; every mention without a link "
  "gets a polite ask — the easiest links that exist.  "
  "3) Resource-page + broken-link outreach: university careers pages, PM resource lists, association "
  "link pages — pitch the research report, calculators and guides as the replacement or addition.  "
  "4) Every guest post, journalist quote, podcast appearance and directory listing from the other "
  "techniques gets its own Link Building row — one system of record for off-page.  "
  "5) Anchors natural: brand or bare URL for most links; exact-match keywords sparingly.  "
  "6) NEVER: paid links, PBNs, bulk reciprocal swaps, comment/forum spam, fiverr link gigs — that is "
  "the link spam Google's SpamBrain demotes, and it burns the domain permanently.  "
  "7) Weekly rhythm: 10 new prospects logged, 5 outreach emails, follow-ups cleared; log the day's "
  "work in DAILY ENTRY as 'Link building / outreach (off-page)'.",
  "Weekly: 10 prospects, 5 outreach; monthly mention-reclaim; quarterly gap analysis",
  "Referring domains trend (Search Console links report + 'Earned - live' rows)",
  "Link Building"),
]
header_band(gp, 4, ["Technique", "Why it works", "Exactly how — numbered steps",
                    "Cadence", "KPI that proves it", "Logs into"])
r = 5
for i, row_ in enumerate(PLAYS):
    for c, v in enumerate(row_, 1):
        cell = gp.cell(r, c); cell.value = v
        cell.font = BODYB if c == 1 else BODY
        cell.alignment = WRAP; cell.border = BOX
        if i % 2: cell.fill = ZEBRA
    # the steps column (C) drives the height — nothing may clip in a manual
    _steps_lines = len(str(row_[2])) / 58 + 1
    gp.row_dimensions[r].height = max(118, min(320, int(_steps_lines) * 11 + 10))
    r += 1
gp.freeze_panes = "A5"
gp.protection.sheet = True

# ------------------------------------------ 4b. PR & TARGET DIRECTORY sheet
td = wb.create_sheet("PR & Target Directory", 3)
title_row(td, 7, "PR & TARGET DIRECTORY  —  named routes, verified August 2026",
          "The actual places, people and submission routes behind the generic platform "
          "rows — researched and verified live. Yellow columns are the team's tracker. "
          "Re-verify a route before relying on it if more than 6 months have passed.")
td.sheet_view.showGridLines = False
td.sheet_properties.tabColor = "FF548235"
header_band(td, 4, ["Category", "Target", "Route / where", "What to do",
                    "Owner", "Status", "Notes / result"],
            [22, 30, 34, 46, 14, 13, 26])
TARGETS = [
 ("UK course directories", "Reed Courses", "reed.co.uk/courses/providers", "Apply as provider; commercial terms on application"),
 ("UK course directories", "findcourses.co.uk", "via their sales team", "List PCL-AI / PFL-AI / PML-AI with UTM links"),
 ("UK course directories", "Emagister UK", "emagister.co.uk", "Lead-gen listing; compare cost per lead after 90 days"),
 ("UK course directories", "Coursecheck", "coursecheck.com/training-providers", "Provider signup + review capture at end of each cohort"),
 ("Gulf directories", "Laimoon", "providers.laimoon.com", "Free provider listing (UAE + Saudi)"),
 ("Gulf directories", "Coursetakers.ae", "coursetakers.ae", "Provider signup; UAE course aggregator"),
 ("Gulf directories", "Edarabia", "edarabia.com", "Submit institution listing + reviews"),
 ("Global directories", "Class Central", "classcentral.com", "Request provider indexing AND pitch a free intro certificate into their editorial lists"),
 ("CPD accreditation", "The CPD Certification Service", "cpduk.co.uk/become-accredited", "Largest/oldest UK accreditor — start here"),
 ("CPD accreditation", "CPD Standards Office", "cpdstandards.com/become-accredited", "Research-based accreditor; alternative quote"),
 ("CPD accreditation", "The CPD Group", "thecpd.group", "9,900+ providers; certified by The CPD Register"),
 ("Podcasts (guesting)", "Project Chatter", "projectchatter.com/be-a-guest", "TOP TARGET — the project controls podcast; direct guest form; 45-60 min remote"),
 ("Podcasts (guesting)", "Beyond Deadlines", "beyond-deadlines.com", "Construction planning/scheduling audience"),
 ("Podcasts (guesting)", "PM Happy Hour", "pmhappyhour.com", "Practitioner-guest format"),
 ("Podcasts (guesting)", "People and Projects", "peopleandprojectspodcast.com", "Long-running interview format"),
 ("Podcasts (guesting)", "Manage This (Velociteach)", "velociteach.com", "1st/3rd Tuesdays; pitch via contact form"),
 ("Publications", "PM World Journal", "pmworldjournal.com/authors", "PRIORITY — pitch a named-author monthly series to the editor"),
 ("Publications", "Project Times", "projecttimes.com/contribute", "Word-doc submission; evergreen career articles"),
 ("Publications", "Training Industry", "trainingindustry.com/collaborate-with-us + editor@trainingindustry.com", "Bylined article + research Top Training Companies criteria"),
 ("Publications", "eLearning Industry", "elearningindustry.com/post-here", "Author profile + L&D-angle articles (no HR content)"),
 ("Publications", "PBC Today", "pbctoday.co.uk", "700-word UK construction opinion pieces"),
 ("Publications", "The Digital Project Manager", "thedigitalprojectmanager.com", "Expert contribution + their podcast + certification roundups"),
 ("Publications (Gulf)", "Construction Week ME", "constructionweekonline.com", "Journalist roster on Muck Rack; offer expert commentary"),
 ("Publications (Gulf)", "CBNME", "cbnme.com", "Gulf construction business news; pitch editors"),
 ("Publications (academic)", "Journal of Project Management", "growingscience.com/jpm", "Open access, covers project controls, no submission charge"),
 ("Journalist requests", "Source of Sources (SOS)", "free subscription", "3 on-expertise answers weekly, 80-120 words, data-led"),
 ("Journalist requests", "Help a B2B Writer", "free, B2B-only", "Respond same day"),
 ("Journalist requests", "ResponseSource (UK)", "responsesource.com", "Paid; Business/Education/Construction categories; 2-hour answers"),
 ("Journalist requests", "#journorequest", "X and Bluesky, twice daily", "Fastest channel; search saved daily"),
 ("Events & awards", "Project Controls Expo UK", "projectcontrolexpo.com/uk", "London (Wembley) 2-4 Nov 2026, 2,000+ delegates — speaker application + Awards entry"),
 ("Events & awards", "Project Controls Expo USA", "projectcontrolexpo.com", "Washington DC 5-7 Oct 2026"),
 ("Events & awards", "Project Controls Expo Australia", "projectcontrolexpo.com", "Melbourne (MCG) 23-25 Nov 2026"),
 ("Events & awards", "Project Control Summit", "projectcontrolsummit.com", "Apply to present"),
 ("Events & awards", "Advancing Project Controls", "advancing-project-controls.com", "US conference; speaking/sponsor slots"),
 ("Speaker directories", "Sessionize", "sessionize.com — free", "Profiles for 2 PCI leaders + 3 CFP applications/quarter"),
 ("Speaker directories", "SpeakerHub", "speakerhub.com", "Directory + CFP board"),
 ("Podcast booking tools", "PodMatch", "podmatch.com", "Guest profile; 2 bookings/month"),
 ("Podcast booking tools", "MatchMaker.fm", "matchmaker.fm", "Freemium alternative"),
 ("Partnership targets", "APM (Assoc. for Project Management)", "apm.org.uk/membership", "Explore Corporate Partner programme — 42,000+ members"),
 ("Partnership targets", "ACostE", "acoste.org.uk", "Cost engineers association; natural collaboration partner"),
 ("Authority registries", "Credential Engine Registry", "credentialengine.org · accounts@credentialengine.org", "FREE — publish PCL-AI/PFL-AI/PML-AI in CTDL; auto-creates the credentialfinder.org listing; consumed by ~21 US state systems"),
 ("Authority registries", "CareerOneStop Certification Finder", "info@careeronestop.org", "US DoL-sponsored directory that syndicates via public API — email the certification details"),
 ("Authority registries", "D&B D-U-N-S", "dnb.com/en-us/smb/duns/get-a-duns.html", "Free, ~30 business days; one per legal entity"),
 ("Authority registries", "ISNI via Bowker", "myidentifiers.com", "~$5 one-time; add the ISNI to Wikidata + schema"),
 ("Authority registries", "OpenCorporates", "opencorporates.com/companies/us_de/", "Verify the Delaware record exists and matches the canonical name — 10 minutes, once"),
 ("Authority registries", "Google Knowledge Panel", "search the brand → 'Claim this knowledge panel'", "Claim once it appears; built by the entity-graph work, not by application"),
 ("Startup listings", "Tracxn", "tracxn.com/listyourstartup", "Free, analyst-reviewed; Perplexity observed citing it — list PCI and Certuvo"),
 ("Startup listings", "Dealroom", "dealroom.co/for-builders", "Free self-serve add + claim"),
 ("Startup listings", "Magnitt (Gulf)", "magnitt.com → Create/Claim company", "The MENA startup dataset governments cite"),
 ("Startup listings", "F6S", "f6s.com", "Free profile; watch grants/accelerators quarterly"),
 ("Startup listings", "AlternativeTo (Certuvo)", "alternativeto.net", "Free dofollow; alternatives pages get AI-cited"),
 ("Startup listings", "SaaSHub (Certuvo)", "saashub.com", "Free listing with dofollow"),
 ("Startup listings", "There's An AI For That (Certuvo)", "theresanaiforthat.com", "Only if genuine AI features ship"),
 ("Gulf job boards", "Bayt.com", "bayt.com employer registration", "Free employer profile; largest Arab-world job platform"),
 ("Gulf job boards", "Naukrigulf", "naukrigulf.com employer zone", "India-to-Gulf pipeline; light presence"),
 ("Gulf job boards", "GulfTalent", "gulftalent.com", "Only after Bayt shows demand"),
 ("Institution authority (strategic)", "ISO/IEC 17024 accreditation", "ANAB (anab.ansi.org, form PCAC-FR-503) / UKAS / IAS", "THE endgame authority move for a certification body — workshop → eligibility → application → assessment; realistic 12-24 months, five-figure annual cost. A leadership programme, not a marketing task; PCI already states 'developed with reference to 17024'"),
 ("Institution authority (strategic)", "Regulatory recognition — honest read", "Ofqual / Dubai KHDA / Abu Dhabi ACTVET / Saudi TVTC", "Ofqual needs an England-operating awarding organisation with governance + reserves; KHDA permit is for Dubai training institutes (AED 15-25k/yr + premises); TVTC licenses foreign providers in-Kingdom. Realistic route today: partner institutes that already hold the local licence deliver PCI prep — log them in Partnership Pipeline"),
 ("Institution authority (strategic)", "buildingSMART International", "buildingsmart.org member directory", "Paid membership with a public member-directory listing; consider when the BIM/data-standards angle goes live"),
 ("Institution authority (strategic)", "UNESCO Global Skills Academy / ILO skills partnerships", "unesco.org/en/global-education-coalition", "Partnership application, big-brand halo (IBM, Microsoft, Coursera are partners); long shot — one well-prepared approach, then park"),
 ("Institution authority (strategic)", "The Conversation", "theconversation.com/become-an-author", "Only university-affiliated authors (adjunct/honorary roles and PhD candidates qualify; independent orgs excluded). Route: PCI leaders holding visiting/adjunct roles, or honorary fellows in academia writing on project controls + AI"),
 ("Institution authority (strategic)", "UK construction awards (CN Awards, NFB Awards of Excellence, NCE)", "constructionnews.co.uk / builders.org.uk / nceawards.newcivilengineer.com", "Entries with published, indexed shortlist pages — authority mentions on trade-press domains; enter the training/skills categories"),
 ("Guest publications (PM)", "PM Column", "pmcolumn.com/contribution-guidelines", "Live guest-blogging guidelines; practical PM audience"),
 ("Guest publications (PM)", "Projectcubicle", "projectcubicle.com (guest post)", "~900-word guest posts with byline + link; running since 2007"),
 ("Guest publications (PM)", "IAPM 'Write for IAPM'", "iapm.net/en/service/write-for-iapm", "Structured guest-article process; European PM body audience"),
 ("Guest publications (PM)", "ProProfs PM articles", "proprofs.com/c/submit-project-management-article", "Simple submission; leadership/PM topics"),
 ("Guest publications (PM)", "Rebel's Guide to PM", "rebelsguidetopm.com — pitch directly", "Award-winning PM blog; accepts guest posts, no formal page — personalised pitch to Elizabeth Harrin"),
 ("Guest publications (PM)", "Vendor blogs: Deltek Project Nation + Plan Academy", "deltek.com Project Nation guest author; planacademy.com", "PM-software vendor blogs that publish expert contributions — exactly the project-controls buyer; pitch P6/EVM how-tos"),
 ("News distribution (verified)", "PRLog", "prlog.org — free", "Best free wire: unlimited cadence, images/video, fast Google indexing via RSS. Every credential/programme announcement goes here after the own-site newsroom post"),
 ("News distribution (verified)", "openPR", "openpr.com — free", "1 free release per 30 days; Google News indexed; strongest EU reach of the free wires"),
 ("News distribution (verified)", "IssueWire", "issuewire.com", "Free tier = first release only, NO Google News (that needs the $21-45 tiers); use paid only when a specific spread matters"),
 ("News distribution (verified)", "EIN Presswire", "einpresswire.com", "No free tier; $149/release transparent pricing — the paid option when a launch justifies it"),
 ("News distribution (verified)", "The rule for every release", "own newsroom first", "Google treats distributed press-release links as nofollow (link-scheme policy) — the value is indexed corroboration AI engines read, never link equity. Newsworthy facts, quotes, numbers; no keyword stuffing"),
 ("Job posting boards (verified)", "Careers page + Google for Jobs", "JobPosting JSON-LD per role page", "FREE and fully live 2026 — do this before paying any board; surfaces in Google's job box"),
 ("Job posting boards (verified)", "LinkedIn Jobs", "Company page → post free", "1 free job at a time (pauses day 14, expires day 30); promote only if stalled (~$300/30 days typical)"),
 ("Job posting boards (verified)", "Indeed", "post directly in the dashboard", "Free organic degraded by the 31 Mar 2026 policy (feed-posted jobs need Indeed Apply); direct posts keep residual visibility — sponsor small amounts only when needed"),
 ("Job posting boards (verified)", "Wellfound", "wellfound.com — free unlimited + ATS", "Free core forever; remote marketing/growth roles"),
 ("Job posting boards (verified)", "Bayt / Naukrigulf / Naukri.com", "employer accounts", "Bayt ~$158/post (the Gulf board); Naukrigulf 5 free posts for new employers; Naukri.com from ~Rs 400 — the India pipeline"),
 ("Job posting boards (verified)", "jobs.ac.uk / HigherEdJobs", "recruiter products", "GBP 250+ / $325 per post — only for UK/US academic SMEs, examiners and item writers"),
 ("Job posting boards (verified)", "Internshala", "employer dashboard — 1 free post/month", "India marketing interns and junior ambassadors; free tier only"),
 ("Verified SKIPs — do not spend time", "Golden.com", "—", "DEAD — acquired by ComplyAdvantage 2024; public wiki gone"),
 ("Verified SKIPs — do not spend time", "BBB accreditation", "—", "$500-1,000+/yr pay-to-play seal; reviews have replaced it"),
 ("Verified SKIPs — do not spend time", "Futurepedia Verified", "—", "$497 — pay only after free directories prove qualified traffic"),
 ("Verified SKIPs — do not spend time", "UKRLP / UKPRN", "—", "Requires a UK legal entity; becomes a KEEP the day a UK Ltd exists"),
 ("Verified SKIPs — do not spend time", "ROR", "—", "Eligible only once PCI appears as affiliation in published research"),
 ("Verified SKIPs — do not spend time", "CPD Register 'free listing'", "—", "Does not exist — listing is bundled with paid accreditation only"),
 ("Verified SKIPs — do not spend time", "Coursera / edX partnership", "—", "Selective BD partnerships; unrealistic at current scale — revisit later"),
 ("Verified SKIPs — do not spend time", "llms.txt", "—", "Re-verified evidence-negative: 97% of files never fetched; no provider parses it"),
 ("Verified SKIPs — do not spend time", "ESCO qualification listing", "—", "EU database fed only by Member-State national qualification registers referenced to EQF — no route for a non-EU private credential"),
 ("Verified SKIPs — do not spend time", "Wikibooks / Wikiversity", "—", "Same conflict-of-interest rules as Wikipedia — organisational self-contribution is treated as promotional; never self-edit"),
 ("Verified SKIPs — do not spend time", "BusinessBalls", "—", "Content library (Accipio-owned), no contributor route found"),
 ("Verified SKIPs — do not spend time", "engineering.com contributor route", "—", "No verified 2026 submission page — do not chase until the 6-monthly re-verify finds one"),
 ("Verified SKIPs — do not spend time", "PR.com / 24-7PressRelease / PRFree / PRUnderground", "—", "Unverified this round — stick to the verified PRLog + openPR stack; re-check at the 6-monthly review"),
 ("Verified SKIPs — do not spend time", "Xing (DACH)", "—", "Groups + events permanently shut Jan 2023 — now a pure DACH jobs board; no community surface for cert marketing"),
 ("Verified SKIPs — do not spend time", "VKontakte", "—", "EU-sanctioned Jul 2026 — compliance and reputational risk; do not touch"),
 ("Verified SKIPs — do not spend time", "WeChat / Weibo / Line / KakaoTalk", "—", "Functioning but wrong markets/language for an English certification body; Line OA can't even be registered by a non-JP/TW/TH entity"),
 ("Verified SKIPs — do not spend time", "Akhtaboot / Oliv / Dubizzle Jobs / apna", "—", "Regional job boards fully shadowed by Bayt (MENA) or too junior/blue-collar (apna, Oliv) for this audience"),
 ("Verified SKIPs — do not spend time", "We Work Remotely / RemoteOK", "—", "$299+/post with a tech-lean audience — wrong cost/audience for a small institute"),
]
r = 5
prev_cat = None
for cat, tgt, route, action in TARGETS:
    td.cell(r, 1).value = cat if cat != prev_cat else ""
    td.cell(r, 1).font = BODYB
    td.cell(r, 2).value = tgt; td.cell(r, 2).font = BODYB
    td.cell(r, 3).value = route; td.cell(r, 3).font = BODY
    td.cell(r, 4).value = action; td.cell(r, 4).font = BODY
    for c in range(1, 8):
        td.cell(r, c).border = BOX
        td.cell(r, c).alignment = WRAP
        if r % 2: td.cell(r, c).fill = ZEBRA
    for c in (5, 6, 7):
        td.cell(r, c).fill = YELLOW
        td.cell(r, c).protection = Protection(locked=False)
    _need = max(len(str(action)) / 42, len(str(route)) / 30, len(str(tgt)) / 27, 1.0)
    td.row_dimensions[r].height = max(24, int(_need + 1) * 10 + 4)
    prev_cat = cat
    r += 1
dv_own = DataValidation(type="list", formula1="'START HERE'!$B$55:$B$64", allow_blank=True)
dv_own.showErrorMessage = True
td.add_data_validation(dv_own); dv_own.add(f"E5:E{r-1}")
dv_st = DataValidation(type="list", formula1='"Not started,In progress,Submitted,Placed,Declined,Recurring"',
                       allow_blank=True)
dv_st.showErrorMessage = True
td.add_data_validation(dv_st); dv_st.add(f"F5:F{r-1}")
td.freeze_panes = "A5"
td.protection.sheet = True

# ------------------------------------ 4c. OBJECTIVE PERFORMANCE sheet
# The campaign dimension made testable: every log row carries an Objective
# (honorary outreach / certification sales / authority building / ...), and
# this sheet totals the results per category so each effort is judged on its
# own numbers — plus a person x objective minutes matrix for the manager.
from canonical_lov import OBJECTIVES as _OBJ_LIST

op = wb.create_sheet("Objective Performance")
op.sheet_view.showGridLines = False
title_row(op, 11, "OBJECTIVE PERFORMANCE  —  results by campaign and by brand",
          "Every log row carries two tags (dropdown columns in DAILY ENTRY, LinkedIn "
          "Outreach, Content Calendar, Community & PR and Partnership Pipeline): the "
          "Objective (which campaign) and For (brand) (who it is for — the institute, "
          "PCL-AI / PFL-AI / PML-AI, PCI World or Certuvo). The tables below split the "
          "same results both ways. The red '(no objective set)' / '(no brand set)' rows "
          "are untagged work — chase both to zero in the Monday review. The Value rank "
          "column says which effort is worth the most per hour (1 = top): compare it with "
          "Share of minutes — if high-value efforts get a low share, rebalance the week. "
          "Platforms carry the same ranking on Platform Setup column P.")
header_band(op, 3, ["Objective (effort category)", "Activities logged", "Minutes logged",
                    "Content published", "Community actions", "Leads logged",
                    "Meetings booked", "Revenue + signed deals (USD)", "Share of minutes",
                    "Minutes per meeting"],
            widths=[34, 11, 11, 11, 11, 11, 11, 14, 11, 12])

_DE_M = "'DAILY ENTRY'!$M$7:$M$1006"; _DE_G = "'DAILY ENTRY'!$G$7:$G$1006"
_DE_B = "'DAILY ENTRY'!$B$7:$B$1006"
_CC_F = "'Content Calendar'!$F$5:$F$403"; _CC_J = "'Content Calendar'!$J$5:$J$403"
_CC_B = "'Content Calendar'!$B$5:$B$403"
_CPR_T = "'Community & PR'!$T$5:$T$403"; _CPR_N = "'Community & PR'!$N$5:$N$403"
_CPR_B = "'Community & PR'!$B$5:$B$403"
_LO_AO = "'LinkedIn Outreach'!$AO$5:$AO$1203"; _LO_S = "'LinkedIn Outreach'!$S$5:$S$1203"
_LO_AL = "'LinkedIn Outreach'!$AL$5:$AL$1203"; _LO_B = "'LinkedIn Outreach'!$B$5:$B$1203"
_PP_V = "'Partnership Pipeline'!$V$5:$V$403"; _PP_W = "'Partnership Pipeline'!$W$5:$W$403"
_PP_X = "'Partnership Pipeline'!$X$5:$X$403"; _PP_B = "'Partnership Pipeline'!$B$5:$B$403"

_BR_DE = "'DAILY ENTRY'!$N$7:$N$1006"
_BR_CC = "'Content Calendar'!$S$5:$S$403"
_BR_CPR = "'Community & PR'!$U$5:$U$403"
_BR_LO = "'LinkedIn Outreach'!$AP$5:$AP$1203"
_BR_PP = "'Partnership Pipeline'!$Y$5:$Y$403"

from openpyxl.formatting.rule import ColorScaleRule as _CSR, FormulaRule as _FR

# Every row of these tables tests the SAME existence column that the untagged
# row does — the sheet's own identity column, not the employee name. Two
# defects came out of getting this wrong: a DAILY ENTRY row with minutes but
# no name and no tag fell through every bucket and vanished from the TOTAL,
# and a single stray dropdown click on an empty row was counted as a real
# lead. Objective/brand tags are DROPDOWNS — pre-clicking one is normal
# behaviour, so a tag alone can never make a row exist.
_ID_DE = "'DAILY ENTRY'!$A$7:$A$1006"
_ID_LO = "'LinkedIn Outreach'!$C$5:$C$1203"
_ID_CPR = "'Community & PR'!$C$5:$C$403"
_ID_PP = "'Partnership Pipeline'!$C$5:$C$403"
_LO_AG2 = "'LinkedIn Outreach'!$AG$5:$AG$1203"
_CPR_P2 = "'Community & PR'!$P$5:$P$403"


def _perf_table(first, cats, untag_label, d_de, d_cc, d_cpr, d_lo, d_pp):
    """One results table (categories + untagged + TOTAL) over a tag dimension."""
    untag = first + len(cats)
    total = untag + 1
    for i, cat in enumerate(cats):
        rr = first + i
        op.cell(rr, 1).value = cat
        op.cell(rr, 2).value = f'=COUNTIFS({_ID_DE},"<>",{d_de},$A{rr})'
        op.cell(rr, 3).value = f'=SUMIFS({_DE_G},{_ID_DE},"<>",{d_de},$A{rr})'
        op.cell(rr, 4).value = (f'=COUNTIFS({d_cc},$A{rr},{_CC_J},"Published")'
                                f'+COUNTIFS({d_cc},$A{rr},{_CC_J},"Repurposed")')
        op.cell(rr, 5).value = f'=COUNTIFS({_ID_CPR},"<>",{d_cpr},$A{rr})'
        op.cell(rr, 6).value = f'=COUNTIFS({_ID_LO},"<>",{d_lo},$A{rr})'
        # a meeting is counted from the meeting DATE. Counting the outcome
        # dropdown erased every meeting the moment it progressed to
        # "Application Started" or "Converted" — success deleted the evidence.
        op.cell(rr, 7).value = (f'=COUNTIFS({_LO_AG2},">0",{d_lo},$A{rr})'
                                f'+COUNTIFS({_CPR_P2},">0",{d_cpr},$A{rr})')
        # ">0" not "<>": a typed note in the signed-date cell was booking
        # revenue that then reconciled to no week and no month.
        op.cell(rr, 8).value = (f'=SUMIFS({_LO_AL},{_ID_LO},"<>",{d_lo},$A{rr})'
                                f'+SUMIFS({_PP_V},{d_pp},$A{rr},{_PP_W},">0")')
    rr = untag
    op.cell(rr, 1).value = untag_label
    op.cell(rr, 2).value = f'=COUNTIFS({_ID_DE},"<>",{d_de},"")'
    op.cell(rr, 3).value = f'=SUMIFS({_DE_G},{_ID_DE},"<>",{d_de},"")'
    op.cell(rr, 4).value = (f'=COUNTIFS({d_cc},"",{_CC_J},"Published")'
                            f'+COUNTIFS({d_cc},"",{_CC_J},"Repurposed")')
    op.cell(rr, 5).value = f'=COUNTIFS({_ID_CPR},"<>",{d_cpr},"")'
    op.cell(rr, 6).value = f'=COUNTIFS({_ID_LO},"<>",{d_lo},"")'
    op.cell(rr, 7).value = (f'=COUNTIFS({_LO_AG2},">0",{d_lo},"")'
                            f'+COUNTIFS({_CPR_P2},">0",{d_cpr},"")')
    op.cell(rr, 8).value = (f'=SUMIFS({_LO_AL},{_ID_LO},"<>",{d_lo},"")'
                            f'+SUMIFS({_PP_V},{_ID_PP},"<>",{d_pp},"",{_PP_W},">0")')
    op.cell(total, 1).value = "TOTAL"
    for _c in range(2, 9):
        _L = get_column_letter(_c)
        op.cell(total, _c).value = f'=SUM({_L}{first}:{_L}{untag})'
    for rr in range(first, total + 1):
        # Share of minutes divides by the REAL DAILY ENTRY total, so the TOTAL
        # row reads 100% only when nothing has leaked. Dividing by its own
        # total made that cell tautologically 100% and hid the one defect it
        # was placed there to expose.
        op.cell(rr, 9).value = f'=IF(SUM({_DE_G})=0,"",C{rr}/SUM({_DE_G}))'
        op.cell(rr, 10).value = f'=IF(G{rr}=0,"",ROUND(C{rr}/G{rr},0))'
    for rr in range(first, total + 1):
        for _c in range(1, 11):
            cell = op.cell(rr, _c)
            cell.font = BODYB if rr == total else BODY
            cell.border = BOX
            cell.alignment = Alignment(wrap_text=(_c == 1), vertical="center")
            if rr < total and rr % 2 == 0:
                cell.fill = ZEBRA
            if _c == 8:
                cell.number_format = '"$"#,##0'
            elif _c == 9:
                cell.number_format = "0%"
            elif _c > 1:
                cell.number_format = "#,##0"
        op.row_dimensions[rr].height = 18
    for _c in range(1, 11):
        op.cell(total, _c).fill = PatternFill("solid", fgColor="DDEBF7")
    op.conditional_formatting.add(
        f"C{first}:C{untag - 1}",
        _CSR(start_type="min", start_color="FFFFFF", end_type="max", end_color="9DC3E6"))
    op.conditional_formatting.add(
        f"H{first}:H{untag - 1}",
        _CSR(start_type="min", start_color="FFFFFF", end_type="max", end_color="A9D08E"))
    op.conditional_formatting.add(
        f"A{untag}:J{untag}",
        _FR(formula=[f"$B${untag}>0"], fill=PatternFill("solid", fgColor="F4CCCC")))
    return untag, total

_first = 4
_OBJ_FIRST = 4                      # captured: `_first` is reused as a loop
                                    # variable further down the build
_untagged, _total = _perf_table(_first, _OBJ_LIST, "(no objective set)",
                                _DE_M, _CC_F, _CPR_T, _LO_AO, _PP_X)

# value rank per effort category: which effort is worth the most per hour.
# Management judgment, pre-filled and editable (yellow) — the numbers beside
# it then show whether the team's minutes actually follow the value.
from canonical_lov import OBJECTIVE_RANKS as _ORANKS
_kh = op.cell(3, 11)
_kh.value = "Value rank (1 = most valuable) — edit if strategy changes"
_kh.font = H2; _kh.fill = BANDFILL; _kh.border = BOX
_kh.alignment = Alignment(wrap_text=True, vertical="center")
for _i, _obj in enumerate(_OBJ_LIST):
    rr = _first + _i
    _rank, _why = _ORANKS[_obj]
    c = op.cell(rr, 11)
    c.value = f"{_rank}  ·  {_why}"
    c.font = BODY; c.border = BOX; c.fill = YELLOW
    c.protection = Protection(locked=False)
    c.alignment = Alignment(wrap_text=True, vertical="center")
for rr in (_untagged, _total):
    op.cell(rr, 11).border = BOX
op.column_dimensions["K"].width = 36

# ---- block 2: the same results split by brand / property
from canonical_lov import BRANDS as _BR_LIST
def _reconcile(row, total_row):
    """The one line that proves nothing leaked out of the table above it."""
    c = op.cell(row, 1)
    c.value = ("Reconciliation — DAILY ENTRY minutes not counted above "
               "(rows with no date): ")
    c.font = GREYNOTE
    op.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    v = op.cell(row, 9)
    v.value = f"=SUM({_DE_G})-C{total_row}"
    v.font = Font(name="Arial", size=8, bold=True, color="C00000")
    v.number_format = "#,##0"
    v.alignment = Alignment(horizontal="right")
    op.row_dimensions[row].height = 14


_reconcile(_total + 1, _total)
_b_sec = _total + 3
section(op, _b_sec, 10, "RESULTS BY BRAND / PROPERTY  —  which certification or property the effort serves")
header_band(op, _b_sec + 1, ["Brand / property", "Activities logged", "Minutes logged",
                             "Content published", "Community actions", "Leads logged",
                             "Meetings booked", "Revenue + signed deals (USD)",
                             "Share of minutes", "Minutes per meeting"])
_BR_FIRST = _b_sec + 2
_b_untag, _b_total = _perf_table(_b_sec + 2, _BR_LIST, "(no brand set)",
                                 _BR_DE, _BR_CC, _BR_CPR, _BR_LO, _BR_PP)

_reconcile(_b_total + 1, _b_total)

# ---- block 3: minutes by person x objective (who carries which campaign)
_m_hdr = _b_total + 3
section(op, _m_hdr, 11, "MINUTES BY PERSON × OBJECTIVE  —  who is carrying which campaign")
_m_names = _m_hdr + 1                              # 19: roster headers
c0 = op.cell(_m_names, 1); c0.value = "Objective"
c0.font = H2; c0.fill = BANDFILL; c0.border = BOX
c0.alignment = Alignment(vertical="center")
for _s in range(10):                               # roster slots B55..B64
    cell = op.cell(_m_names, 2 + _s)
    cell.value = f"=IF('START HERE'!$B${55 + _s}=\"\",\"\",'START HERE'!$B${55 + _s})"
    cell.font = H2; cell.fill = BANDFILL; cell.border = BOX
    cell.alignment = Alignment(wrap_text=True, vertical="center")
op.row_dimensions[_m_names].height = 24
_m_first = _m_names + 1                            # 20
_m_untag = _m_first + len(_OBJ_LIST)               # 31
_m_total = _m_untag + 1                            # 32
for _i, _obj in enumerate(_OBJ_LIST):
    op.cell(_m_first + _i, 1).value = _obj
op.cell(_m_untag, 1).value = "(no objective set)"
op.cell(_m_total, 1).value = "TOTAL"
for _s in range(10):
    _L = get_column_letter(2 + _s)
    for _i in range(len(_OBJ_LIST)):
        rr = _m_first + _i
        op.cell(rr, 2 + _s).value = (f'=IF({_L}${_m_names}="","",'
                                     f'SUMIFS({_DE_G},{_DE_M},$A{rr},{_DE_B},{_L}${_m_names}))')
    op.cell(_m_untag, 2 + _s).value = (f'=IF({_L}${_m_names}="","",'
                                       f'SUMIFS({_DE_G},{_DE_M},"",{_DE_B},{_L}${_m_names}))')
    op.cell(_m_total, 2 + _s).value = (f'=IF({_L}${_m_names}="","",'
                                       f'SUM({_L}{_m_first}:{_L}{_m_untag}))')
for rr in range(_m_first, _m_total + 1):
    for _c in range(1, 12):
        cell = op.cell(rr, _c)
        cell.font = BODYB if rr == _m_total else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c == 1), vertical="center")
        if rr < _m_total and rr % 2 == 0:
            cell.fill = ZEBRA
        if _c > 1:
            cell.number_format = "#,##0"
    op.row_dimensions[rr].height = 18
for _c in range(1, 12):
    op.cell(_m_total, _c).fill = PatternFill("solid", fgColor="DDEBF7")
# minutes logged under a name that is not on the roster belong to nobody in
# this grid — say so on the grid rather than letting the columns quietly
# fail to add up to the team's minutes
_m_lost = _m_total + 1
op.cell(_m_lost, 1).value = "Not in this grid (name blank or not on the roster)"
op.cell(_m_lost, 2).value = f"=SUM({_DE_G})-SUM(B{_m_total}:K{_m_total})"
for _c in range(1, 12):
    cell = op.cell(_m_lost, _c)
    cell.font = Font(name="Arial", size=9, bold=(_c == 2), color="C00000")
    cell.border = BOX
    cell.alignment = Alignment(wrap_text=(_c == 1), vertical="center")
    if _c > 1:
        cell.number_format = "#,##0"
op.row_dimensions[_m_lost].height = 18
op.conditional_formatting.add(
    f"A{_m_lost}:K{_m_lost}",
    _FR(formula=[f"$B${_m_lost}>0"], fill=PatternFill("solid", fgColor="F4CCCC")))
_note = op.cell(_m_total + 3, 1)
_note.value = ("A blank column = an empty roster slot on START HERE §6. Minutes come from "
               "DAILY ENTRY only — outreach, content and community results are in the tables "
               "above. Set BOTH tags (Objective + For (brand)) on every row you log; the red "
               "rows show anything untagged. If the last row is above zero, somebody typed a "
               "name that does not match the roster exactly — fix the spelling on DAILY ENTRY.")
op.merge_cells(start_row=_m_total + 2, start_column=1, end_row=_m_total + 2, end_column=11)
_note.font = GREYNOTE; _note.alignment = WRAP
op.row_dimensions[_m_total + 2].height = 28
op.column_dimensions["A"].width = 34
op.freeze_panes = "B4"
op.protection.sheet = True
op.protection.selectLockedCells = False
op.protection.selectUnlockedCells = False
op.print_area = f"A1:K{_m_total + 3}"
op.sheet_properties.tabColor = "FF1F3864"

# ------------------------------------ 4d. JOB POSTINGS sheet
# One row per position per platform. A job post is also brand marketing:
# candidates read it, Google for Jobs indexes it, and stale posts look bad.
from canonical_lov import BRANDS as _JB_BRANDS, PLATFORMS as _JB_PLATFORMS

jp = wb.create_sheet("Job Postings")
jp.sheet_view.showGridLines = False
title_row(jp, 13, "JOB POSTINGS  —  every open position, on every platform that matters",
          "One row per position per platform. Log the day's posting work in DAILY ENTRY "
          "as 'Job post published' too, so it counts in activity. Close filled roles the "
          "same day — a stale listing reads as an unprofessional institute.")
header_band(jp, 3, ["Date posted", "Posted by", "Position title", "Position type",
                    "For (brand)", "Platform", "Country / region", "Job post URL",
                    "Status", "Applicants", "Shortlisted", "Hired", "Notes"],
            widths=[12, 16, 26, 20, 22, 24, 16, 30, 14, 11, 11, 9, 26])
_JP_LAST = 203
import datetime as _jdt
_jex = ["", "Employee 1", "Regional Marketing Executive (Gulf)", "Marketing",
        "PCI AI - Institute (umbrella)", "LinkedIn Company Page", "UAE + Gulf",
        "[post URL]", "Open", 4, 1, 0, "EXAMPLE ROW - delete before use."]
jp.cell(4, 1).value = _jdt.datetime(2026, 8, 17)
jp.cell(4, 1).number_format = "yyyy-mm-dd"
for _c, _v in enumerate(_jex[1:], 2):
    jp.cell(4, _c).value = _v
for rr in range(4, _JP_LAST + 1):
    for _c in range(1, 14):
        cell = jp.cell(rr, _c)
        cell.font = BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (3, 13)), vertical="center")
        cell.fill = PatternFill("solid", fgColor="E7E6E6") if rr == 4 else YELLOW
        cell.protection = Protection(locked=False)
    if rr == 4:
        continue
for _c in range(1, 14):
    jp.cell(4, _c).fill = PatternFill("solid", fgColor="E7E6E6")

_jp_dvs = [
    ("'START HERE'!$B$55:$B$64", "B4:B203"),
    ('"Marketing,Subject-matter expert / Examiner,Regional ambassador,Operations / Admin,Internship,Other"', "D4:D203"),
    (f"Lists!$O$4:$O${3 + len(_JB_BRANDS)}", "E4:E203"),
    (f"Lists!$J$4:$J${3 + len(_JB_PLATFORMS)}", "F4:F203"),
    ('"Draft,Open,Paused,Closed - filled,Closed - not filled"', "I4:I203"),
]
for _f1, _sq in _jp_dvs:
    _dv = DataValidation(type="list", formula1=_f1, allow_blank=True)
    _dv.error = "Pick a value from the dropdown — free-typed values fall out of every report."
    _dv.showErrorMessage = True
    jp.add_data_validation(_dv); _dv.add(_sq)
_dvd = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True)
_dvd.error = "Enter a real date (not text)"; _dvd.showErrorMessage = True
jp.add_data_validation(_dvd); _dvd.add("A4:A203")
_dvw = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
_dvw.error = "Enter a whole number"; _dvw.showErrorMessage = True
jp.add_data_validation(_dvw); _dvw.add("J4:L203")
jp.protection.sheet = True
jp.protection.deleteRows = False
jp.freeze_panes = "C4"
jp.print_area = "A1:M40"
jp.print_title_rows = "1:3"
jp.sheet_properties.tabColor = "FF2E75B6"

# ------------------------------------ 4d2. LINK BUILDING sheet (off-page SEO)
# The off-page engine gets its own log: every link prospect, outreach, and
# earned backlink — one row each, with anchor and dofollow tracking. Search
# Console shows what Google found; THIS shows what the team actually did.
lb = wb.create_sheet("Link Building")
lb.sheet_view.showGridLines = False
title_row(lb, 14, "LINK BUILDING  —  the off-page SEO engine, link by link",
          "One row per link prospect, from first contact to live. Referring domains are "
          "the currency of off-page SEO — but only earned ones: NEVER buy links, join "
          "link farms or swap bulk reciprocals (that is the link spam Google penalises). "
          "Tactics and weekly cadence live in GROWTH PLAYBOOK technique 23. Our target "
          "URL must be on an estate domain (START HERE §11). Cross-check earned links "
          "against Search Console monthly.")
header_band(lb, 3, ["Date", "Employee", "Tactic", "Target site / domain",
                    "Their page URL", "Our target URL", "Contact / route",
                    "Anchor text (planned)", "Status", "Link type", "Date live",
                    "Objective", "For (brand)", "Notes"],
            widths=[11, 14, 22, 20, 26, 26, 18, 18, 14, 12, 11, 26, 22, 20])
_LB_LAST = 403
lb.cell(4, 1).value = _jdt.datetime(2026, 8, 17)
lb.cell(4, 1).number_format = "yyyy-mm-dd"
_lbex = ["Employee 1", "Unlinked mention reclaimed", "pmworldjournal.com",
         "[article that mentions PCI]", "https://projectcontrolsinstitute.org/",
         "editor email", "Project Controls Institute", "Earned - live", "Dofollow",
         None, "Authority & Entity Building", "PCI AI - Institute (umbrella)",
         "EXAMPLE ROW - delete before use."]
for _c, _v in enumerate(_lbex, 2):
    lb.cell(4, _c).value = _v
lb.cell(4, 11).value = _jdt.datetime(2026, 8, 17)
lb.cell(4, 11).number_format = "yyyy-mm-dd"
for rr in range(4, _LB_LAST + 1):
    for _c in range(1, 15):
        cell = lb.cell(rr, _c)
        cell.font = BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (5, 6, 14)), vertical="center")
        cell.fill = PatternFill("solid", fgColor="E7E6E6") if rr == 4 else YELLOW
        cell.protection = Protection(locked=False)
    if rr > 4:
        lb.cell(rr, 1).number_format = "yyyy-mm-dd"
        lb.cell(rr, 11).number_format = "yyyy-mm-dd"
_lb_dvs = [
    ("'START HERE'!$B$55:$B$64", "B4:B403"),
    ('"Competitor backlink gap,Unlinked mention reclaimed,Broken link replacement,'
     'Resource page pitch,Guest post,Digital PR / journalist quote,Directory / listing,'
     'Partner / .edu link,Podcast / interview,Syndication canonical,Other"', "C4:C403"),
    ('"Prospect,Outreach sent,Follow-up,Earned - live,Rejected,Lost"', "I4:I403"),
    ('"Dofollow,Nofollow,Mention only,Unknown"', "J4:J403"),
    (f"Lists!$Y$4:$Y${3 + len(_OBJ_LIST)}", "L4:L403"),
    (f"Lists!$O$4:$O${3 + len(_JB_BRANDS)}", "M4:M403"),
]
for _f1, _sq in _lb_dvs:
    _dv = DataValidation(type="list", formula1=_f1, allow_blank=True)
    _dv.error = "Pick a value from the dropdown — free-typed values fall out of every report."
    _dv.showErrorMessage = True
    lb.add_data_validation(_dv); _dv.add(_sq)
_lbd = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True)
_lbd.error = "Enter a real date (not text)"; _lbd.showErrorMessage = True
lb.add_data_validation(_lbd)
for _part in ("A4:A403", "K4:K403"):
    _lbd.add(_part)
lb.protection.sheet = True
lb.protection.deleteRows = False
lb.freeze_panes = "C4"
lb.print_area = "A1:N40"
lb.print_title_rows = "1:3"
lb.sheet_properties.tabColor = "FF2E75B6"

# ------------------------------------ 4d3. CONTENT SCHEDULER sheet
# Scheduling as a managed system: what runs where, at what cadence, from when
# to when — with planned-vs-published tracking computed live from the
# Content Calendar. The reference block below the grid holds the researched
# truth about which platforms can schedule, where, and with what limits.
SCHED_TOOLS = ('"Native scheduler,Meta Business Suite,TikTok Studio,YouTube Studio,'
               'Telegram native,Metricool,Buffer,Later,Publer,Hootsuite,ESP scheduler,Manual"')
cs = wb.create_sheet("Content Scheduler")
cs.sheet_view.showGridLines = False
title_row(cs, 15, "CONTENT SCHEDULER  —  what runs where, at what cadence, fully tracked",
          "One row per running schedule (platform × cadence × window). Planned posts and "
          "published-in-window compute live from this row and the Content Calendar — "
          "coverage below 100% means the schedule is slipping. The reference table under "
          "the grid says which platforms support scheduling, where, and with what limits "
          "(researched Aug 2026). Dashboard shows active schedules and coverage.")
header_band(cs, 3, ["Platform", "For (brand)", "Objective", "Cadence", "Posts per cycle",
                    "Content format", "Scheduling tool", "Start date", "End date",
                    "Planned posts (auto)", "Published in window (auto)", "Coverage (auto)",
                    "Owner", "Status", "Notes"],
            widths=[24, 20, 24, 11, 10, 15, 16, 11, 11, 11, 12, 10, 13, 11, 22])
_CS_LAST = 103
_cs_ex = ["LinkedIn Company Page", "PCI AI - Institute (umbrella)", "Content & SEO Growth",
          "3x Weekly", 1, "Single image", "Native scheduler", None, None, None, None, None,
          "Employee 1", "Active", "EXAMPLE ROW - delete before use."]
for _c, _v in enumerate(_cs_ex, 1):
    cs.cell(4, _c).value = _v
cs.cell(4, 8).value = _jdt.datetime(2026, 8, 17)
cs.cell(4, 9).value = _jdt.datetime(2026, 9, 13)
for rr in range(4, _CS_LAST + 1):
    # Planned posts run off the cadence's REAL cycle length in days. The
    # earlier version treated "Monthly" as 0.25 per week, which is 13.04 posts
    # a year, not 12 — an 8.3% over-count that made every monthly schedule
    # look behind. The OR(...) guard also catches an end date before the start
    # date, which would otherwise produce a negative "planned" figure.
    cs.cell(rr, 10).value = (
        f'=IF(OR($H{rr}="",$I{rr}="",$D{rr}="",$I{rr}<$H{rr}),"",'
        f'ROUND((($I{rr}-$H{rr})+1)'
        f'/IF($D{rr}="Daily",1,IF($D{rr}="3x Weekly",7/3,IF($D{rr}="Weekly",7,'
        f'IF($D{rr}="Fortnightly",14,IF($D{rr}="Monthly",30.44,1000000)))))'
        f'*IF($E{rr}="",1,$E{rr}),0))')
    # Published-in-window matches the schedule's OWN brand and objective, not
    # just its platform. Matching on platform alone counted the same five real
    # posts against three different LinkedIn schedules — twelve times over.
    # A blank brand or objective on the schedule row means "any".
    # ISNUMBER() keeps a text-typed published date out: text compares as
    # greater than any number in Excel, so it would otherwise always match.
    cs.cell(rr, 11).value = (
        f'=IF(OR($A{rr}="",$H{rr}="",$I{rr}="",$I{rr}<$H{rr}),"",SUMPRODUCT('
        f"('Content Calendar'!$C$5:$C$403=$A{rr})"
        f"*ISNUMBER('Content Calendar'!$K$5:$K$403)"
        f"*('Content Calendar'!$K$5:$K$403>=$H{rr})"
        f"*('Content Calendar'!$K$5:$K$403<=$I{rr})"
        f'*((($B{rr}="")+' + "('Content Calendar'!$S$5:$S$403=$B" + str(rr) + "))>0)"
        f'*((($C{rr}="")+' + "('Content Calendar'!$F$5:$F$403=$C" + str(rr) + "))>0)"
        f"))")
    cs.cell(rr, 12).value = (
        f'=IF(OR($J{rr}="",$K{rr}="",N($J{rr})<=0),"",$K{rr}/$J{rr})')
    for _c in range(1, 16):
        cell = cs.cell(rr, _c)
        cell.font = BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c == 15), vertical="center")
        if _c in (10, 11, 12):
            pass                                   # formulas stay locked
        else:
            cell.protection = Protection(locked=False)
        if rr == 4:
            cell.fill = PatternFill("solid", fgColor="E7E6E6")
        elif _c not in (10, 11, 12):
            cell.fill = YELLOW
        if _c in (8, 9):
            cell.number_format = "yyyy-mm-dd"
        if _c in (10, 11):
            cell.number_format = "#,##0"
        if _c == 12:
            cell.number_format = "0%"
_cs_dvs = [
    (f"Lists!$J$4:$J${3 + len(_JB_PLATFORMS)}", "A4:A103"),
    (f"Lists!$O$4:$O${3 + len(_JB_BRANDS)}", "B4:B103"),
    (f"Lists!$Y$4:$Y${3 + len(_OBJ_LIST)}", "C4:C103"),
    ('"Daily,3x Weekly,Weekly,Fortnightly,Monthly"', "D4:D103"),
    (f"Lists!$H$4:$H$21", "F4:F103"),
    (SCHED_TOOLS, "G4:G103"),
    ("'START HERE'!$B$55:$B$64", "M4:M103"),
    ('"Planned,Active,Paused,Completed"', "N4:N103"),
]
for _f1, _sq in _cs_dvs:
    _dv = DataValidation(type="list", formula1=_f1, allow_blank=True)
    _dv.error = "Pick a value from the dropdown — free-typed values fall out of every report."
    _dv.showErrorMessage = True
    cs.add_data_validation(_dv); _dv.add(_sq)
_csd = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True)
_csd.error = "Enter a real date (not text)"; _csd.showErrorMessage = True
cs.add_data_validation(_csd)
for _part in ("H4:H103", "I4:I103"):
    _csd.add(_part)
_csw = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
_csw.error = "Whole number of posts per cadence cycle"; _csw.showErrorMessage = True
cs.add_data_validation(_csw); _csw.add("E4:E103")
_cov_ok = PatternFill("solid", fgColor="C6EFCE")
_cov_mid = PatternFill("solid", fgColor="FFE699")
_cov_bad = PatternFill("solid", fgColor="F4CCCC")
cs.conditional_formatting.add("L5:L103", _FR(formula=['AND(ISNUMBER($L5),$L5>1.25)'],
                                             fill=PatternFill("solid", fgColor="BDD7EE")))
cs.conditional_formatting.add("L5:L103", _FR(formula=['AND(ISNUMBER($L5),$L5>=1,$L5<=1.25)'], fill=_cov_ok))
cs.conditional_formatting.add("L5:L103", _FR(formula=['AND(ISNUMBER($L5),$L5>=0.7,$L5<1)'], fill=_cov_mid))
cs.conditional_formatting.add("L5:L103", _FR(formula=['AND(ISNUMBER($L5),$L5<0.7)'], fill=_cov_bad))
# a reversed date range is a data-entry error — say so instead of failing quietly
cs.conditional_formatting.add(
    "H5:I103", _FR(formula=['AND($H5<>"",$I5<>"",$I5<$H5)'], fill=_cov_bad))
# ---- reference block: which platforms can schedule, where, with what limits
SCHED_REF = [
 ("LinkedIn (Page + personal)", "YES — built into the composer, free",
  "10 min to 3 months ahead", "Text, single image, single video, link",
  "No polls, documents/carousels or multi-image; Pages cannot schedule reshares",
  "Page 3 posts/wk; personal daily"),
 ("Facebook Page", "YES — Meta Business Suite Planner",
  "75 days; 25 posts/day Meta family cap", "Posts, Reels, Stories, bulk",
  "Live video", "3-5 quality posts/wk"),
 ("Instagram", "YES — in-app (Business/Creator) + Meta Business Suite",
  "25/day, 75 days ahead", "Feed, carousels, Reels; STORIES via Business Suite only",
  "Stories from the IG app itself; Live", "3-5 feed/wk + Stories near-daily"),
 ("X (Twitter)", "YES — x.com desktop, free tier", "Up to 18 months",
  "Single posts (text, image, video, link)",
  "Threads (multi-post) not schedulable natively — use Buffer/Publer; no mobile; no polls",
  "1-2 posts/day; 1-2 threads/wk"),
 ("Threads", "YES — native since Jan 2025 (three-dot menu)", "75 days",
  "Standard posts (text, image, video)", "Replies cannot be scheduled",
  "1 post/day, mirroring X"),
 ("TikTok", "YES — TikTok Studio, desktop, Business/Creator accounts",
  "15 min to 10 DAYS only; scheduled posts cannot be edited", "Videos",
  "Anything from the mobile app; personal accounts", "1-3 videos/wk"),
 ("YouTube", "YES — Studio scheduled publish + Premieres", "About a year ahead",
  "Long-form, Shorts (scheduled individually), Premieres", "Batch-scheduling Shorts",
  "1 long-form/wk + Shorts more often"),
 ("Pinterest", "YES — native Pin scheduler (Business)", "30 days max; 10 pins queued max",
  "Standard image or video Pins, one at a time (Idea Pins were retired in 2023)",
  "Bulk scheduling", "Monthly batch — one Pin per published article"),
 ("Bluesky", "NO native scheduler (open feature request)", "—",
  "Third-party only: Buffer (approved partner), Fedica — threads schedulable there",
  "Everything natively", "Watching brief — mirror 2 posts/wk"),
 ("Snapchat", "PARTIAL — Public Story posts schedulable in-app", "Window undocumented",
  "Public Story snaps", "Spotlight has no native scheduling (Later via official API)",
  "Regular Story presence on posting days"),
 ("Telegram Channel", "YES — long-press send, then Schedule",
  "365 days ahead; 100 queued per channel", "All message types",
  "Recurring posts (needs bots)", "1 value post/wk, pin the campaign"),
 ("WhatsApp Channel", "NO — Channels have no scheduling", "—", "—",
  "Manual posting only (chat scheduling beta does not cover Channels)",
  "1 value broadcast/wk maximum"),
 ("Google Business Profile", "YES — 'Schedule this post' toggle, desktop (new — confirm in your dashboard)",
  "No published caps", "Updates, Offers, Events", "Mobile; recurring posts",
  "1 post/wk + answer reviews and questions"),
 ("Reddit", "Mod tools only, in communities you moderate",
  "Recurring hourly/daily/weekly", "Text/link posts",
  "Image/video scheduled posts; ordinary-user scheduling (Postpone if ever needed)",
  "Value-first — no fixed cadence"),
 ("Medium", "YES — Publish menu, 'Schedule for later'", "Publishes within ~5 min of set time",
  "Stories/articles incl. into publications accepting published posts",
  "Publications that only accept drafts", "1-4 articles/month"),
 ("Newsletter / ESP", "YES — every modern ESP, with send-time optimisation",
  "Effectively unlimited", "Campaigns / broadcasts", "—", "Weekly or fortnightly"),
]
_ref_hdr = _CS_LAST + 3                            # 106
section(cs, _ref_hdr, 15, "WHERE SCHEDULING WORKS  —  researched Aug 2026, re-verify 6-monthly")
_rh = _ref_hdr + 1
for _c, _h in enumerate(["Platform", "Native scheduler", "Window / limits",
                         "Schedulable", "Not schedulable / third-party route",
                         "PCI cadence (same figure as PLATFORM GUIDE)"], 1):
    hc = cs.cell(_rh, _c); hc.value = _h; hc.font = H2; hc.fill = BANDFILL
    hc.border = BOX; hc.alignment = Alignment(wrap_text=True, vertical="center")
cs.row_dimensions[_rh].height = 18
_rr2 = _rh + 1
for _row in SCHED_REF:
    for _c, _v in enumerate(_row, 1):
        cell = cs.cell(_rr2, _c)
        cell.value = _v
        cell.font = BODYB if _c == 1 else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if _rr2 % 2 == 0:
            cell.fill = ZEBRA
    _need = max(len(_row[1]) / 20, len(_row[3]) / 24, len(_row[4]) / 24, 2.0)
    cs.row_dimensions[_rr2].height = max(28, int(_need + 1) * 10 + 4)
    _rr2 += 1
_csn = cs.cell(_rr2 + 1, 1)
_csn.value = ("RECOMMENDED $0 STACK (researched): Metricool Free — 1 brand, 9 networks free "
              "including Threads, Bluesky, Google Business Profile, TikTok, Pinterest and "
              "YouTube — plus the NATIVE schedulers for LinkedIn (3 months), X (18 months), "
              "Telegram (365 days), Medium and the ESP. LinkedIn and X sit outside Metricool's "
              "free tier but have the strongest native schedulers, so the gap costs nothing. "
              "Nothing mainstream schedules WhatsApp Channels or ordinary Reddit posts — those "
              "stay manual. Flags: the GBP scheduler is newly rolled out (confirm in your own "
              "dashboard before relying on it); Metricool's free monthly post cap is reported "
              "inconsistently (20 vs 50) — check in-app.")
_csn.font = GREYNOTE
cs.merge_cells(start_row=_rr2 + 1, start_column=1, end_row=_rr2 + 1, end_column=15)
_csn.alignment = WRAP
cs.row_dimensions[_rr2 + 1].height = 52
cs.protection.sheet = True
cs.protection.deleteRows = False
cs.freeze_panes = "C4"
cs.print_area = "A1:O45"
cs.print_title_rows = "1:3"
cs.sheet_properties.tabColor = "FF2E75B6"

# ------------------------------------ 4e. WEEKLY PULSE sheet
# Real performance tracking: this week against last week and the 4-week
# average, computed live from the dated logs. Monday-anchored weeks.
wp = wb.create_sheet("Weekly Pulse")
wp.sheet_view.showGridLines = False
title_row(wp, 6, "WEEKLY PULSE  —  this week against last week, live from the logs",
          "Nothing is typed here. Weeks run Monday-Sunday; 'This week' counts from the "
          "Monday shown below. Green = up on last week, red = down. If a high-value "
          "effort (see Objective Performance ranks) is falling week on week, rebalance "
          "before month-end — this page is the early warning, the Dashboard is the "
          "cumulative record.")
wp.cell(3, 1).value = "Week starts (Monday)"
wp.cell(3, 1).font = BODYB
wp.cell(3, 2).value = "=TODAY()-MOD(TODAY()-2,7)"
wp.cell(3, 2).number_format = "yyyy-mm-dd"
wp.cell(3, 2).font = BODYB
header_band(wp, 5, ["KPI", "This week", "Last week", "Change",
                    "4-week avg (completed weeks)", "Where it comes from"],
            widths=[32, 12, 12, 11, 16, 40])

_W = "$B$3"
def _win(rng, datecol, extra=""):
    return (f"{rng},{datecol},\">=\"&{_W},{datecol},\"<\"&{_W}+7{extra}")
_DEA = "'DAILY ENTRY'!$A$7:$A$1006"; _DEF = "'DAILY ENTRY'!$F$7:$F$1006"
_DED = "'DAILY ENTRY'!$D$7:$D$1006"; _DEG2 = "'DAILY ENTRY'!$G$7:$G$1006"
_DEB2 = "'DAILY ENTRY'!$B$7:$B$1006"
_CCK = "'Content Calendar'!$K$5:$K$403"
_CPRA = "'Community & PR'!$A$5:$A$403"; _CPRB2 = "'Community & PR'!$B$5:$B$403"
_CPRP = "'Community & PR'!$P$5:$P$403"
_LOAG = "'LinkedIn Outreach'!$AG$5:$AG$1203"; _LOAK = "'LinkedIn Outreach'!$AK$5:$AK$1203"
_LOAL2 = "'LinkedIn Outreach'!$AL$5:$AL$1203"; _LOA2 = "'LinkedIn Outreach'!$A$5:$A$1203"
_LOC2 = "'LinkedIn Outreach'!$C$5:$C$1203"
_JPA = "'Job Postings'!$A$5:$A$203"; _JPC = "'Job Postings'!$C$5:$C$203"
_LBK = "'Link Building'!$K$5:$K$403"
_PPV2 = "'Partnership Pipeline'!$V$5:$V$403"; _PPW2 = "'Partnership Pipeline'!$W$5:$W$403"

def _sumw(vals, datecol, off, extra=""):
    lo_ = f"{_W}{'-' + str(off) if off else ''}"
    hi_ = f"{_W}{'-' + str(off - 7) if off > 7 else ('+7' if off == 0 else '')}"
    return f'SUMIFS({vals},{datecol},">="&{lo_},{datecol},"<"&{hi_}{extra})'
def _cntw(datecol, off, extra=""):
    lo_ = f"{_W}{'-' + str(off) if off else ''}"
    hi_ = f"{_W}{'-' + str(off - 7) if off > 7 else ('+7' if off == 0 else '')}"
    return f'COUNTIFS({datecol},">="&{lo_},{datecol},"<"&{hi_}{extra})'

_act = lambda t: f',{_DED},"{t}"'
PULSE = [
 ("Minutes logged",            lambda o: _sumw(_DEG2, _DEA, o),                     "DAILY ENTRY minutes, by entry date"),
 ("Activities logged",         lambda o: _cntw(_DEA, o, f',{_DEB2},"<>"'),          "DAILY ENTRY rows with an employee"),
 ("Leads researched",          lambda o: _sumw(_DEF, _DEA, o, _act("Lead researched (Sales Navigator)")), "DAILY ENTRY 'how many', lead research rows"),
 ("Connection requests sent",  lambda o: _sumw(_DEF, _DEA, o, _act("Connection request sent")), "DAILY ENTRY 'how many'"),
 ("Messages sent (first + follow-up)", lambda o: _sumw(_DEF, _DEA, o, _act("First message sent")) + "+" + _sumw(_DEF, _DEA, o, _act("Follow-up message sent")), "DAILY ENTRY 'how many'"),
 # "Published" must mean the same thing here as on the Dashboard and Objective
 # Performance: status Published/Repurposed — not merely a date in the cell.
 ("Content published",         lambda o: (_cntw(_CCK, o, f',{_CC_J},"Published"') + "+"
                                          + _cntw(_CCK, o, f',{_CC_J},"Repurposed"')),
  "Content Calendar: status Published/Repurposed, by published date"),
 ("Community & PR actions",    lambda o: _cntw(_CPRA, o, f',{_CPRB2},"<>"'),        "Community & PR rows by date"),
 ("Email campaigns sent (recipients)", lambda o: _sumw(_DEF, _DEA, o, _act("Email campaign sent")), "DAILY ENTRY 'how many' = emails delivered"),
 ("WhatsApp / Telegram / SMS sent", lambda o: _sumw(_DEF, _DEA, o, _act("WhatsApp / Telegram / SMS sent")), "DAILY ENTRY 'how many' = messages delivered"),
 ("Job posts published",       lambda o: _cntw(_JPA, o, f',{_JPC},"<>"'),           "Job Postings rows by date posted"),
 ("Backlinks gone live",       lambda o: _cntw(_LBK, o),                            "Link Building rows by the date the link went live"),
 ("Meetings booked",           lambda o: _cntw(_LOAG, o) + "+" + _cntw(_CPRP, o),   "LinkedIn Outreach meeting dates + Community & PR meeting dates"),
 ("Revenue recorded (USD)",    lambda o: _sumw(_LOAL2, _LOAK, o) + "+" + _sumw(_PPV2, _PPW2, o), "Outreach revenue by purchase date + signed partnership deals"),
 ("New leads logged",          lambda o: _cntw(_LOA2, o, f',{_LOC2},"<>"'),         "LinkedIn Outreach rows by date added"),
]
_pr = 6
_PULSE_ROW = {}          # label -> row, so Dashboard tiles can never drift again
for _name, _fn, _src in PULSE:
    _PULSE_ROW[_name] = _pr
    wp.cell(_pr, 1).value = _name
    wp.cell(_pr, 2).value = "=" + _fn(0)
    wp.cell(_pr, 3).value = "=" + _fn(7)
    wp.cell(_pr, 4).value = (f'=IF(C{_pr}=0,IF(B{_pr}=0,"","new"),(B{_pr}-C{_pr})/C{_pr})')
    # The 4-week average must span four COMPLETED weeks. The earlier version
    # ran a 28-day window ending mid-week and still divided by 4, so on a
    # Monday morning a team that never varied read 20% down on its own
    # average — and the size of the lie moved with the day of the week.
    _t = _fn(28)
    _wide = _t.replace(f'"<"&{_W}-21', f'"<"&{_W}')
    wp.cell(_pr, 5).value = f"=({_wide})/4"
    wp.cell(_pr, 6).value = _src
    for _c in range(1, 7):
        cell = wp.cell(_pr, _c)
        cell.font = BODYB if _c == 1 else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (1, 6)), vertical="center")
        if _pr % 2 == 0:
            cell.fill = ZEBRA
        if _c in (2, 3, 5):
            cell.number_format = "#,##0"
        if _c == 4:
            cell.number_format = "+0%;-0%"
    if "Revenue" in _name:
        for _c in (2, 3, 5):
            wp.cell(_pr, _c).number_format = '"$"#,##0'
    wp.cell(_pr, 6).font = GREYNOTE
    wp.row_dimensions[_pr].height = 20
    _pr += 1
_p_end = _pr - 1
_up = PatternFill("solid", fgColor="C6EFCE"); _dn = PatternFill("solid", fgColor="F4CCCC")
op_ = wp.conditional_formatting
op_.add(f"D6:D{_p_end}", _FR(formula=[f'AND(ISNUMBER($D6),$D6>0)'], fill=_up))
op_.add(f"D6:D{_p_end}", _FR(formula=[f'AND(ISNUMBER($D6),$D6<0)'], fill=_dn))

_pr += 1
section(wp, _pr, 6, "MINUTES THIS WEEK BY PERSON  —  who showed up"); _pr += 1
for _c, _h in ((1, "Person"), (2, "This week"), (3, "Last week"), (4, "Change")):
    hc = wp.cell(_pr, _c); hc.value = _h; hc.font = H2; hc.fill = BANDFILL
    hc.border = BOX; hc.alignment = Alignment(vertical="center")
wp.row_dimensions[_pr].height = 18
_pr += 1
for _s in range(10):
    rr = _pr + _s
    wp.cell(rr, 1).value = f"=IF('START HERE'!$B${55 + _s}=\"\",\"\",'START HERE'!$B${55 + _s})"
    wp.cell(rr, 2).value = (f'=IF($A{rr}="","",'
                            + _sumw(_DEG2, _DEA, 0, f",{_DEB2},$A{rr}") + ")")
    wp.cell(rr, 3).value = (f'=IF($A{rr}="","",'
                            + _sumw(_DEG2, _DEA, 7, f",{_DEB2},$A{rr}") + ")")
    wp.cell(rr, 4).value = (f'=IF(OR($A{rr}="",C{rr}=""),"",'
                            f'IF(C{rr}=0,IF(B{rr}=0,"","new"),(B{rr}-C{rr})/C{rr}))')
    for _c in range(1, 5):
        cell = wp.cell(rr, _c)
        cell.font = BODY; cell.border = BOX
        if rr % 2 == 0: cell.fill = ZEBRA
        if _c in (2, 3): cell.number_format = "#,##0"
        if _c == 4: cell.number_format = "+0%;-0%"
op_.add(f"D{_pr}:D{_pr + 9}", _FR(formula=[f'AND(ISNUMBER($D{_pr}),$D{_pr}>0)'], fill=_up))
op_.add(f"D{_pr}:D{_pr + 9}", _FR(formula=[f'AND(ISNUMBER($D{_pr}),$D{_pr}<0)'], fill=_dn))
_ppl_first = _pr
_pr += 10
# The per-person rows could never add up to the team row: a DAILY ENTRY row
# logged with no name — or with a name that does not match the roster exactly —
# lands in the team total and in nobody's row, with nothing on the page to say
# so. These two rows make the block reconcile or visibly fail to.
_unat = _pr
wp.cell(_unat, 1).value = "Unattributed (name blank or not on the roster)"
wp.cell(_unat, 2).value = ("=" + _sumw(_DEG2, _DEA, 0)
                           + "-SUM(B" + str(_ppl_first) + ":B" + str(_ppl_first + 9) + ")")
wp.cell(_unat, 3).value = ("=" + _sumw(_DEG2, _DEA, 7)
                           + "-SUM(C" + str(_ppl_first) + ":C" + str(_ppl_first + 9) + ")")
_tot = _pr + 1
wp.cell(_tot, 1).value = "TOTAL  (must equal 'Minutes logged' above)"
wp.cell(_tot, 2).value = f"=SUM(B{_ppl_first}:B{_unat})"
wp.cell(_tot, 3).value = f"=SUM(C{_ppl_first}:C{_unat})"
for _rr in (_unat, _tot):
    for _c in range(1, 5):
        cell = wp.cell(_rr, _c)
        cell.font = BODYB if _rr == _tot else BODY
        cell.border = BOX
        if _c in (2, 3):
            cell.number_format = "#,##0"
    wp.cell(_rr, 1).alignment = Alignment(wrap_text=True, vertical="center")
op_.add(f"B{_unat}:C{_unat}", _FR(formula=[f"$B${_unat}>0"], fill=_dn))
_pr += 2
_wn = wp.cell(_pr + 1, 1)
_wn.value = ("Meetings and revenue count by their own dates (meeting date, purchase date, "
             "contract date) — log those columns or this page under-reports. The 4-week "
             "average covers the four COMPLETED weeks before this one, so it does not move "
             "with the day of the week. Anything on the 'Unattributed' row is minutes "
             "logged under a name that is not on the roster — fix the spelling on DAILY "
             "ENTRY and it will return to its owner.")
_wn.font = GREYNOTE
wp.merge_cells(start_row=_pr + 1, start_column=1, end_row=_pr + 1, end_column=6)
wp.freeze_panes = "A6"
wp.protection.sheet = True
wp.protection.selectLockedCells = False
wp.protection.selectUnlockedCells = False
wp.print_area = f"A1:F{_pr + 2}"
wp.sheet_properties.tabColor = "FF1F3864"

# ------------------------------------ 4f. How-To Guides: the missing workstreams
htg = wb["How-To Guides"]
_HTG_NEW = [
 ("Email marketing (ESP)",
  "The owned list is the highest-ROI channel in marketing — the audience you do not rent from an algorithm.",
  "Pick the ESP, authenticate the domain (SPF, DKIM, DMARC), add signup forms to the site and lead magnets, build the welcome automation.",
  "Newsletter every fortnight; automated nurture sequences always-on (playbook technique 9), segmented by funnel stage. Log every campaign in DAILY ENTRY as 'Email campaign sent' with How many = emails delivered.",
  "Open rate above 35%, click rate above 2.5%, unsubscribes under 0.3%.",
  "Buying lists (illegal under GDPR and kills deliverability). Sending the same email to everyone. No double opt-in for EU contacts.",
  "Campaign report link or screenshot", "3 h"),
 ("WhatsApp / Telegram / SMS",
  "The Gulf audience lives on WhatsApp — 90%+ open rates for reminders. Telegram carries the same role in several markets.",
  "WhatsApp Business API via a provider, the WhatsApp Channel, the Telegram channel, an SMS provider; record consent for every contact.",
  "Exam, webinar and application reminders as they occur; at most one value broadcast per week. Log sends in DAILY ENTRY as 'WhatsApp / Telegram / SMS sent'.",
  "Opt-outs under 1%; replies arriving — a broadcast that gets replies is working.",
  "Messaging without opt-in (blocked by the platform and illegal in most markets). More than one broadcast a week — people mute, then you have nothing.",
  "Broadcast stats screenshot", "1.5 h"),
 ("Job postings & hiring",
  "Every job post is also brand marketing: candidates read it, Google for Jobs indexes it, and rivals see an institute that is growing.",
  "Careers page with JobPosting schema so Google for Jobs picks roles up free; employer profiles (LinkedIn, Bayt, Indeed) come from Platform Setup.",
  "Keep the Job Postings tab current — one row per platform per position. Post each open role on 3+ platforms including one Gulf board. Close filled roles the same day.",
  "Every open role visible on 3+ platforms; applicant counts logged weekly.",
  "Dead listings left open — it reads as a disorganised institute. Salary stated on one platform and hidden on another.",
  "Post URL in the Job Postings tab", "1 h"),
 ("Press releases & free news",
  "Indexed third-party references are what AI engines read when they decide PCI is real. Journalists are not the audience of free wires — the index is.",
  "Own-site newsroom page with NewsArticle schema; free accounts on PRLog and openPR (see PR & Target Directory for the verified verdicts).",
  "When there is real news (credential launch, partnership, milestone): newsroom post first, then PRLog, then openPR (1 per 30 days). Playbook technique 16 has the full pipeline.",
  "Each release indexed on 2+ non-PCI domains within a week (search the headline).",
  "Expecting journalist coverage from free wires. Keyword-stuffed releases — links are nofollow by Google policy, so stuffing gains nothing and looks desperate.",
  "Release URLs logged in Community & PR", "1 h / month"),
 ("Educational publishing & DOIs",
  "Frameworks with DOIs and open educational resources put PCI in the academic-adjacent record — the strongest legitimacy signal a young institute can buy for free.",
  "Zenodo account + 'PCI AI' community (free DOIs); OER Commons and MERLOT member accounts; TrainingZone/HRZone contributor profile.",
  "Quarterly: one CC-licensed exam-prep primer to OER Commons + MERLOT, one framework PDF to Zenodo with named authors, one practical article to TrainingZone/HRZone.",
  "DOIs citable and cited in PCI articles; resources approved by the library curators.",
  "Posting marketing copy as 'education' — curators remove it and the account with it.",
  "DOI links and resource URLs", "2 h / month"),
 ("SEO & keywords",
  "Search is the compounding channel: rankings earned this quarter pay for years. The Keyword Plan tells you exactly which battles to pick.",
  "Keyword Plan sheet reviewed with the manager; Google Search Console verified; SEO Clusters set up with the pillar topics.",
  "Pick from the Keyword Plan by difficulty — Easy first, Medium next, Hard only via pillar pages. One spoke article per week, logged in Content Calendar and mapped in SEO Clusters.",
  "Easy keywords in the top 10 within 90 days; impressions rising month on month in Search Console.",
  "Chasing Hard head terms first (PMI and AACE own them — earn authority on Easy/Medium first). Writing content without checking the plan.",
  "Monthly Search Console screenshot", "4 h"),
]
_hr = htg.max_row + 1
for _row in _HTG_NEW:
    for _c, _v in enumerate(_row, 1):
        cell = htg.cell(_hr, _c)
        cell.value = _v
        cell._style = copymod.copy(htg.cell(4 + (_hr % 2), _c)._style)
    htg.row_dimensions[_hr].height = 103.5   # match the existing table rhythm
    _hr += 1

# ------------------------------------ 4g. Dashboard polish
db8 = wb["Dashboard"]
db8.sheet_view.showGridLines = False
for _col, _w in (("A", 40), ("B", 14), ("C", 16), ("D", 12), ("E", 48)):
    db8.column_dimensions[_col].width = _w
_v2 = db8.cell(2, 1).value
if isinstance(_v2, str) and "Weekly Pulse" not in _v2:
    db8.cell(2, 1).value = (_v2.rstrip() + "  This page is the cumulative record; "
                            "week-on-week movement lives on Weekly Pulse.")
_att = PatternFill("solid", fgColor="C6EFCE")
_beh = PatternFill("solid", fgColor="FCE4D6")
db8.conditional_formatting.add(
    "D42:D49", _FR(formula=['AND(ISNUMBER($D42),$D42>=1)'], fill=_att))
db8.conditional_formatting.add(
    "D42:D49", _FR(formula=['AND(ISNUMBER($D42),$D42<0.7,$C42>0)'], fill=_beh))

# ------------------------------------ 4h. KEYWORD PLAN sheet
# 76 keywords with SERP-grounded difficulty so SEO effort goes where it can
# win: Easy first, Medium next, Hard only via pillars.
from keywords_data import KEYWORDS, P1_WHY

kw = wb.create_sheet("Keyword Plan")
kw.sheet_view.showGridLines = False
title_row(kw, 14, "KEYWORD PLAN  —  where the SEO effort wins (researched Aug 2026)",
          "76 keywords, difficulty graded from live SERP sampling (sampled rows marked ✓): EASY = "
          "thin/forum/programmatic pages rank — attack now. MEDIUM = mid-authority mix — "
          "winnable with a genuinely better page. HARD = PMI / AACE / RICS / big-vendor "
          "territory — pillar pages only, patience. Volume bands are honest editorial "
          "estimates, not tool metrics. CRITICAL: projectcontrolsinstitute.com (a "
          "competitor training provider) owns PCI's exact name in Google today — the "
          "brand-defense row is attack keyword #1. Re-verify quarterly in Search "
          "Console; re-check the P1 SERPs monthly.")
header_band(kw, 3, ["Keyword", "Cluster", "Intent", "Funnel", "Volume band",
                    "Difficulty", "Who ranks today (Aug 2026)", "Asset to build",
                    "Priority", "SERP ✓", "Owner", "Status", "Published URL", "Notes"],
            widths=[34, 11, 12, 8, 11, 10, 44, 34, 8, 7, 14, 14, 26, 22])
_kwr = 4
for (k, cl, intent, fun, vol, diff, who, asset, pri, sampled) in KEYWORDS:
    vals = [k, cl, intent, fun, vol, diff, who, asset, pri, "✓" if sampled else ""]
    for _c, _v in enumerate(vals, 1):
        cell = kw.cell(_kwr, _c)
        cell.value = _v
        cell.font = BODYB if _c in (1, 6, 9) else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (1, 7, 8)), vertical="center",
                                   horizontal="center" if _c in (9, 10) else "general")
        if _kwr % 2 == 0:
            cell.fill = ZEBRA
    for _c in (11, 12, 13, 14):
        cell = kw.cell(_kwr, _c)
        cell.border = BOX
        cell.fill = YELLOW
        cell.font = BODY
        cell.protection = Protection(locked=False)
        cell.alignment = WRAP
    kw.row_dimensions[_kwr].height = 26
    _kwr += 1
_kw_last = _kwr - 1
_easy = PatternFill("solid", fgColor="C6EFCE"); _med = PatternFill("solid", fgColor="FFE699")
_hard = PatternFill("solid", fgColor="F4CCCC")
kw.conditional_formatting.add(f"F4:F{_kw_last}", _FR(formula=['$F4="Easy"'], fill=_easy))
kw.conditional_formatting.add(f"F4:F{_kw_last}", _FR(formula=['$F4="Medium"'], fill=_med))
kw.conditional_formatting.add(f"F4:F{_kw_last}", _FR(formula=['$F4="Hard"'], fill=_hard))
kw.conditional_formatting.add(f"A4:A{_kw_last}",
                              _FR(formula=['$I4="P1"'], fill=PatternFill("solid", fgColor="FFE699")))
_kw_dvs = [("'START HERE'!$B$55:$B$64", f"K4:K{_kw_last}"),
           ('"Not started,Brief written,In production,Published,Ranking top 10,Needs refresh"',
            f"L4:L{_kw_last}")]
for _f1, _sq in _kw_dvs:
    _dv = DataValidation(type="list", formula1=_f1, allow_blank=True)
    _dv.showErrorMessage = True
    kw.add_data_validation(_dv); _dv.add(_sq)

_kwr += 1
section(kw, _kwr, 14, "THE 10 P1 ATTACK KEYWORDS  —  start here, in this order")
_kwr += 1
for _i, (_k, _why) in enumerate(P1_WHY, 1):
    kw.cell(_kwr, 1).value = f"{_i}.  {_k}"
    kw.cell(_kwr, 1).font = BODYB
    kw.cell(_kwr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    kw.merge_cells(start_row=_kwr, start_column=2, end_row=_kwr, end_column=14)
    kw.cell(_kwr, 2).value = _why
    kw.cell(_kwr, 2).font = BODY
    kw.cell(_kwr, 2).alignment = WRAP
    if _kwr % 2 == 0:
        for _c in range(1, 15):
            kw.cell(_kwr, _c).fill = ZEBRA
    kw.row_dimensions[_kwr].height = 24
    _kwr += 1
_mn = kw.cell(_kwr + 1, 1)
_mn.value = ("Method: difficulty from live SERP sampling of 33 queries + competition logic "
             "(Hard = PMI/AACE/APMG/RICS/university/big-vendor page 1; Easy = forums, thin "
             "programmatic pages, solo niche sites, or a visible gap). No tool metrics were "
             "used — never quote the volume bands as data. Work Easy first; a new domain "
             "beats nobody at Hard head terms until authority accrues.")
_mn.font = GREYNOTE
kw.merge_cells(start_row=_kwr + 1, start_column=1, end_row=_kwr + 1, end_column=14)
kw.row_dimensions[_kwr + 1].height = 30
kw.freeze_panes = "B4"
kw.protection.sheet = True
kw.print_area = f"A1:P{_kwr + 2}"
kw.print_title_rows = "1:3"
kw.sheet_properties.tabColor = "FF2E75B6"

# ------------------------------------ 4h2. ARTICLE BANK sheet (5,000+ briefs)
# The editorial engine: every brief carries its keywords and a ready-to-use
# AI writing prompt, so nobody ever asks "what should I write?" again.
from article_bank import generate as _ab_generate

_AB = _ab_generate(5000)
ab = wb.create_sheet("Article Bank")
ab.sheet_view.showGridLines = False
title_row(ab, 15, "ARTICLE BANK  —  " + f"{len(_AB):,} article briefs with keywords and AI prompts",
          "Built from competitor mining, question-pattern research, verified standards "
          "(AACE RPs, EIA-748, NEC4, FIDIC, IFRS) and the Keyword Plan. Two independent "
          "columns: EFFORT is how much work the piece is; PRIORITY is the order to write in "
          "(P1 gold flagships first). SERP difficulty lives on the Keyword Plan. Claim a row "
          "with Owner + Status, paste the AI prompt into your writing tool, then EDIT like an "
          "expert — the prompt bans invented statistics, and data-study briefs must be verified "
          "against named sources at writing time.")
header_band(ab, 3, ["ID", "Article title", "Pillar (same 7 as SEO Clusters)", "Topic cluster",
                    "Format", "Audience", "Funnel",
                    "Effort", "Priority", "Primary keyword", "Supporting keywords", "Words",
                    "AI writing prompt (copy the cell)", "Owner", "Status", "Published URL"],
            widths=[10, 62, 26, 22, 14, 34, 9, 9, 9, 30, 40, 13, 52, 14, 15, 26])
from article_bank import pillar_for as _pillar_for
_ab_first = 4
_AB_ID = {}                      # primary keyword -> AB id, for the Keyword Plan join
_AB_TITLE_ID = {}                # title -> AB id, for the SEO Clusters join
for _i, (t, cl, fmt, eff, pri, pk, sk, aud, fun, words, prompt) in enumerate(_AB):
    rr = _ab_first + _i
    _id = f"AB-{_i + 1:05d}"
    _AB_ID.setdefault(pk.lower(), _id)
    _AB_TITLE_ID.setdefault(t.lower(), _id)
    vals = [_id, t, _pillar_for(cl, t), cl, fmt, aud, fun, eff, pri, pk, sk, words, prompt]
    for _c, _v in enumerate(vals, 1):
        cell = ab.cell(rr, _c)
        cell.value = _v
        cell.font = BODY
        cell.alignment = Alignment(vertical="top", wrap_text=(_c in (2, 3, 6, 10, 11)))
    ab.cell(rr, 2).font = BODYB
    for _c in (14, 15, 16):
        cell = ab.cell(rr, _c)
        cell.fill = YELLOW
        cell.font = BODY
        cell.protection = Protection(locked=False)
    # rows are sized for the longest wrapped cell so nothing is clipped
    ab.row_dimensions[rr].height = max(
        15, 12 * max(1, -(-len(t) // 60), -(-len(aud) // 33), -(-len(sk) // 39)) + 3)
_ab_last = _ab_first + len(_AB) - 1
ab.conditional_formatting.add(
    f"A{_ab_first}:P{_ab_last}",
    _FR(formula=["MOD(ROW(),2)=0"], fill=ZEBRA))
ab.conditional_formatting.add(
    f"A{_ab_first}:B{_ab_last}",
    _FR(formula=[f'$I{_ab_first}="P1"'], fill=PatternFill("solid", fgColor="FFE699")))
ab.conditional_formatting.add(                       # published rows go green
    f"A{_ab_first}:B{_ab_last}",
    _FR(formula=[f'$O{_ab_first}="Published"'], fill=PatternFill("solid", fgColor="E2EFDA")))
_ab_dvs = [("'START HERE'!$B$55:$B$64", f"N{_ab_first}:N{_ab_last}"),
           ('"Not started,Assigned,Drafted,Published,Ranking top 10,Needs refresh"',
            f"O{_ab_first}:O{_ab_last}")]
for _f1, _sq in _ab_dvs:
    _dv = DataValidation(type="list", formula1=_f1, allow_blank=True)
    _dv.showErrorMessage = True
    ab.add_data_validation(_dv); _dv.add(_sq)
ab.auto_filter.ref = f"A3:P{_ab_last}"
ab.freeze_panes = "C4"
ab.protection.sheet = True
ab.print_area = "A1:L43"
ab.print_title_rows = "1:3"
ab.sheet_properties.tabColor = "FF2E75B6"

# ---- the join: Keyword Plan and SEO Clusters now point AT briefs by ID
# An independent review found the SEO chain was three islands: 70% of the
# planned keywords had no brief at all (including all ten of the P1 attack
# terms), the seven cluster spokes had none, and the word "Cluster" meant
# three unrelated things on the three sheets. The bank is now generated FROM
# the plan, both sheets carry the brief's ID, and one shared Pillar
# vocabulary runs across all three.
_kp_hdr = ["Pillar (same 7 as SEO Clusters)", "Article Bank ID (the brief)"]
for _i, _h in enumerate(_kp_hdr):
    _c = 15 + _i
    _hc = kw.cell(3, _c)
    _hc.value = _h
    _hc.font = H2; _hc.fill = BANDFILL; _hc.border = BOX
    _hc.alignment = Alignment(wrap_text=True, vertical="center")
    kw.column_dimensions[get_column_letter(_c)].width = 26 if _i == 0 else 16
from article_bank import PILLARS as _PILLARS
_KP_PILLAR = {"Brand": "Certification and careers", "Core": "Certification and careers",
              "Role": "Certification and careers", "Conquest": "Certification and careers",
              "Skills": "Project controls fundamentals", "AI": "AI in project controls",
              "Geo": "Certification and careers", "Comparison": "Certification and careers",
              "Exam prep": "Certification and careers",
              "Honorary": "Certification and careers"}
_kp_missing = []
for _i, _kwrow in enumerate(KEYWORDS):
    rr = 4 + _i
    _term = _kwrow[0]
    kw.cell(rr, 15).value = _KP_PILLAR.get(_kwrow[1], "Project controls fundamentals")
    _abid = _AB_ID.get(_term.lower())
    kw.cell(rr, 16).value = _abid or "—"
    if not _abid:
        _kp_missing.append(_term)
    for _c in (15, 16):
        cell = kw.cell(rr, _c)
        cell.font = BODY if _c == 15 else BODYB
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c == 15), vertical="center")
        if rr % 2 == 0:
            cell.fill = ZEBRA
if _kp_missing:
    raise SystemExit(f"Keyword Plan rows with no brief: {_kp_missing[:5]}")
kw.auto_filter.ref = f"A3:P{3 + len(KEYWORDS)}"

_sc9 = wb["SEO Clusters"]
_sch = _sc9.cell(3, 6)
_sch.value = "Article Bank ID (the brief)"
_sch._style = copymod.copy(_sc9.cell(3, 5)._style)
_sc9.column_dimensions["F"].width = 18
from article_bank import SPOKES as _SPOKES
for _i, (_pil, _title) in enumerate(_SPOKES):
    rr = 4 + _i
    _sc9.cell(rr, 6).value = _AB_TITLE_ID.get(_title.lower(), "—")
    _sc9.cell(rr, 6).font = BODYB
    _sc9.cell(rr, 6).border = BOX
_scn = _sc9.cell(12, 1)
_scn.value = ("Pillar names here, on Keyword Plan column O and in Article Bank column C are "
              "the SAME seven — filter any of the three by the same word. Column F is the "
              "brief for the supporting article: find that ID on Article Bank, claim it with "
              "your name, and the prompt is ready to use.")
_scn.font = GREYNOTE
_scn.alignment = WRAP
_sc9.merge_cells(start_row=12, start_column=1, end_row=12, end_column=6)
_sc9.row_dimensions[12].height = 30

# ------------------------------------ 4i. PLATFORM GUIDE sheet
# The step-by-step usage manual for every platform: setup lives on Platform
# Setup column E; THIS is what you do every week once the account exists.
# Identity columns are formulas so the guide can never drift from the estate.
from usage_guides import USAGE
from canonical_lov import PLATFORMS as _PG_PLATFORMS

_missing = [p[0] for p in _PG_PLATFORMS if p[0] not in USAGE]
if _missing:
    raise SystemExit(f"usage_guides.py missing entries for: {_missing}")

pg = wb.create_sheet("PLATFORM GUIDE")
pg.sheet_view.showGridLines = False
title_row(pg, 9, "PLATFORM GUIDE  —  how to use every platform, step by step",
          "One row per platform, the whole estate. Account setup lives on Platform Setup "
          "(column E); this sheet is the WEEKLY PLAY once the account exists — what to do, "
          "how often, and the number that proves it is working. Identity columns update "
          "from Platform Setup automatically. Gold rows are the top-10 value platforms: "
          "if time is short, work this sheet top rank first.")
header_band(pg, 3, ["Platform", "Area", "Priority", "Value rank", "Strongest in (countries)",
                    "What it's for (logging rule)", "How to use it — step by step",
                    "KPI that proves it", "Time / week"],
            widths=[24, 12, 9, 8, 20, 30, 56, 22, 11])
_pg_notes = {p[0]: p[4] for p in _PG_PLATFORMS}
for _i, (_name, _area, _prio, _new, _note) in enumerate(_PG_PLATFORMS):
    rr = 4 + _i
    src = 4 + _i                     # Platform Setup rows map 1:1
    pg.cell(rr, 1).value = f"='Platform Setup'!$B${src}"
    pg.cell(rr, 2).value = f"='Platform Setup'!$C${src}"
    pg.cell(rr, 3).value = f"='Platform Setup'!$D${src}"
    pg.cell(rr, 4).value = f"='Platform Setup'!$P${src}"
    pg.cell(rr, 5).value = f"='Platform Setup'!$Q${src}"
    pg.cell(rr, 6).value = _note or "—"
    steps, kpi, tpw = USAGE[_name]
    pg.cell(rr, 7).value = steps
    pg.cell(rr, 8).value = kpi
    pg.cell(rr, 9).value = tpw
    for _c in range(1, 10):
        cell = pg.cell(rr, _c)
        cell.font = BODYB if _c == 1 else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (1, 5, 6, 7, 8)), vertical="top")
        if rr % 2 == 0:
            cell.fill = ZEBRA
    _need = max(len(steps) / 52, len(str(_note or "")) / 28, len(kpi) / 20, 2.0)
    pg.row_dimensions[rr].height = max(46, min(170, int(_need + 1) * 10 + 6))
_pg_last = 3 + len(_PG_PLATFORMS)
pg.conditional_formatting.add(
    f"A4:A{_pg_last}",
    _FR(formula=['AND(ISNUMBER($D4),$D4<=10)'], fill=PatternFill("solid", fgColor="FFE699")))
pg.conditional_formatting.add(
    f"C4:C{_pg_last}", _FR(formula=['$C4="Critical"'],
                           fill=PatternFill("solid", fgColor="F8CBAD")))
_pgn = pg.cell(_pg_last + 2, 1)
_pgn.value = ("Cadences are per active platform — nobody runs all of them at once. The "
              "manager assigns platforms by Value rank and the team's hours; anything "
              "assigned gets its weekly play logged in DAILY ENTRY. Deeper technique "
              "detail: GROWTH PLAYBOOK. Named routes and contacts: PR & Target Directory.")
_pgn.font = GREYNOTE
pg.merge_cells(start_row=_pg_last + 2, start_column=1, end_row=_pg_last + 2, end_column=9)
pg.row_dimensions[_pg_last + 2].height = 30
pg.freeze_panes = "B4"
pg.protection.sheet = True
pg.protection.selectLockedCells = False
pg.protection.selectUnlockedCells = False
pg.print_area = f"A1:I{_pg_last + 3}"
pg.print_title_rows = "1:3"
pg.sheet_properties.tabColor = "FF548235"

# ------------------------------------ 4j. Dashboard headline-KPI panel
# Live tiles in the free F:G columns beside sections 1-2 — the numbers a
# manager checks before anything else, visible on the first screen.
def _pulse(label):
    """Reference a Weekly Pulse KPI by NAME — never by a hardcoded row."""
    return f"='Weekly Pulse'!$B${_PULSE_ROW[label]}"

_KPIS = [
 ("REVENUE TO DATE", "=B95", '"$"#,##0'),
 ("REVENUE THIS WEEK", _pulse("Revenue recorded (USD)"), '"$"#,##0'),
 ("MEETINGS BOOKED", "=B101", "#,##0"),
 ("MEETINGS THIS WEEK", _pulse("Meetings booked"), "#,##0"),
 ("LEADS LOGGED", "=B26", "#,##0"),
 ("ACCEPTANCE RATE", "=B29", "0%"),
 ("FOLLOW-UPS DUE TODAY", "=B37", "#,##0"),
 ("MINUTES THIS WEEK", _pulse("Minutes logged"), "#,##0"),
 ("BACKLINKS THIS WEEK", _pulse("Backlinks gone live"), "#,##0"),
 ("CONTENT THIS WEEK", _pulse("Content published"), "#,##0"),
 ("PLATFORM COVERAGE", "=B21", "0%"),
 ("ACTIVE SCHEDULES", "=COUNTIFS('Content Scheduler'!$H$5:$H$103,\"<=\"&TODAY(),"
  "'Content Scheduler'!$I$5:$I$103,\">=\"&TODAY())", "#,##0"),
 # weighted by size — an unweighted mean let one tiny schedule at 0% drag a
 # hundred delivered posts down, and counted a broken row as a real 0%
 ("SCHEDULE COVERAGE", "=IFERROR(SUM('Content Scheduler'!$K$5:$K$103)"
  "/SUM('Content Scheduler'!$J$5:$J$103),\"\")", "0%"),
]
db8.column_dimensions["F"].width = 19
db8.column_dimensions["G"].width = 19
db8.merge_cells("F3:G3")
_kh8 = db8["F3"]
_kh8.value = "HEADLINE KPIs — live"
_kh8.font = H2; _kh8.fill = BANDFILL
_kh8.alignment = Alignment(horizontal="center", vertical="center")
db8.row_dimensions[3].height = 22
_kr = 4
_KPI_VALROW = {}                 # label -> row of its value cell (never hardcode)
for _label, _formula, _fmt in _KPIS:
    db8.merge_cells(start_row=_kr, start_column=6, end_row=_kr, end_column=7)
    lc = db8.cell(_kr, 6)
    lc.value = _label
    lc.font = Font(name="Arial", size=9, bold=True, color="5A6472")
    lc.alignment = Alignment(horizontal="center", vertical="center")
    db8.merge_cells(start_row=_kr + 1, start_column=6, end_row=_kr + 1, end_column=7)
    vc = db8.cell(_kr + 1, 6)
    vc.value = _formula
    vc.number_format = _fmt
    vc.font = Font(name="Arial", size=17, bold=True, color=NAVY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    _KPI_VALROW[_label] = _kr + 1
    for _rr in (_kr, _kr + 1):
        for _cc in (6, 7):
            db8.cell(_rr, _cc).border = BOX
            db8.cell(_rr, _cc).fill = PatternFill("solid", fgColor="F2F6FB")
    db8.row_dimensions[_kr].height = 15
    db8.row_dimensions[_kr + 1].height = 26
    _kr += 2
# the "work waiting" tile turns red when it is non-zero — addressed BY NAME so
# inserting a KPI above it can never repoint the rule at the wrong tile.
_fu = _KPI_VALROW["FOLLOW-UPS DUE TODAY"]
db8.conditional_formatting.add(
    f"F{_fu}:G{_fu}",
    _FR(formula=[f"$F${_fu}>0"], fill=PatternFill("solid", fgColor="F4CCCC")))

# ------------------------------------ 4k. Judge-2 content corrections
# Targeted text fixes on base sheets: misdirected pointers, stale wording,
# and instructions that contradicted the current design.
def _patch(ws, coord, old_frag, new_text):
    v = ws[coord].value
    if isinstance(v, str) and old_frag in v:
        ws[coord].value = new_text

# Same label, two sources, two numbers on ONE sheet destroys trust: §3 counts
# leads in the Outreach log, §4 sums what was logged in DAILY ENTRY.
for _r in range(24, 40):
    if str(db8.cell(_r, 1).value or "").strip() == "Connection requests sent":
        db8.cell(_r, 1).value = "Connection requests sent (leads in the Outreach log)"
        break
for _r in range(24, 40):
    if str(db8.cell(_r, 1).value or "").strip() == "Connections accepted":
        db8.cell(_r, 1).value = "Connections accepted (leads in the Outreach log)"
        break
for _r in range(40, 52):
    if str(db8.cell(_r, 1).value or "").strip() == "Connection requests sent":
        db8.cell(_r, 1).value = "Connection requests sent (logged in DAILY ENTRY)"
        break

# Day-one experience: a wall of zeros reads as "broken" unless the page says
# why. These lines appear only while the file is genuinely empty.
db8.merge_cells("A3:E3")
_ez = db8["A3"]
_ez.value = ('=IF(AND($B$41=0,$B$26=0),"FIRST RUN  —  every figure below is zero '
             'because nothing has been logged yet. They fill in automatically as the '
             'team uses DAILY ENTRY and the log tabs. Nothing is broken.","")')
_ez.font = Font(name="Arial", size=9, bold=True, color="9C5700")
_ez.fill = PatternFill("solid", fgColor="FFF2CC")
_ez.alignment = Alignment(vertical="center")
db8.row_dimensions[3].height = 18
wp.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
_ezp = wp.cell(4, 1)
_ezp.value = ('=IF($B$6=0,"No activity logged for this week yet  —  minutes and counts '
              'appear here as the team logs during the week.","")')
_ezp.font = Font(name="Arial", size=9, bold=True, color="9C5700")
_ezp.alignment = Alignment(vertical="center")
wp.row_dimensions[4].height = 16

_sm8 = wb["Summary"]
_patch(_sm8, "B38", "Daily Log", "Fill in DAILY ENTRY. Unlogged work does not count.")
_patch(_sm8, "C12", "Days the employee recorded activity",
       "Days on which anyone recorded activity.")
_lp8 = wb["LinkedIn Playbook"]
for _r in range(4, 20):
    v = _lp8.cell(_r, 5).value
    if isinstance(v, str):
        if "Daily Log, column K" in v:
            _lp8.cell(_r, 5).value = v.replace(
                "Daily Log, column K",
                "DAILY ENTRY — activity type 'Engagement (commented on someone else)'")
        elif "Daily Log - Blocker column" in v:
            _lp8.cell(_r, 5).value = v.replace(
                "Daily Log - Blocker column",
                "DAILY ENTRY — Notes or blocker column (column J)")
_ls8 = wb["Lists"]
_v = _ls8["A1"].value
if isinstance(_v, str) and "bottom of a column" in _v and "data-validation" not in _v:
    _ls8["A1"].value = (_v.rstrip() + "  NOTE: after adding a value, the manager must also "
                        "extend the matching data-validation range — dropdowns are capped "
                        "at the current last value and a value outside the range cannot "
                        "be selected anywhere.")
_htg8 = wb["How-To Guides"]
_patch(_htg8, "D5", "24h",
       "Set the page CTA, add a banner and pin a strong post. 3 page posts per week; "
       "reply to every comment within 60 minutes of posting (first-hour velocity gates "
       "reach) and clear stragglers within 24h.")
_htg8["D6"].value = ("One article per week: republish the site original via Medium's "
                     "Import tool (canonical set automatically) once the original is "
                     "indexed, or write a Medium-first piece if none is ready.")
_htg8["F6"].value = ("Republishing without the Import tool / canonical. Thin 300-word "
                     "posts. Missing tags.")
_patch(_htg8, "D15", "Approach 5 associations",
       "Approach 7 associations, 4 universities, 7 employers, 4 podcasts and 3 media "
       "contacts per week (25 total — matching the daily target of 5).")
_sh8 = wb["START HERE"]
_v = _sh8["B5"].value
if isinstance(_v, str) and "Six of the ten columns" in _v:
    _sh8["B5"].value = _v.replace(
        "Six of the ten columns are dropdowns",
        "Every row also carries the two tag columns (Objective + For (brand)); six of "
        "the twelve columns are dropdowns")
for _r in range(24, 33):
    if str(_sh8.cell(_r, 1).value or "").startswith("Community answers"):
        _sh8.cell(_r, 1).value = ("Community answers / comments "
                                  "(incl. engagement; 3-5 full answers/week)")
        _sh8.cell(_r, 1).alignment = Alignment(wrap_text=True, vertical="center")
        _sh8.row_dimensions[_r].height = 26
_patch(_sh8, "B103", "manager initials", "Done (Yes / No — manager confirms)")
_patch(wb["Daily Log"], "A2", "one quarter at 10 people",
       "This tab is optional and auto-calculated. It holds 120 formula rows — the "
       "manager copies the last row down when more are needed. The tab everyone "
       "fills is DAILY ENTRY.")
_ex8 = wb["Experiments"]
_v = _ex8["A2"].value
if isinstance(_v, str) and "significance" in _v:
    _ex8["A2"].value = _v.replace("Rates and significance are calculated",
                                  "Rates and a size guard are calculated")
_wr8 = wb["Weekly Review"]
_v = _wr8["D3"].value
if isinstance(_v, str) and "(daily log)" in _v.lower():
    _wr8["D3"].value = "Leads researched (DAILY ENTRY)"
_utm8 = wb["UTM Builder"]
_v = _utm8["A4"].value
if isinstance(_v, str) and "example" in _v.lower():
    _utm8["A4"].value = "https://projectcontrolsinstitute.org/honorary-certification"

# ------------------------------------ 4l. Judge-3 formatting & UX fixes
GREYEX = PatternFill("solid", fgColor="E7E6E6")
# UTM example row gets the standard example-grey (inputs only)
for _c in range(1, 6):
    _utm8.cell(4, _c).fill = GREYEX
# Summary genuinely fits one page
wb["Summary"].page_setup.fitToHeight = 1
# Dashboard: hide the chart-data helper block in plain sight (charts still read it)
for _r in range(2, 19):
    for _c in (8, 9):
        db8.cell(_r, _c).font = Font(name="Arial", size=8, color="FFFFFF")
# charts move right so they never cover the KPI panel or helper cells
for _ch in db8._charts:
    try:
        _ch.anchor._from.col += 4
        _ch.anchor.to.col += 4
    except AttributeError:
        pass
# money formats on §8 (match Weekly Pulse)
for _coord in ("B95", "B98", "B99", "B100"):
    db8[_coord].number_format = '"$"#,##0'
# leftover bordered empty rows under §6
_noborder = Border()
for _r in (78, 80):
    for _c in range(1, 6):
        if db8.cell(_r, _c).value is None:
            db8.cell(_r, _c).border = _noborder
# Channel Costs: bordered table + money format
_cc8 = wb["Channel Costs"]
for _r in range(3, 24):
    for _c in range(1, 5):
        _cc8.cell(_r, _c).border = BOX
for _r in range(4, 24):
    _cc8.cell(_r, 2).number_format = '"$"#,##0'
# QA & Compliance: the live-signal column joins the table frame
_qa8 = wb["QA & Compliance"]
_qa8.column_dimensions["I"].width = 38
for _r in range(4, 19):
    c = _qa8.cell(_r, 9)
    c.border = BOX
    c.font = BODY
    c.alignment = WRAP
# LinkedIn Outreach example message fully visible
wb["LinkedIn Outreach"].row_dimensions[4].height = 66
# print hygiene: repeated headers + helper block off-print
wb["PR & Target Directory"].print_title_rows = "1:4"
wb["How-To Guides"].print_title_rows = "1:3"
wb["Weekly Review"].print_area = "A1:Q36"
# Daily Log banner matches the tab name
_dl8 = wb["Daily Log"]
_v = _dl8["A1"].value
if isinstance(_v, str) and "DAILY SUMMARY" in _v:
    _dl8["A1"].value = _v.replace("DAILY SUMMARY", "DAILY LOG")
# Message Bank approval note becomes a bounded, wrapped cell
_mb8 = wb["Message Bank"]
if isinstance(_mb8["G2"].value, str):
    _mb8.merge_cells("G2:I2")
    _mb8["G2"].font = GREYNOTE
    _mb8["G2"].alignment = WRAP
    _mb8["G2"].border = BOX
# Lists title band spans the data
if "A1:O1" in [str(m) for m in wb["Lists"].merged_cells.ranges]:
    wb["Lists"].unmerge_cells("A1:O1")
    wb["Lists"].merge_cells("A1:Y1")
# START HERE: §9/§10/§11 get the same band treatment as §1-§8;
# §10 title no longer collides with the Done header; §7 note wraps fully
_sh8["A103"] = "10.  DAY-1 CHECKLIST  —  every new team member"
for _r in (93, 103, 115):
    for _c in (1, 2, 3):
        cell = _sh8.cell(_r, _c)
        cell.fill = BANDFILL
        cell.font = Font(name="Arial", size=10 if _c > 1 else 11, bold=True, color="FFFFFF")
_sh8.merge_cells("A70:C70")
_sh8["A70"].alignment = WRAP
_sh8.row_dimensions[70].height = 44
# Judge 10: START HERE never named the navigation or the guides — a new joiner
# had no stated path. Row 3 becomes the signpost.
_sh8["A3"] = ("NEW HERE?  Read this page, then TEAM GUIDE (where to log what), then "
              "PLATFORM GUIDE (how to use each platform).  ·  Lost? The MAP tab lists "
              "every sheet with one click.  ·  Need something to write? Article Bank.")
_sh8.merge_cells("A3:F3")
_sh8["A3"].font = Font(name="Arial", size=9, bold=True, color=NAVY)
_sh8["A3"].fill = PatternFill("solid", fgColor="FFF2CC")
_sh8["A3"].alignment = Alignment(vertical="center", wrap_text=True)
_sh8["A3"].border = BOX
_sh8.row_dimensions[3].height = 26

# ------------------------------------ 4m. Numerical-oracle corrections
# An independent numerical audit injected a four-person, four-week dataset,
# recalculated, and re-derived every aggregate from the raw rows in Python.
# The arithmetic was sound; what was being COUNTED was not. Everything below
# is a definition fix, and every one of them changed a number a manager acts
# on. Rows are found by their label so this block can never drift.
_LO_R = "'LinkedIn Outreach'!$R$5:$R$1203"
_LO_M = "'LinkedIn Outreach'!$M$5:$M$1203"
_LO_N = "'LinkedIn Outreach'!$N$5:$N$1203"


def _dbrow(label):
    for _r in range(1, 130):
        if str(db8.cell(_r, 1).value or "").strip() == label:
            return _r
    raise SystemExit(f"Dashboard row not found: {label!r}")


# --- rates could exceed 100%, and under-logging RAISED a person's score.
# Marking a lead accepted without having logged the request pushed acceptance
# to 123% and handed +32.5 points on Employee Score — the sheet rewarded the
# commonest logging slip. The denominator now cannot be smaller than the
# numerator, and the count of contradictory rows is reported.
_r29 = _dbrow("Acceptance rate")
db8.cell(_r29, 2).value = (f'=IFERROR(COUNTIF({_LO_N},"Yes")'
                           f'/MAX(COUNTIF({_LO_M},"Yes"),COUNTIF({_LO_N},"Yes")),0)')
_r32 = _dbrow("Reply rate on messages sent")
_pos = (f'COUNTIF({_LO_S},"Interested")+COUNTIF({_LO_S},"Info Requested")'
        f'+COUNTIF({_LO_S},"Meeting Booked")+COUNTIF({_LO_S},"Converted")')
db8.cell(_r32, 2).value = f'=IFERROR(({_pos})/MAX(COUNTIF({_LO_R},"Yes"),{_pos}),0)'

# --- a meeting was counted from the Outcome dropdown, which is single-valued:
# the moment a meeting progressed to "Application Started" or "Converted" it
# was erased from the cumulative count. Five of eight test meetings vanished,
# and the cumulative tile read LOWER than the weekly one.
_r33 = _dbrow("Meetings / calls booked")
db8.cell(_r33, 1).value = "Currently at meeting-booked stage"
_r101 = _dbrow("Meetings booked to date")
db8.cell(_r101, 2).value = (f'=COUNTIF({_LO_AG2},">0")+COUNTIF({_CPR_P2},">0")')
db8.cell(_r101, 1).value = "Meetings booked to date (by meeting date)"

# --- work waiting today must include the leads who replied and are owed the
# next thing, not only the ones who went quiet
_r37 = _dbrow("Follow-ups due today or earlier")
db8.cell(_r37, 2).value = "+".join(
    f'COUNTIFS(\'LinkedIn Outreach\'!$U$5:$U$1203,"<="&TODAY(),'
    f'\'LinkedIn Outreach\'!$U$5:$U$1203,">0",{_LO_S},"{_o}")'
    for _o in ("Awaiting Reply", "", "No Response", "Interested", "Info Requested"))
db8.cell(_r37, 2).value = "=" + db8.cell(_r37, 2).value
db8.cell(_r37, 1).value = "Follow-ups and replies due today or earlier"

# --- "Expected to date" charged every person for every day ANYBODY worked:
# headcount x distinct calendar days, not person-days. Attainment was
# understated 2.9x in the test, so every "Behind target" verdict was produced
# against an inflated bar. DAILY ENTRY column L already holds person-days.
_PD = "SUM('DAILY ENTRY'!$L$7:$L$1006)"
for _r in range(42, 50):
    db8.cell(_r, 3).value = f"='START HERE'!$B${_r - 17}*{_PD}"
db8.cell(40, 3).value = "Expected to date (target × person-days)"

# --- three different "community" numbers shared one word; name the units
db8.cell(_dbrow("Community answers / comments"), 1).value = \
    "Community answers / comments (items counted in DAILY ENTRY)"
db8.cell(_dbrow("Community and PR actions logged"), 1).value = \
    "Community & PR rows logged (one row per action)"

# --- a typed note in the contract-signed cell was booking revenue. "<>"
# accepts any text; ">0" accepts only a real date, so signed value and open
# pipeline can never both claim the same deal — or both disown it.
_r98 = _dbrow("Enterprise pipeline value (open)")
_r99 = _dbrow("Enterprise value signed")
db8.cell(_r99, 2).value = (f"=SUMIFS({_PP_V},{_PP_W},\">0\")")
db8.cell(_r98, 2).value = f"=SUM({_PP_V})-B{_r99}"
db8.cell(_r98, 1).value = "Enterprise pipeline value (not yet signed)"
_r95 = _dbrow("Certification revenue recorded")
db8.cell(_r95, 1).value = "Certification revenue recorded (component)"
# one revenue definition, stated once, and the KPI tile reads it. The old
# footnote sat on this row; it moves down under the new total.
_rtot = _r101 + 1
for _mr in [str(m) for m in list(db8.merged_cells.ranges)
            if m.min_row in (_rtot, _rtot + 1)]:
    db8.unmerge_cells(_mr)
for _c in range(1, 8):
    db8.cell(_rtot, _c).value = None
    db8.cell(_rtot + 1, _c).value = None
db8.cell(_rtot, 1).value = "TOTAL REVENUE RECORDED  (certifications + signed enterprise)"
db8.cell(_rtot, 2).value = f"=B{_r95}+B{_r99}"
db8.cell(_rtot, 1)._style = copymod.copy(db8.cell(_r95, 1)._style)
db8.cell(_rtot, 2)._style = copymod.copy(db8.cell(_r95, 2)._style)
db8.cell(_rtot, 1).font = BODYB
db8.cell(_rtot, 2).font = Font(name="Arial", size=9, bold=True, color=NAVY)
db8.cell(_rtot, 2).number_format = '"$"#,##0'
for _c in (1, 2):
    db8.cell(_rtot, _c).fill = PatternFill("solid", fgColor="DDEBF7")
_dbnote = db8.cell(_rtot + 1, 1)
_dbnote.value = ("Revenue figures reconcile against PCI platform order references — a "
                 "Converted row without a PCI ref is unverified. 'Recorded' means a real "
                 "purchase date on Outreach or a real contract-signed date on Partnerships; "
                 "a typed note in either cell is not a sale.")
_dbnote.font = GREYNOTE
_dbnote.alignment = WRAP
db8.merge_cells(start_row=_rtot + 1, start_column=1, end_row=_rtot + 1, end_column=5)
db8.row_dimensions[_rtot + 1].height = 26

# --- SECTION 9: DATA HEALTH. Six numbers that must all read zero. Every one
# of them corresponds to a defect the audit found by hand; the workbook now
# finds them itself. A wrong-year typo alone moved cumulative leads from 195
# to 1,194 with nothing anywhere to flag it.
_dh = _rtot + 3
section(db8, _dh, 5, "9.  DATA HEALTH  —  every number here should read 0")
for _c, _h in enumerate(["Check", "Count", "Why it matters", "", ""], 1):
    hc = db8.cell(_dh + 1, _c)
    hc.value = _h or None
    hc.font = H2; hc.fill = BANDFILL; hc.border = BOX
    hc.alignment = Alignment(vertical="center")
db8.merge_cells(start_row=_dh + 1, start_column=3, end_row=_dh + 1, end_column=5)
db8.row_dimensions[_dh + 1].height = 18
_HEALTH = [
 ("Rows with an unreadable (text) date",
  "=COUNTA('DAILY ENTRY'!$A$7:$A$1006)-COUNT('DAILY ENTRY'!$A$7:$A$1006)"
  "+COUNTA('Content Calendar'!$K$5:$K$403)-COUNT('Content Calendar'!$K$5:$K$403)",
  "A date typed as text lands in the cumulative totals and in no week — the two "
  "can then never be reconciled."),
 ("Future-dated rows",
  "=COUNTIF('DAILY ENTRY'!$A$7:$A$1006,\">\"&TODAY())"
  "+COUNTIF('LinkedIn Outreach'!$A$5:$A$1203,\">\"&TODAY())",
  "One wrong-year typo multiplies every cumulative figure and the 'days since' "
  "ages go negative."),
 ("Accepted without a logged connection request",
  f'=COUNTIFS({_LO_N},"Yes",{_LO_M},"<>Yes")',
  "Acceptance and reply rates are only trustworthy while every accepted lead has "
  "its request logged."),
 ("Minutes logged under a name not on the roster",
  "=COUNTIFS('DAILY ENTRY'!$B$7:$B$1006,\"<>\")"
  "-SUMPRODUCT(COUNTIFS('DAILY ENTRY'!$B$7:$B$1006,'START HERE'!$B$55:$B$64))",
  "Those minutes count for the team and for nobody — the per-person block cannot "
  "add up to the team row."),
 ("Content marked Published with no published date",
  "=COUNTIFS('Content Calendar'!$J$5:$J$403,\"Published\",'Content Calendar'!$K$5:$K$403,\"\")"
  "+COUNTIFS('Content Calendar'!$J$5:$J$403,\"Repurposed\",'Content Calendar'!$K$5:$K$403,\"\")",
  "Counted by the Dashboard, invisible to Weekly Pulse and to every schedule's "
  "coverage figure."),
 ("Published with a date but not marked published",
  "=SUMPRODUCT(ISNUMBER('Content Calendar'!$K$5:$K$403)"
  "*('Content Calendar'!$J$5:$J$403<>\"Published\")"
  "*('Content Calendar'!$J$5:$J$403<>\"Repurposed\"))",
  "The mirror error — counted by Weekly Pulse, invisible to the Dashboard."),
 ("Signed deals with no deal value",
  f'=COUNTIFS({_PP_W},">0",{_PP_V},"")',
  "A signed deal worth nothing is in neither open pipeline nor booked revenue."),
 ("Duplicate leads flagged in the Outreach log",
  "=COUNTIF('LinkedIn Outreach'!$AN$5:$AN$1203,\"DUPLICATE*\")",
  "The same person counted twice inflates every funnel rate."),
 ("Published content with no platform named",
  "=COUNTIFS('Content Calendar'!$J$5:$J$403,\"Published\",'Content Calendar'!$C$5:$C$403,\"\")"
  "+COUNTIFS('Content Calendar'!$J$5:$J$403,\"Repurposed\",'Content Calendar'!$C$5:$C$403,\"\")",
  "It counts as published but belongs to no schedule, so it silently depresses "
  "every coverage figure."),
 ("Impossible (negative) values — investigate",
  "=COUNTIF('DAILY ENTRY'!$G$4:$G$1006,\"<0\")"
  "+COUNTIF('DAILY ENTRY'!$F$4:$F$1006,\"<0\")"
  "+COUNTIF('LinkedIn Outreach'!$AL$4:$AL$1203,\"<0\")"
  "+COUNTIF('Partnership Pipeline'!$V$4:$V$403,\"<0\")",
  "Pasting bypasses every dropdown and range check in Excel. A negative minute "
  "or a negative fee turns a headline KPI negative with nothing else to show "
  "for it."),
]
_hr = _dh + 2
_row_style = copymod.copy(db8.cell(_r95, 1)._style)
for _lab, _f, _why in _HEALTH:
    db8.cell(_hr, 1).value = _lab
    db8.cell(_hr, 2).value = _f
    db8.merge_cells(start_row=_hr, start_column=3, end_row=_hr, end_column=5)
    db8.cell(_hr, 3).value = _why
    for _c in range(1, 6):
        cell = db8.cell(_hr, _c)
        cell.font = BODY if _c != 2 else BODYB
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c in (1, 3)), vertical="center")
        if _hr % 2 == 0:
            cell.fill = ZEBRA
    db8.cell(_hr, 2).number_format = "#,##0"
    db8.cell(_hr, 2).alignment = Alignment(horizontal="center", vertical="center")
    db8.cell(_hr, 3).font = GREYNOTE
    db8.row_dimensions[_hr].height = 26
    _hr += 1
db8.conditional_formatting.add(
    f"A{_dh + 2}:E{_hr - 1}",
    _FR(formula=[f"$B{_dh + 2}>0"], fill=PatternFill("solid", fgColor="F4CCCC")))
db8.conditional_formatting.add(
    f"B{_dh + 2}:B{_hr - 1}",
    _FR(formula=[f"$B{_dh + 2}=0"], fill=PatternFill("solid", fgColor="C6EFCE")))
_dhn = db8.cell(_hr, 1)
_dhn.value = ("Anything red here is a logging problem, not a formula problem — fix it on "
              "the source sheet and this section goes green on its own. Check it before "
              "quoting any number on this page to anyone outside the team.")
_dhn.font = GREYNOTE
_dhn.alignment = WRAP
db8.merge_cells(start_row=_hr, start_column=1, end_row=_hr, end_column=5)
db8.row_dimensions[_hr].height = 26
_DB_LAST = _hr

# --- the KPI tile now reads the one revenue definition
for _rr in range(4, 40):
    if str(db8.cell(_rr, 6).value or "") == "REVENUE TO DATE":
        db8.cell(_rr + 1, 6).value = f"=B{_rtot}"
        break

# --- Weekly Review lost every Sunday afternoon. Week end is a bare date, so
# "<=$C4" is Sunday 00:00 — anything logged with a time of day on the last day
# of the week fell outside the week it belonged to, while Weekly Pulse (which
# uses a correct half-open window) counted it. Two pages, same week, different
# answers.
_wr9 = wb["Weekly Review"]
for _r in range(4, 30):
    for _c in range(4, 23):
        _v = _wr9.cell(_r, _c).value
        if isinstance(_v, str) and f'"<="&$C{_r}' in _v:
            _wr9.cell(_r, _c).value = _v.replace(f'"<="&$C{_r}', f'"<"&$C{_r}+1')
# --- and it could never show work done before the programme start date, with
# no total row to reveal the gap. Both are now on the page.
_wr9.cell(30, 1).value = "Before programme start"
_wr9.cell(30, 2).value = "='START HERE'!$B$20-3650"
_wr9.cell(30, 3).value = "='START HERE'!$B$20-1"
for _c in range(4, 23):
    _src = _wr9.cell(4, _c).value
    if isinstance(_src, str) and _src.startswith("="):
        _wr9.cell(30, _c).value = _src.replace("$B4", "$B30").replace("$C4", "$C30")
        _wr9.cell(30, _c)._style = copymod.copy(_wr9.cell(4, _c)._style)
_wr_bench = _wr9.cell(31, 1).value          # the benchmarks note lives here
_wr_bstyle = copymod.copy(_wr9.cell(31, 1)._style)
for _mr in [str(m) for m in list(_wr9.merged_cells.ranges) if m.min_row == 31]:
    _wr9.unmerge_cells(_mr)
for _c in range(1, 23):
    _wr9.cell(31, _c).value = None
_wr9.cell(31, 1).value = "TOTAL (all weeks + before start)"
for _c in range(4, 23):
    if _wr9.cell(4, _c).value:
        _L9 = get_column_letter(_c)
        _wr9.cell(31, _c).value = f"=SUM({_L9}4:{_L9}30)"
        _wr9.cell(31, _c)._style = copymod.copy(_wr9.cell(4, _c)._style)
        _wr9.cell(31, _c).font = BODYB
for _r in (30, 31):
    _wr9.cell(_r, 1)._style = copymod.copy(_wr9.cell(4, 1)._style)
    _wr9.cell(_r, 1).font = BODYB
    for _c in (2, 3):
        _wr9.cell(_r, _c)._style = copymod.copy(_wr9.cell(4, _c)._style)
    for _c in range(1, 23):
        _wr9.cell(_r, _c).fill = PatternFill("solid", fgColor="DDEBF7")
for _c in (7, 10):                                  # the two % columns
    _wr9.cell(31, _c).value = None
_wrn = _wr9.cell(33, 1)
_wrn.value = ("Weeks run from the programme start date on START HERE §2. The 'Before "
              "programme start' row catches anything logged earlier so the TOTAL always "
              "reconciles with the Dashboard's cumulative figures. The person selector "
              "(top right) filters every column except the three all-team columns; '*' "
              "means everyone whose name was logged — rows logged with no name are not "
              "included, and the Dashboard's DATA HEALTH section counts them.")
_wrn.font = GREYNOTE
_wrn.alignment = WRAP
_wr9.merge_cells(start_row=33, start_column=1, end_row=33, end_column=17)
_wr9.row_dimensions[33].height = 32
if _wr_bench:                                # benchmarks note moves down
    _wb35 = _wr9.cell(35, 1)
    _wb35.value = _wr_bench
    _wb35._style = _wr_bstyle
    _wb35.alignment = WRAP
    _wr9.merge_cells(start_row=35, start_column=1, end_row=35, end_column=17)
    _wr9.row_dimensions[35].height = 30

# --- Employee Score spent 30 of its 100 points on ONE ratio: "Reply rate
# (20)" and "Positive replies (10)" were the same number divided the same way,
# so they saturated and zeroed in lockstep. The 10 points now measure output
# (positive replies as a count against a target), which is a different thing.
_es9 = wb["Employee Score"]
_es9.cell(3, 5).value = "Positive replies\n(10)\ncount vs target 10"
for _r in range(4, 14):
    _es9.cell(_r, 5).value = (f'=IF($A{_r}="","",'
                              f"ROUND(MIN(1,'Team Scorecard'!H{_r}/10)*10,1))")
    # a TOTAL printed beside "Too little data to score fairly yet" reads as a
    # verdict. It is not one until the gate opens.
    _es9.cell(_r, 9).value = (f'=IF($A{_r}="","",IF($K{_r}<>"Enough data","",'
                              f'ROUND(SUM(B{_r}:H{_r}),1)))')

# --- Team Scorecard rates had the same containment hole as the Dashboard:
# 333% acceptance and 400% reply rate for one person in the test data.
_ts9 = wb["Team Scorecard"]
for _r in range(4, 14):
    _ts9.cell(_r, 6).value = (f'=IF($A{_r}="","",IFERROR(E{_r}/MAX(D{_r},E{_r}),""))')
    _ts9.cell(_r, 9).value = (f'=IF($A{_r}="","",IFERROR(H{_r}/MAX(G{_r},H{_r}),""))')

# --- Platform Progress: a future-dated row produced a NEGATIVE age that read
# as freshly worked, and the "who worked on it last" column names whichever
# matching row comes first in sheet order, not the last person.
# --- dates must be dates. A date typed as text ("2026-08-18") is counted by
# every cumulative total and by no weekly window, so the two can never be
# reconciled; a wrong-year typo multiplies the cumulative figures outright.
# Columns that legitimately hold future dates (due dates, plan dates, schedule
# windows) only get the is-it-a-real-date test.
def _drop_date_dv(sheet, col):
    """Remove any older date rule on this column so exactly one rule applies."""
    ws_ = wb[sheet]
    keep = []
    for d in ws_.data_validations.dataValidation:
        if d.type == "date" and any(
                re.sub(r"\d+", "", p) in (col, f"{col}:{col}")
                for part in str(d.sqref).split() for p in part.split(":")):
            continue
        keep.append(d)
    ws_.data_validations.dataValidation = keep


_PAST_DATES = [("DAILY ENTRY", "A", 4, 1006), ("LinkedIn Outreach", "A", 5, 1203),
               ("LinkedIn Outreach", "AD", 5, 1203), ("LinkedIn Outreach", "AE", 5, 1203),
               ("LinkedIn Outreach", "AK", 5, 1203), ("Content Calendar", "K", 5, 403),
               ("Community & PR", "A", 5, 403), ("Daily Log", "A", 4, 403),
               ("Job Postings", "A", 5, 203), ("Link Building", "K", 5, 403),
               ("Partnership Pipeline", "A", 5, 403), ("Partnership Pipeline", "W", 5, 403)]
_ANY_DATES = [("LinkedIn Outreach", "AG", 5, 1203), ("LinkedIn Outreach", "AI", 5, 1203),
              ("Community & PR", "O", 5, 403), ("Community & PR", "P", 5, 403),
              ("Content Calendar", "A", 5, 403), ("Master Tasks", "L", 4, 66),
              ("Partnership Pipeline", "Q", 5, 403), ("Partnership Pipeline", "S", 5, 403)]
for _sn, _col, _r1, _r2 in _PAST_DATES:
    if _sn not in wb.sheetnames:
        continue
    _drop_date_dv(_sn, _col)
    _d = DataValidation(type="date", operator="lessThanOrEqual", formula1="TODAY()",
                        allow_blank=True)
    _d.error = ("Enter a real date, not text, and not in the future. A date typed as text "
                "lands in the running totals but in no week.")
    _d.errorTitle = "That is not a usable date"
    _d.showErrorMessage = True
    wb[_sn].add_data_validation(_d)
    _d.add(f"{_col}{_r1}:{_col}{_r2}")
for _sn, _col, _r1, _r2 in _ANY_DATES:
    if _sn not in wb.sheetnames:
        continue
    _drop_date_dv(_sn, _col)
    _d = DataValidation(type="date", operator="greaterThan", formula1="36526",
                        allow_blank=True)
    _d.error = "Enter a real date (not text) — a text date is invisible to every weekly view."
    _d.errorTitle = "That is not a usable date"
    _d.showErrorMessage = True
    wb[_sn].add_data_validation(_d)
    _d.add(f"{_col}{_r1}:{_col}{_r2}")

_pp9 = wb["Platform Progress"]
_pp9.cell(3, 9).value = "Who last logged on it (one of)"
for _r in range(4, 137):
    _v = _pp9.cell(_r, 11).value
    if isinstance(_v, str) and _v.startswith("="):
        _pp9.cell(_r, 11).value = _v.replace(
            "=IF(", f'=IF($H{_r}<0,"Future-dated entry — check the date on DAILY ENTRY",IF(',
            1) + ")"

# ------------------------------------ 4n. Day-one calm + channel economics
# A workbook that opens as a wall of red teaches the team to ignore red. On a
# fresh file nothing has been done yet, so nothing is behind yet — the warning
# colours only earn their place once there is something to warn about.
_ps8 = wb["Platform Setup"]
for _r in range(4, 137):
    _ps8.cell(_r, 9).value = None            # 2FA is unknown, not "No"
for _rng in [k for k in list(_ps8.conditional_formatting._cf_rules)
             if str(k.sqref).startswith("I4:")]:
    del _ps8.conditional_formatting._cf_rules[_rng]
_ps8.conditional_formatting.add(
    "I4:I136",
    _FR(formula=['AND($I4="No",$J4<>"Not Started",$J4<>"")'],
        fill=PatternFill("solid", fgColor="F4CCCC")))
_ps8.cell(3, 9).value = "2FA on?"
# Platform Progress read "Setup in progress" for 133 platforms nobody had
# touched — the status of an untouched estate is "not started".
for _r in range(4, 137):
    _v = _pp9.cell(_r, 11).value
    if isinstance(_v, str) and _v.startswith("="):
        _pp9.cell(_r, 11).value = _v.replace(
            f'IF($C{_r}="Not Started","Not set up yet","Setup in progress")',
            f'IF(OR($C{_r}="Not Started",$C{_r}=0,$C{_r}=""),"Not set up yet",'
            f'"Setup in progress")')

# The Dashboard kept per-row "below 80% = red" rules from an earlier version
# alongside the guarded ones, so every attainment cell opened red before a
# single day had been logged. The guarded rules (which require a target to
# exist) are the ones that stay.
for _rng in [k for k in list(db8.conditional_formatting._cf_rules)
             if re.fullmatch(r"D4[2-9]", str(k.sqref))]:
    del db8.conditional_formatting._cf_rules[_rng]

# Master Tasks' brand column arrived without the styling every other column
# has, so it read as a hole in the table.
_mt8 = wb["Master Tasks"]
for _r in range(4, 67):
    _c8 = _mt8.cell(_r, 16)
    _c8._style = copymod.copy(_mt8.cell(_r, 15)._style)
    _c8.fill = YELLOW
    _c8.protection = Protection(locked=False)
_mt8.column_dimensions["P"].width = 26
_mt8.column_dimensions["O"].width = 30

# --- Channel Costs promised cost-per-outcome and never computed it.
_cc8 = wb["Channel Costs"]
_cc8.cell(15, 1).value = "CHANNEL ECONOMICS  (computed — nothing to type here)"
_cc8.cell(15, 1).font = H2
_cc8.cell(15, 1).fill = BANDFILL
_cc8.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)
_cc8.row_dimensions[15].height = 20
_ECON = [
 ("Total monthly channel cost", "=SUM($B$4:$B$13)", '"$"#,##0',
  "Everything the programme spends in a month."),
 ("Meetings booked to date", f"='Dashboard'!$B${_r101}", "#,##0",
  "Counted from real meeting dates on Outreach and Community & PR."),
 ("Cost per meeting (one month of cost)", f'=IF(Dashboard!$B${_r101}=0,"",'
  f"SUM($B$4:$B$13)/Dashboard!$B${_r101})", '"$"#,##0.00',
  "One month of spend divided by every meeting booked so far — it falls as the "
  "programme matures. Compare months, not the absolute figure."),
 ("Total revenue recorded", f"='Dashboard'!$B${_rtot}", '"$"#,##0',
  "Certification revenue plus signed enterprise value."),
 ("Revenue per $1 of monthly cost", f'=IF(SUM($B$4:$B$13)=0,"",'
  f"Dashboard!$B${_rtot}/SUM($B$4:$B$13))", "0.0",
  "Above 1.0 means one month of channel cost is already covered."),
]
_er = 16
for _lab, _f, _fmt, _why in _ECON:
    _cc8.cell(_er, 1).value = _lab
    _cc8.cell(_er, 2).value = _f
    _cc8.cell(_er, 4).value = _why
    for _c in range(1, 5):
        cell = _cc8.cell(_er, _c)
        cell.font = BODYB if _c == 2 else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=(_c == 4), vertical="center")
        if _er % 2 == 0:
            cell.fill = ZEBRA
    _cc8.cell(_er, 2).number_format = _fmt
    _cc8.cell(_er, 4).font = GREYNOTE
    _cc8.row_dimensions[_er].height = 26
    _er += 1
_cc8.column_dimensions["A"].width = 34
_cc8.column_dimensions["D"].width = 52
_cc8.print_area = f"A1:D{_er}"

# --- the Article Bank prompts name the .org. A competitor training provider
# owns the .com, and a writer who links to it hands them the traffic.
ab.cell(2, 1).value = (str(ab.cell(2, 1).value or "") +
                       "  ·  Every link in a published article goes to "
                       "projectcontrolsinstitute.ORG — the .com belongs to a competitor "
                       "training provider (see Keyword Plan).")

# ------------------------------------------------ 5. UPGRADE NOTES redesign
un = wb["V7 UPGRADE NOTES"]
un.title = "UPGRADE NOTES"
un.cell(4, 5).value = "What this workbook does"
for _r in range(5, un.max_row + 1):
    for _c in (5,):
        v = un.cell(_r, _c).value
        if isinstance(v, str) and "V7 " in v:
            un.cell(_r, _c).value = v.replace("V7 adds", "This version adds").replace("V7 cannot", "The workbook cannot")
# superseded/stale V7-era rows corrected in place (Judge-2 findings 2 and 9)
for _r in range(5, un.max_row + 1):
    v = un.cell(_r, 5).value
    if isinstance(v, str) and "Only the Content Calendar has a Brand column" in v:
        un.cell(_r, 5).value = ("SUPERSEDED by the brand dimension: every log tab now "
                                "carries a For (brand) column, so Certuvo work is logged "
                                "in this file and tagged 'Certuvo (exam prep)' — results "
                                "split cleanly on Objective Performance.")
        un.cell(_r, 6).value = ("Tag Certuvo rows with the brand dropdown; no separate "
                                "copy of the file is needed.")
    if isinstance(v, str) and "one canonical 71-platform list" in v:
        un.cell(_r, 5).value = v.replace(
            "All lists now come from one canonical 71-platform list",
            "All lists come from one canonical platform list (71 at that round; "
            "the estate has since grown — see later entries)")
# V8 rows appended before the restyle so they inherit the table formatting
V8_NOTES = [
 ("Single-employee residue removed", "Process", "Major",
  "START HERE section 2 asked for one 'Employee name' — a leftover from a single-user draft that made the whole file read as one person's tracker. It is now Team Settings (team name + manager); people exist only on the roster, targets are explicitly per person, and the Dashboard multiplies by headcount.",
  "Manager fills Team Settings and the roster; nothing else defines who is on the team."),
 ("Platform coverage doubled on live research", "Platform", "Major",
  "Web research (August 2026, 26 searches) added 29 vetted organic-growth platforms — the ChatGPT-visibility pipeline (Bing Webmaster Tools + IndexNow), AI answer-engine auditing, PM World Journal, Project Controls Expo, the post-HARO journalist stack, UK/Gulf course directories, CPD accreditors, podcast booking tools and more — bringing the estate at that round to 100 platforms across 12 areas (later rounds grew it to 133 across 13), each with setup steps.",
  "Work Platform Setup by priority; the named routes live on PR & Target Directory."),
 ("Dead tactics documented so nobody wastes time", "Process", "Minor",
  "The research also verified what NOT to do: HARO is dead, article directories and bulk PDF submission are link spam, llms.txt is evidence-negative, engagement pods and automated LinkedIn tools get accounts penalised. GROWTH PLAYBOOK technique 15 carries the full list.",
  "Read playbook technique 15 before acting on any 'SEO tips' list from the internet. Note: the original HARO shut Dec 2024; the brand was relaunched under Featured.com in 2025 — treat it as part of the Featured stack."),
 ("GROWTH PLAYBOOK added — now 22 techniques in working detail", "Process", "Major",
  "Numbered steps, cadence, KPI and pitfalls for every technique: SEO clusters, E-E-A-T, answer-engine optimisation, digital PR, the 2026 LinkedIn engine, community growth, syndication with verified canonical rules (Substack and LinkedIn have none), webinars, email, the Credly badge loop, directories, YouTube Shorts-first, measurement, the dead-tactics guardrail — plus, from later rounds: entity authority (14), the free publishing & news pipeline (16), and six world-class additions: original research & data studies, free tools & calculators, Arabic/Gulf localisation, internal linking architecture, employee advocacy & social proof, and guarded programmatic geo×role pages (17-22).",
  "Each technique names the sheet it logs into; the manager assigns owners per technique."),
 ("TEAM GUIDE added — the end-to-end explainer", "Process", "Major",
  "A team that does not know how to run the system now has the answer in one tab: where to log what, first day, every day, how scoring works, and an FAQ.",
  "Every new joiner reads TEAM GUIDE before anything else (Day-1 Checklist item 1)."),
 ("Accounts Register added — and why it holds no passwords", "Process", "Major",
  "Every marketing account on one page: URL, login, vault entry name, 2FA, owner, status — built by formula from Platform Setup so there is one source of truth. Passwords stay in the team password manager: a shared, synced, version-historied workbook can never hold credentials safely, and the register names the exact vault entry instead.",
  "Manager fills the vault location cell and grants vault access on day 1."),
 ("PR & Target Directory added — named, verified routes", "Platform", "Major",
  "40 researched targets with live routes at that round — course directories (UK/Gulf/global), CPD accreditors, five PM podcasts that take guests, seven publications with submission pages, the journalist-request stack, 2026 expo dates, speaker directories and partnership targets; later research rounds grew the directory to ~100 rows including the verified skip list.",
  "Assign an owner per category; statuses tracked in the yellow columns."),
 ("Authority & AI-visibility research round", "Platform", "Major",
  "A second deep-research pass (34 verified searches) added 26 authority and listing platforms under a new Authority & Listings area: the Credential Engine Registry (free, consumed by US state systems), CareerOneStop, the Knowledge Panel entity-graph stack, D-U-N-S, ISNI, Tracxn/Dealroom/Magnitt startup databases, Gulf job boards and Certuvo tool directories — plus corrected education schema (the Course Info rich result was retired in 2025; Course List remains). GROWTH PLAYBOOK technique 14 carries the evidence-ranked top-10 AI-visibility actions; the verified skip-list (Golden dead, BBB pay-to-play, UKRLP needs a UK entity) is on PR & Target Directory so nobody re-litigates it.",
  "Work the Critical rows first: Credential Engine, the entity graph, education schema. The 20-prompt AI citation audit runs monthly."),
 ("Institution authority + free publishing + free news research round", "Platform", "Major",
  "A third research pass (33 verified searches) covered three lanes. (1) Institution authority: the ISO/IEC 17024 accreditation path (ANAB/UKAS/IAS) and an honest read on Ofqual/KHDA/ACTVET/TVTC regulatory recognition — partner-institute route today; The Conversation's academic-affiliation rule; buildingSMART; UNESCO Global Skills Academy; UK award shortlists. (2) Free educational publishing: Zenodo DOIs, OER Commons + MERLOT (non-university contributors verified), TrainingZone/HRZone, six PM guest-post routes. (3) Free news: PRLog + openPR verified as the free stack (IssueWire free tier has no Google News; EIN has no free tier); press-release links are nofollow — value is AI-readable corroboration. Dead ends recorded on the skip list (ESCO, Wikibooks/Wikiversity COI, BusinessBalls, unverified PR wires).",
  "Playbook technique 16 is the pipeline; the strategic 17024 row is leadership's call, not marketing's."),
 ("Objective dimension added — performance is now testable per campaign", "Platform", "Major",
  "Every log tab carries an Objective dropdown (11 effort categories: honorary certification outreach, per-credential certification sales, authority & entity building, content & SEO, community, partnerships & PR, events, Certuvo, brand). The Content Calendar's existing Objective column became the dropdown — no duplicate column. The new Objective Performance sheet totals activities, minutes, content, community actions, leads, meetings and revenue per category, plus a person × objective minutes matrix, so each campaign is judged on its own numbers.",
  "Tag every row; the red '(no objective set)' line must be zero by each Monday review."),
 ("Brand / property dimension + the web domain map", "Platform", "Major",
  "Every log tab (and Platform Setup and Master Tasks) now carries a For (brand) column: PCI AI - Institute, PCL-AI, PFL-AI or PML-AI certification, PCI World, Certuvo, or All/shared — so results are testable per property as well as per campaign. The Content Calendar's old 3-value Brand column became this dropdown; Certuvo-only platforms (G2, Capterra, AlternativeTo, Udemy…) come pre-tagged on Platform Setup. Objective Performance gained a RESULTS BY BRAND table. START HERE §11 now maps the five web domains — projectcontrolsinstitute.org, pciai.org, pciglobal.ai, pciworld.org, mypci.org — to the brand each serves, and every UTM/press link must land on one of them.",
  "Tag both columns on every row; the manager corrects the domain map if a purpose is wrong."),
 ("Value ranking — so the team knows what an hour is worth", "Process", "Major",
  "Every effort category now carries a value rank on Objective Performance (1 = most valuable: the three certification-sales campaigns first, then honorary outreach as the strategic wedge, then partnerships, authority, content, events, Certuvo, community, brand). Every platform carries a value rank across the whole estate on Platform Setup column P (1 = most valuable; priority tier first, curated order within tier; the top 10 glow gold, and Platform Progress mirrors the rank beside actual activity). Compare rank against Share of minutes / minutes spent each Monday: when low-value work eats high share, rebalance.",
  "Ranks are pre-filled management judgment in editable yellow cells — re-rank when strategy changes, do not silently ignore them."),
 ("Direct channels, job postings, and a country map for every platform", "Platform", "Major",
  "DAILY ENTRY gained activity types for the missing channels — 'Email campaign sent', 'WhatsApp / Telegram / SMS sent', 'Job post published' — with how-many counts. A Job Postings tab tracks every open position on every platform (one row per platform per position, applicant counts). Country research (30 searches) mapped where each of 133 platforms is strongest — Platform Setup column Q — and added Snapchat (about 72% of Saudi internet users — the biggest KSA channel the tracker was missing), Careers Page + Google for Jobs (free, live 2026), Naukri.com and Jobberman; verified job-board verdicts (LinkedIn free post rules, Indeed's 31 Mar 2026 organic degradation, Wellfound free ATS) are on PR & Target Directory; Xing, VK, WeChat-class and junior boards are documented skips.",
  "Post every open role on 3+ platforms including one Gulf board; check column Q before picking a channel for a country campaign."),
 ("Keyword Plan — SEO effort graded Easy / Medium / Hard", "Platform", "Major",
  "76 researched keywords across 10 clusters, difficulty graded from 33 live SERP samples so effort goes where it can win: Easy (thin/forum SERPs — attack now), Medium (winnable), Hard (PMI/AACE/RICS territory — pillars only). Ten P1 attack keywords are ordered with reasons. Critical finding: projectcontrolsinstitute.com — a competitor training provider — owns PCI's exact name in Google today, so brand defense is attack keyword #1. Owner/Status tracker columns per keyword.",
  "SEO work starts from this sheet, Easy first; re-verify quarterly in Search Console, P1 SERPs monthly."),
 ("Weekly Pulse — real performance tracking", "Process", "Major",
  "A live this-week-vs-last-week page: 14 KPIs (minutes, activities, leads, connections, messages, content, community, email recipients, WhatsApp/SMS, job posts, backlinks live, meetings by meeting date, revenue by purchase date, new leads) with week-on-week change, 4-week average, and minutes-per-person — green up, red down, Monday-anchored. The Dashboard stays the cumulative record; the Pulse is the early warning. How-To Guides gained the six missing workstream rows (email, WhatsApp/SMS, job postings, press releases, educational publishing, SEO & keywords).",
  "Manager opens Weekly Pulse every Monday before the team meeting; falling high-value efforts get rebalanced the same day."),
 ("PLATFORM GUIDE + six world-class techniques + headline KPIs", "Process", "Major",
  "The PLATFORM GUIDE sheet is the step-by-step usage manual for the entire estate: one row per platform (all 133) with the weekly play, the KPI that proves it, and time per week — identity columns update from Platform Setup automatically, gold rows mark the top-10 value platforms. GROWTH PLAYBOOK grew to 22 techniques with the missed world-class organic plays: original research & data studies (the link magnet), free tools & calculators, Arabic/Gulf localisation, internal linking architecture, employee advocacy & social-proof flywheel, and guarded programmatic geo × role pages. The Dashboard gained a live HEADLINE KPIs panel (revenue, meetings, leads, acceptance, follow-ups due, this-week numbers from Weekly Pulse). A three-judge audit (formulas / content / formatting) ran over the full file and every finding was fixed — including the superseded Certuvo separate-copy rule, misdirected Daily Log pointers, and dropdowns that undershot their lists.",
  "New joiners read TEAM GUIDE then PLATFORM GUIDE; managers assign platforms by value rank and check the KPI panel daily."),
 ("Off-page SEO at full strength — the Link Building engine", "Platform", "Major",
  "Off-page SEO now has its own system of record: the Link Building tab logs every link prospect from first contact to live (tactic, target site, our URL, anchor, Dofollow/Nofollow, status), with GROWTH PLAYBOOK technique 23 supplying the weekly programme — competitor backlink gap analysis, monthly unlinked-mention reclaim, resource-page and broken-link outreach, natural anchor rules, and hard guardrails (no paid links, PBNs or swaps — the link spam Google demotes). Guest posts, journalist quotes, podcast appearances and directory listings from the other techniques all flow into the same tracker. DAILY ENTRY gained the 'Link building / outreach (off-page)' activity type and Weekly Pulse tracks 'Backlinks gone live' week on week.",
  "Weekly: 10 prospects, 5 outreach. Cross-check earned links against Search Console monthly — the tracker is what we did, Search Console is what Google saw."),
 ("Article Bank — 5,000+ researched briefs so nobody asks what to write", "Platform", "Major",
  "A full editorial engine: 5,000+ article briefs across 25 clusters (brand and comparison flagships, EVM, IFRS and project accounting, budgeting and cash flow, dashboards, AI-in-everything, careers by role and country, industries, tools, standards spines from AACE RPs / EIA-748 / NEC4 / FIDIC / IFRS, exam prep, case studies, FAQ answers). Every row carries its primary and supporting keywords and a format-specific AI writing prompt that bans invented statistics and demands cited sources. Built on a 31-search research round: competitor blog mining, real question patterns, verified standards references, and sourced jobs-market angles. One honest correction from that research: the defensible claim is that NO certification combines project management + finance + AI (each pairwise space has entrants) — the flagship articles are framed accordingly.",
  "Writers filter Easy-difficulty rows in their cluster, claim with Owner + Status, paste the prompt, then edit like an expert and verify every number before publishing."),
 ("Content Scheduler + MAP navigation + scheduling research", "Platform", "Major",
  "Scheduling is now a managed system: the Content Scheduler holds one row per platform × cadence × window (daily / 3x weekly / weekly / fortnightly / monthly, start and end dates, posts per cycle) with planned posts computed from the cadence, published-in-window counted live from the Content Calendar, and colour-coded coverage — the Dashboard gained ACTIVE SCHEDULES and SCHEDULE COVERAGE tiles. A research round verified which platforms support scheduling natively in 2026, through what, with what window limits and formats — that truth lives in the reference table under the scheduler grid. And because the workbook now spans 41 tabs, the MAP sheet (second tab) gives one-click navigation to every sheet with a one-line description of each.",
  "Every recurring posting commitment gets a scheduler row; coverage under 100% gets fixed before the Friday review."),
 ("Final QA audit — ten lenses, and the bug it caught", "Process", "Major",
  "A ten-lens audit ran over the finished file: formulas and ranges, dropdown/list integrity, content cross-references, rendered formatting, protection and lock, Article Bank quality (all 5,013 rows), independent KPI re-derivation, guides coherence, print/export, and a fresh-eyes day-one user walk. It caught a genuine defect: adding the 'Backlinks gone live' row to Weekly Pulse had shifted the rows beneath it, so the Dashboard's REVENUE THIS WEEK tile was silently displaying meetings and MEETINGS THIS WEEK was displaying backlinks. Dashboard tiles now reference Weekly Pulse KPIs BY NAME (never by a fixed row), two more tiles were added (backlinks, content this week), and a regression check asserts every tile reads the row it claims. Also fixed: formula cells on Message Bank and Lists were editable (a stray paste could have destroyed a character count or the roster selector) — every formula is now locked except inside the example rows that must stay deletable; and START HERE never named the navigation, so row 3 now signposts TEAM GUIDE, PLATFORM GUIDE, MAP and Article Bank.",
  "The audit script ships with the build: 62 structural checks plus a 52-assertion live simulation must pass before any release."),
 ("Independent critical re-audit — treated as if built by someone else", "Process", "Major",
  "A fresh adversarial audit assumed nothing and trusted nothing: an independent numerical oracle, a static 'silent-zero' analysis of all 2,078 conditional criteria (a SUMIFS whose criteria text no longer exists returns 0 forever with no error — none found), Excel-integrity checks for repair-prompt triggers (none), a cached-error sweep (none), and cadence maths proved across every scheduling frequency. Real defects found and fixed: (1) 'Content published' meant different things on Weekly Pulse (by published date) and the Dashboard (by status) — now both mean status Published/Repurposed, so the two pages can never disagree; (2) 'Connection requests sent' appeared twice on the Dashboard from two different sources with the same label — now each names its source; (3) a schedule row with the end date before the start date produced NEGATIVE planned posts and poisoned coverage — now blank, with the dates flagged red; (4) formula cells on Message Bank and Lists were editable; (5) 24 sheets still showed raw gridlines while the rest did not — the whole workbook now shares one look; (6) a first-run screen of zeros looked broken — Dashboard and Weekly Pulse now explain themselves until data arrives.",
  "The release gate is now 66 structural checks + a 51-assertion live simulation + the static and integrity analyses — all must pass."),
 ("Premium formatting pass", "Process", "Minor",
  "Everything now matches the workbook's own design system (Arial, navy titles, banded headers), tabs are colour-grouped and reordered by workflow, panes freeze on every long sheet, and every sheet prints one page wide with repeating headers.",
  ""),
]
_r = un.max_row + 1
_start_n = 26
for _i, (t, ty, sv, w, td_) in enumerate(V8_NOTES):
    un.cell(_r, 1).value = _start_n + _i
    un.cell(_r, 2).value = t
    un.cell(_r, 3).value = ty
    un.cell(_r, 4).value = sv
    un.cell(_r, 5).value = w
    un.cell(_r, 6).value = td_
    _r += 1
un.sheet_view.showGridLines = False
un["A1"] = "UPGRADE NOTES  —  what was tested, what was missing, what changed"
title_row(un, 6, un["A1"].value,
          "Three independent audit passes ran end to end — every formula, every platform "
          "list, and the whole system as a product a team runs daily — then three judges "
          "verified the rebuild. Type: PLATFORM = a capability the workbook lacked; "
          "PROCESS = a workflow rule no spreadsheet can enforce alone.")
for c in range(1, 7):
    h = un.cell(4, c); h.font = H2; h.fill = BANDFILL
    h.alignment = Alignment(wrap_text=True, vertical="center")
    h.border = BOX
for r in range(5, un.max_row + 1):
    if not un.cell(r, 2).value:
        continue
    for c in range(1, 7):
        cell = un.cell(r, c)
        cell.font = Font(name="Arial", size=8.5, bold=(c in (2, 3)))
        cell.alignment = WRAP; cell.border = BOX
        if r % 2: cell.fill = ZEBRA
    # height follows the longest text so no changelog entry ever clips
    _elen = len(str(un.cell(r, 5).value or ""))
    _flen = len(str(un.cell(r, 6).value or ""))
    un.row_dimensions[r].height = max(44, min(260, int(max(_elen / 56, _flen / 24) + 1) * 10 + 8))
un.freeze_panes = "A5"
un.sheet_properties.tabColor = "FFC00000"
un.protection.sheet = True

# ------------------------------------ 5a. GLOSSARY sheet
# A readability audit rated three sheets 4-5/10 for a non-native English
# reader and counted 73 unexplained terms on PLATFORM GUIDE alone. The terms
# are the right ones — what was missing was somewhere to look them up.
from glossary_data import GLOSSARY as _GLOSS

gl = wb.create_sheet("Glossary")
gl.sheet_view.showGridLines = False
title_row(gl, 2, "GLOSSARY  —  every term this workbook uses without stopping to explain it",
          "The workbook is written for a team spread across several countries, and it uses "
          "the vocabulary the work is actually done in. Every one of those words is here, in "
          "plain English, grouped by where you will meet it. Nothing on this page is typed "
          "or calculated — it is here to be read.")
# The group already has its own band above each block, so a column repeating
# it on every row would be noise. Two columns: the term, and what it means.
header_band(gl, 3, ["Term", "What it means"], widths=[30, 118])
_gr = 4
_last_group = None
for _grp, _term, _def in _GLOSS:
    if _grp != _last_group:
        section(gl, _gr, 2, _grp.upper())
        _gr += 1
        _last_group = _grp
    gl.cell(_gr, 1).value = _term
    gl.cell(_gr, 2).value = _def
    for _c in (1, 2):
        cell = gl.cell(_gr, _c)
        cell.font = BODYB if _c == 1 else BODY
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1 if _c == 1 else 0)
        if _gr % 2 == 0:
            cell.fill = ZEBRA
    gl.row_dimensions[_gr].height = max(17, 11 * (len(_def) // 112 + 1) + 6)
    _gr += 1
_gn = gl.cell(_gr + 1, 1)
_gn.value = ("If a word on any sheet stopped you and it is not here, that is a gap worth "
             "reporting in the Monday review — the next version adds it.")
_gn.font = GREYNOTE
gl.merge_cells(start_row=_gr + 1, start_column=1, end_row=_gr + 1, end_column=2)
gl.freeze_panes = "A4"
gl.protection.sheet = True
gl.print_area = f"A1:B{_gr + 1}"
gl.print_title_rows = "1:3"
gl.sheet_properties.tabColor = "FF808080"

# ------------------------------------ 5b. MAP sheet — one-click navigation
# 40 tabs deserve a contents page: every sheet, grouped, hyperlinked.
mp = wb.create_sheet("MAP")
mp.sheet_view.showGridLines = False
title_row(mp, 4, "MAP  —  every tab, one click",
          "Click any name to jump there. Colour groups: gold = start here, green = "
          "guides, blue = daily work, navy = results, purple = management, grey = "
          "reference. New joiners: START HERE, then TEAM GUIDE, then PLATFORM GUIDE.")
_MAP_GROUPS = [
 ("READ FIRST", [
  ("START HERE", "The operating model: rules, targets, roster, domains"),
  ("TEAM GUIDE", "Where to log what — the 15-minute onboarding"),
  ("GROWTH PLAYBOOK", "23 growth techniques in working detail"),
  ("PLATFORM GUIDE", "How to use every platform, step by step"),
  ("PR & Target Directory", "Named routes: publications, podcasts, boards, skips"),
  ("UPGRADE NOTES", "The changelog: what was tested and changed")]),
 ("DAILY WORK", [
  ("DAILY ENTRY", "The one sheet everyone fills, every day"),
  ("LinkedIn Outreach", "One row per lead — the honorary engine"),
  ("Partnership Pipeline", "Organisations: associations, universities, employers"),
  ("Content Calendar", "Every piece of content, planned to published"),
  ("Content Scheduler", "Cadences, windows and coverage per platform"),
  ("Community & PR", "Answers, mentions, press and communities"),
  ("Job Postings", "Every open role on every board"),
  ("Link Building", "Off-page SEO: every link prospect to live"),
  ("Experiments", "A/B tests with honest conclusions"),
  ("UTM Builder", "Tracked links — the only links we share"),
  ("SEO Clusters", "Pillars and spokes with Search Console numbers"),
  ("Keyword Plan", "76 graded keywords: what to attack, in order"),
  ("Article Bank", "5,000+ article briefs with keywords + AI prompts"),
  ("Glossary", "Every term this workbook uses, in plain English"),
  ("Daily Log", "Optional auto day-summary")]),
 ("RESULTS", [
  ("Weekly Pulse", "This week vs last week, live"),
  ("Dashboard", "The cumulative record + headline KPIs"),
  ("Summary", "One page for management"),
  ("Objective Performance", "Results by campaign and by brand"),
  ("Team Scorecard", "Per-person outreach numbers"),
  ("Employee Score", "The weighted score behind reviews"),
  ("Weekly Review", "Friday wins, misses and next week"),
  ("Platform Progress", "Activity per platform vs value rank"),
  ("Who Did What", "Per-person split per platform"),
  ("Accounts Register", "Every account, its owner and vault entry")]),
 ("MANAGEMENT", [
  ("Master Tasks", "The full workstream list"),
  ("Platform Setup", "All 133 platforms: setup, rank, geography"),
  ("Publishing Plan", "The ten publishing platforms, ranked"),
  ("Channel Costs", "What each paid channel costs"),
  ("QA & Compliance", "The 15 checks that protect reputation"),
  ("Message Bank", "Approved outreach messages only")]),
 ("REFERENCE", [
  ("LinkedIn Playbook", "The outreach method, step by step"),
  ("How-To Guides", "Training manual per workstream"),
  ("Benchmarks", "What good numbers look like in 2026"),
  ("Lists", "Every dropdown's source values")]),
]
from openpyxl.worksheet.hyperlink import Hyperlink as _HL
_mr = 4
for _gname, _entries in _MAP_GROUPS:
    section(mp, _mr, 4, _gname)
    _mr += 1
    for _sname, _desc in _entries:
        c = mp.cell(_mr, 1)
        c.value = _sname
        c.hyperlink = _HL(ref=f"A{_mr}", location=f"'{_sname}'!A1")
        c.font = Font(name="Arial", size=10, bold=True, color="1D4ED8", underline="single")
        mp.merge_cells(start_row=_mr, start_column=2, end_row=_mr, end_column=4)
        d = mp.cell(_mr, 2)
        d.value = _desc
        d.font = BODY
        if _mr % 2 == 0:
            for _c in range(1, 5):
                mp.cell(_mr, _c).fill = ZEBRA
        mp.row_dimensions[_mr].height = 17
        _mr += 1
    _mr += 1
mp.column_dimensions["A"].width = 24
for _cl in ("B", "C", "D"):
    mp.column_dimensions[_cl].width = 22
mp.protection.sheet = True
mp.print_area = f"A1:D{_mr}"
mp.sheet_properties.tabColor = "FFC9A227"

# ------------------------------------------------ 6. Global formatting pass
TABS = {"START HERE": "C9A227", "MAP": "FFC9A227",
        "TEAM GUIDE": "FF548235", "GROWTH PLAYBOOK": "FF548235",
        "PLATFORM GUIDE": "FF548235",
        "UPGRADE NOTES": "FFC00000", "DAILY ENTRY": "FF2E75B6", "LinkedIn Outreach": "FF2E75B6",
        "Partnership Pipeline": "FF2E75B6", "Content Calendar": "FF2E75B6",
        "Content Scheduler": "FF2E75B6",
        "Community & PR": "FF2E75B6", "Job Postings": "FF2E75B6",
        "Link Building": "FF2E75B6",
        "Experiments": "FF2E75B6", "UTM Builder": "FF2E75B6",
        "SEO Clusters": "FF2E75B6", "Keyword Plan": "FF2E75B6", "Article Bank": "FF2E75B6",
        "Daily Log": "FF2E75B6",
        "Weekly Pulse": "FF1F3864",
        "Dashboard": "FF1F3864", "Summary": "FF1F3864", "Objective Performance": "FF1F3864",
        "Employee Score": "FF1F3864",
        "Team Scorecard": "FF1F3864", "Platform Progress": "FF1F3864", "Who Did What": "FF1F3864",
        "Weekly Review": "FF1F3864", "Accounts Register": "FF1F3864",
        "Master Tasks": "FF7030A0", "Platform Setup": "FF7030A0", "Publishing Plan": "FF7030A0",
        "Message Bank": "FF7030A0", "LinkedIn Playbook": "FF808080", "How-To Guides": "FF808080",
        "Benchmarks": "FF808080", "QA & Compliance": "FF7030A0", "Channel Costs": "FF7030A0",
        "Lists": "FF808080", "Glossary": "FF808080"}
FREEZE = {"DAILY ENTRY": "C4", "LinkedIn Outreach": "E4", "Partnership Pipeline": "D4",
          "Content Calendar": "D4", "Community & PR": "D4", "Master Tasks": "D4",
          "Platform Setup": "E4", "Platform Progress": "B4", "Who Did What": "B4",
          "Team Scorecard": "B4", "Employee Score": "B4", "SEO Clusters": "B4",
          "Experiments": "E4", "Daily Log": "C4", "Weekly Review": "D4",
          "Channel Costs": "A4", "Lists": "A4"}
for name, color in TABS.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = color
# belt-and-braces: the helper columns must ship hidden whatever an engine did
de8 = wb["DAILY ENTRY"]
de8.column_dimensions["K"].hidden = True
de8.column_dimensions["L"].hidden = True
for name, cell in FREEZE.items():
    if name in wb.sheetnames:
        wb[name].freeze_panes = cell
# tab order: guides first, then daily work, then management, then reference
ORDER = ["START HERE", "MAP", "TEAM GUIDE", "GROWTH PLAYBOOK", "PLATFORM GUIDE",
         "PR & Target Directory", "UPGRADE NOTES",
         "DAILY ENTRY", "LinkedIn Outreach", "Partnership Pipeline", "Content Calendar",
         "Content Scheduler", "Community & PR", "Job Postings", "Link Building",
         "Experiments", "UTM Builder",
         "SEO Clusters", "Keyword Plan", "Article Bank", "Daily Log",
         "Weekly Pulse", "Dashboard", "Summary", "Objective Performance", "Team Scorecard",
         "Employee Score", "Weekly Review",
         "Platform Progress", "Who Did What", "Accounts Register",
         "Master Tasks", "Platform Setup", "Publishing Plan",
         "Channel Costs", "QA & Compliance", "Message Bank",
         "LinkedIn Playbook", "How-To Guides", "Benchmarks", "Glossary", "Lists"]
wb._sheets = [wb[n] for n in ORDER if n in wb.sheetnames] + \
             [ws for ws in wb._sheets if ws.title not in ORDER]

# page setup: every sheet exports/prints one page wide, landscape when wide,
# with its header rows repeating — a printed pack should look as deliberate
# as the screen.
# One workbook, one look: raw gridlines made half the sheets look like an
# unfinished spreadsheet next to the designed ones. Every table carries its own
# borders, so gridlines are noise. 'Lists' keeps them — it is the raw reference.
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = (ws.title == "Lists")

from openpyxl.worksheet.page import PageMargins
for ws in wb.worksheets:
    ws.page_setup.fitToWidth = 1
    # Summary promises "one page" — everything else may flow down pages
    ws.page_setup.fitToHeight = 1 if ws.title == "Summary" else 0
    # uniform elegant margins + a quiet branded footer on every printed page
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.55, bottom=0.55,
                                  header=0.25, footer=0.3)
    ws.oddFooter.left.text = "PCI AI  —  Growth OS"
    ws.oddFooter.left.size = 7
    ws.oddFooter.left.color = "808080"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 7
    ws.oddFooter.right.color = "808080"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ncols = ws.max_column
    if ncols > 8:
        ws.page_setup.orientation = "landscape"
    if ws.title in ("DAILY ENTRY", "LinkedIn Outreach", "Partnership Pipeline",
                    "Content Calendar", "Community & PR", "Master Tasks",
                    "Platform Setup", "Accounts Register"):
        ws.print_title_rows = "1:3" if ws.title != "Accounts Register" else "1:6"

# ------------------------------------ 7. Padding polish + formula lock
# Cell "padding": lead columns of the wide guide tables get an indent so text
# never touches the border; the tables read like set type, not a data dump.
for _wsn, _col, _first, _last in (("PLATFORM GUIDE", 1, 4, 140),
                                  ("Keyword Plan", 1, 4, 95),
                                  ("PR & Target Directory", 2, 5, 120)):
    _wsx = wb[_wsn]
    for _r in range(_first, _last + 1):
        c = _wsx.cell(_r, _col)
        if c.value is None:
            continue
        _old = c.alignment
        c.alignment = Alignment(wrap_text=_old.wrap_text, vertical=_old.vertical or "top",
                                horizontal=_old.horizontal, indent=1)

# ------------------------------------ 7a. The example-row trap, closed
# An independent reliability review found the worst defect in the file. Every
# log sheet shipped a worked example on its first grid row, and every report
# range started BELOW it. A user who typed over the example — the natural
# reading of "delete before use" — lost 100% of that work from every report,
# silently: the row validated, its dropdowns worked, its own formulas
# computed, and the Dashboard read zero.
#
# The fix removes the trap rather than papering over it. Report ranges now
# start at the first grid row, the example rows are cleared, and the worked
# examples move to TEAM GUIDE where they can teach without being typed over.
_GRID_START = {"DAILY ENTRY": (7, 4, 1006), "LinkedIn Outreach": (5, 4, 1203),
               "Content Calendar": (5, 4, 403), "Community & PR": (5, 4, 403),
               "Partnership Pipeline": (5, 4, 403), "Job Postings": (5, 4, 203),
               "Link Building": (5, 4, 403), "Content Scheduler": (5, 4, 103)}
_EX_ROWS = {"DAILY ENTRY": (4, 5, 6), "LinkedIn Outreach": (4,),
            "Content Calendar": (4,), "Community & PR": (4,),
            "Partnership Pipeline": (4,), "Job Postings": (4,),
            "Link Building": (4,), "Content Scheduler": (4,),
            "Experiments": (4,), "UTM Builder": (4,)}

_range_pat = {}
for _sn, (_old, _new, _last) in _GRID_START.items():
    _range_pat[_sn] = re.compile(
        r"('" + re.escape(_sn) + r"'!\$([A-Z]{1,2})\$)" + str(_old)
        + r"(:\$[A-Z]{1,2}\$" + str(_last) + r")")
_shifted = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            new = v
            for _sn, pat in _range_pat.items():
                if f"'{_sn}'" in new:
                    new = pat.sub(lambda m: m.group(1) + str(_GRID_START[_sn][1])
                                  + m.group(3), new)
            if new != v:
                cell.value = new
                _shifted += 1

# DAILY ENTRY's two day-counting helpers ran from row 7; they now run from the
# first grid row, and their expanding windows are re-anchored to match.
_de7a = wb["DAILY ENTRY"]
for _r in range(4, 1007):
    _de7a.cell(_r, 11).value = (f'=IF($A{_r}="",0,IF(COUNTIF($A$4:$A{_r},$A{_r})=1,1,0))')
    _de7a.cell(_r, 12).value = (f'=IF($A{_r}="",0,'
                                f'IF(COUNTIFS($A$4:$A{_r},$A{_r},$B$4:$B{_r},$B{_r})=1,1,0))')
    for _c in (11, 12):
        _de7a.cell(_r, _c).font = BODY

# the worked examples move to TEAM GUIDE, then the grid rows are cleared
_EX_TEXT = []
for _sn, _rows in _EX_ROWS.items():
    _wsx = wb[_sn]
    for _r in _rows:
        _vals = []
        for _c in range(1, min(_wsx.max_column, 20) + 1):
            _v = _wsx.cell(_r, _c).value
            if _v is None or (isinstance(_v, str) and _v.startswith("=")):
                continue
            _h = _wsx.cell(3, _c).value
            if _h and "EXAMPLE" not in str(_v).upper():
                _vals.append(f"{str(_h).splitlines()[0]}: {_v}")
        if _vals:
            _EX_TEXT.append((_sn, "  ·  ".join(_vals[:8])))
        for _c in range(1, _wsx.max_column + 1):
            _cell = _wsx.cell(_r, _c)
            if not (isinstance(_cell.value, str) and _cell.value.startswith("=")):
                _cell.value = None

# Experiments numbered from a hardcoded 1 on the example row, so deleting it
# started the sequence at EXP-002. It now derives from the row.
_exp7 = wb["Experiments"]
for _r in range(4, 54):
    if _exp7.cell(_r, 2).value is not None or _r <= 5:
        _exp7.cell(_r, 1).value = f'="EXP-"&TEXT(ROW()-3,"000")'
        _exp7.cell(_r, 1).font = BODY

# --- BLOCKER: ONE text date anywhere in DAILY ENTRY column A blanked "Last
# worked on", "Days since" and "Who worked on it last" for ALL 133 platforms.
# The array multiplies a boolean by the date column; one text value makes the
# whole product #VALUE!, and the IFERROR swallowed it into "". A manager read
# that as "no platform has ever been worked on".
_pp7 = wb["Platform Progress"]
_DEA7 = "'DAILY ENTRY'!$A$4:$A$1006"
_DEC7 = "'DAILY ENTRY'!$C$4:$C$1006"
_DEB7 = "'DAILY ENTRY'!$B$4:$B$1006"
_SAFE_DATE = f"IF(ISNUMBER({_DEA7}),{_DEA7},0)"
for _r in range(4, 137):
    _pp7.cell(_r, 7).value = (
        f'=IF($D{_r}=0,"",IFERROR(SUMPRODUCT(MAX(({_DEC7}=$A{_r})*{_SAFE_DATE})),""))')
    _pp7.cell(_r, 9).value = (
        f'=IF($G{_r}="","",IFERROR(INDEX({_DEB7},'
        f'MATCH(1,({_DEC7}=$A{_r})*({_SAFE_DATE}=$G{_r}),0)),""))')
_pp7.conditional_formatting.add(
    "H4:H136", _FR(formula=["$H4<0"], fill=PatternFill("solid", fgColor="F4CCCC")))

# --- a fat-finger had no ceiling: 100000 in "how many" read as 1,250,200% of
# target and the row printed "On target".
for _sn, _col, _r1, _r2, _lo, _hi, _msg in (
        ("DAILY ENTRY", "F", 4, 1006, 0, 2000, "How many — 0 to 2000. A four-figure "
         "entry is almost always a typo; split a genuine bulk send across rows."),
        ("DAILY ENTRY", "G", 4, 1006, 0, 960, "Minutes spent — 0 to 960 (there are "
         "960 minutes in a 16-hour day)."),
        ("Content Scheduler", "E", 4, 103, 1, 50, "Posts per cycle — 1 to 50.")):
    _wsx = wb[_sn]
    _wsx.data_validations.dataValidation = [
        d for d in _wsx.data_validations.dataValidation
        if not (d.type == "whole" and _col in str(d.sqref))]
    _dv7 = DataValidation(type="whole", operator="between", formula1=str(_lo),
                          formula2=str(_hi), allow_blank=True)
    _dv7.error = _msg
    _dv7.errorTitle = "Out of range"
    _dv7.showErrorMessage = True
    _wsx.add_data_validation(_dv7)
    _dv7.add(f"{_col}{_r1}:{_col}{_r2}")

# --- the sheets end somewhere, and typing past the end used to be invisible.
for _sn, (_o7, _n7, _last7) in _GRID_START.items():
    _wsx = wb[_sn]
    _stop = _wsx.cell(_last7 + 1, 1)
    _stop.value = ("LAST ROW — anything typed below this line is invisible to every "
                   "report. Archive a dated copy and start a new file.")
    _stop.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    _stop.fill = PatternFill("solid", fgColor="C00000")
    _stop.alignment = Alignment(vertical="center")
    _wsx.merge_cells(start_row=_last7 + 1, start_column=1,
                     end_row=_last7 + 1, end_column=min(_wsx.max_column, 10))
    _wsx.row_dimensions[_last7 + 1].height = 18

# --- the 4-week average divided by a hard 4 from day one, so week 1 read as
# 4x its own average. The divisor is now however many completed weeks exist.
_WEEKS = ("MIN(4,MAX(1,ROUNDDOWN(($B$3-IF(N('START HERE'!$B$20)=0,$B$3-28,"
          "'START HERE'!$B$20))/7,0)))")
for _r in range(6, 20):
    _v = wp.cell(_r, 5).value
    if isinstance(_v, str) and _v.endswith(")/4"):
        wp.cell(_r, 5).value = _v[:-1] + _WEEKS
wp.cell(5, 5).value = "4-week avg (completed weeks)"

# --- a trailing space on a tag ("honorary certification outreach ") removed
# the row from its category AND from the untagged catch-all: it vanished from
# the table entirely. The untagged row is now a RESIDUAL, so a mistyped tag
# always surfaces somewhere.
_ID_DE7 = "'DAILY ENTRY'!$A$4:$A$1006"
_DE_G7 = "'DAILY ENTRY'!$G$4:$G$1006"
_ID_LO7 = "'LinkedIn Outreach'!$C$4:$C$1203"
for _f7, _u7 in ((_OBJ_FIRST, _untagged), (_BR_FIRST, _b_untag)):
    op.cell(_u7, 2).value = f'=COUNTIF({_ID_DE7},"<>")-SUM(B{_f7}:B{_u7 - 1})'
    op.cell(_u7, 3).value = f'=SUMIFS({_DE_G7},{_ID_DE7},"<>")-SUM(C{_f7}:C{_u7 - 1})'
    op.cell(_u7, 6).value = f'=COUNTIF({_ID_LO7},"<>")-SUM(F{_f7}:F{_u7 - 1})' 
op.cell(_untagged, 1).value = "(no objective set, or a tag that is not on the list)"
op.cell(_b_untag, 1).value = "(no brand set, or a tag that is not on the list)"

# --- and the tag columns get the same red-if-not-on-the-list guard the
# employee-name columns already had
for _sn, _mcol, _bcol, _last8 in (("DAILY ENTRY", "M", "N", 1006),
                                  ("Community & PR", "T", "U", 403),
                                  ("LinkedIn Outreach", "AO", "AP", 1203),
                                  ("Partnership Pipeline", "X", "Y", 403)):
    _wsx = wb[_sn]
    _wsx.conditional_formatting.add(
        f"{_mcol}4:{_mcol}{_last8}",
        _FR(formula=[f'AND(${_mcol}4<>"",COUNTIF(Lists!$Y$4:$Y$14,${_mcol}4)=0)'],
            fill=PatternFill("solid", fgColor="F4CCCC")))
    _wsx.conditional_formatting.add(
        f"{_bcol}4:{_bcol}{_last8}",
        _FR(formula=[f'AND(${_bcol}4<>"",COUNTIF(Lists!$O$4:$O$10,${_bcol}4)=0)'],
            fill=PatternFill("solid", fgColor="F4CCCC")))
_cc9 = wb["Content Calendar"]
_cc9.conditional_formatting.add(
    "F4:F403", _FR(formula=['AND($F4<>"",COUNTIF(Lists!$Y$4:$Y$14,$F4)=0)'],
                   fill=PatternFill("solid", fgColor="F4CCCC")))
_cc9.conditional_formatting.add(
    "S4:S403", _FR(formula=['AND($S4<>"",COUNTIF(Lists!$O$4:$O$10,$S4)=0)'],
                   fill=PatternFill("solid", fgColor="F4CCCC")))

# --- BY AREA showed 0% for areas that carry no tasks at all, reading as
# failure where the truth is "nothing planned here yet"
for _r in range(65, 65 + 13):
    if str(db8.cell(_r, 4).value or "").startswith("=IFERROR(C"):
        db8.cell(_r, 4).value = f'=IF($B{_r}=0,"n/a",$C{_r}/$B{_r})'

# --- Channel Costs section 8 summed ten rows that do not exist
db8.cell(_dbrow("Monthly channel cost (Channel Costs)"), 2).value = \
    "=SUM('Channel Costs'!$B$4:$B$13)"

# --- "Working days per week" drives all eight weekly targets and was painted
# as an input cell, but shipped locked
wb["START HERE"]["B21"].protection = Protection(locked=False)

# --- Lists is the dropdown master for the whole workbook and was fully
# unlocked to everyone. Renaming one platform silently desynchronises every
# row already logged. The manager holds the password; the rows below the
# seeded values stay open so the lists can still be extended.
_ls7 = wb["Lists"]
_LIST_LEN = {}
for _c in range(1, 26):
    _n = 0
    for _r in range(4, 251):
        if _ls7.cell(_r, _c).value not in (None, ""):
            _n = _r
    _LIST_LEN[_c] = _n
for _c in range(1, 26):
    for _r in range(4, 251):
        _ls7.cell(_r, _c).protection = Protection(locked=(_r <= _LIST_LEN[_c]))

# --- two sheets told the user to delete a row that protection forbade
for _sn in ("Experiments", "UTM Builder"):
    wb[_sn].protection.deleteRows = False

# --- formula columns painted as input cells taught the wrong lesson
_lo7 = wb["LinkedIn Outreach"]
for _r in range(4, 1204):
    for _c in (32, 40):                     # follow-up 2 due, duplicate flag
        _lo7.cell(_r, _c).fill = ZEBRA if _r % 2 == 0 else PatternFill()

# --- Who Did What advertised nine empty roster slots
_wdw = wb["Who Did What"]
for _c in range(2, 12):
    _v = _wdw.cell(3, _c).value
    if isinstance(_v, str) and "empty slot" in _v:
        _wdw.cell(3, _c).value = _v.replace('"(empty slot)"', '""')

# --- the word "Status" headed four Dashboard columns and never once carried a
# value. A manager scanning for red found an empty promise. Every metric with
# a target now says where it stands; the rest say so plainly.
_ST_TARGET = {   # row -> (comparison, good direction)
 21: "up", 22: "down", 26: "up", 27: "up", 28: "up", 29: "up", 30: "up", 31: "up",
 33: "up", 34: "up", 36: "down", 37: "down", 53: "up", 54: "up", 55: "up",
 56: "up", 58: "up", 59: "up",
}
for _sec_hdr in (5, 16, 25, 52):
    _r7 = _sec_hdr + 1
    while db8.cell(_r7, 1).value and not str(db8.cell(_r7, 1).value).startswith(
            ("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "MANAGER")):
        _tgt = db8.cell(_r7, 3).value
        if _tgt not in (None, ""):
            _dirn = _ST_TARGET.get(_r7, "up")
            _cmp = f"$B{_r7}>=$C{_r7}" if _dirn == "up" else f"$B{_r7}<=$C{_r7}"
            db8.cell(_r7, 4).value = f'=IF(N($C{_r7})=0,"",IF({_cmp},"On target","Behind"))'
        else:
            db8.cell(_r7, 4).value = '="—"'
        db8.cell(_r7, 4).font = BODY
        db8.cell(_r7, 4).alignment = Alignment(horizontal="center", vertical="center")
        _r7 += 1
    db8.conditional_formatting.add(
        f"D{_sec_hdr + 1}:D{_r7 - 1}",
        _FR(formula=[f'$D{_sec_hdr + 1}="On target"'],
            fill=PatternFill("solid", fgColor="C6EFCE")))
    db8.conditional_formatting.add(
        f"D{_sec_hdr + 1}:D{_r7 - 1}",
        _FR(formula=[f'$D{_sec_hdr + 1}="Behind"'],
            fill=PatternFill("solid", fgColor="F4CCCC")))
    db8.cell(_sec_hdr, 4).value = "Status"
db8.cell(2, 1).value = (str(db8.cell(2, 1).value or "").rstrip()
                        + "  A dash in Status means no target is set for that line — it is a "
                          "volume measure, not a pass/fail.")
db8.cell(2, 1).alignment = WRAP

# --- "THE FIVE NUMBERS THAT MATTER" contained no revenue, and the revenue
# block gave three disconnected figures with no net position.
_sm7 = wb["Summary"]
_sm7.cell(4, 1).value = "THE SIX NUMBERS THAT MATTER"
for _rr in range(38, 30, -1):                  # push the daily rhythm down two
    pass
_sm_style_l = copymod.copy(_sm7.cell(9, 1)._style)
_sm_style_v = copymod.copy(_sm7.cell(9, 2)._style)
_sm7.cell(10, 1).value = "6.  Net position  (revenue + signed deals − one month of channel cost)"
_sm7.cell(10, 2).value = f"=Dashboard!$B${_rtot}-Dashboard!$B${_dbrow('Monthly channel cost (Channel Costs)')}"
_sm7.cell(10, 1)._style = _sm_style_l
_sm7.cell(10, 2)._style = _sm_style_v
_sm7.cell(10, 2).number_format = '"$"#,##0'

# --- QA & Compliance promised a live signal on 15 checks and delivered two.
# Everything the workbook can already answer, it now answers here.
_qa7 = wb["QA & Compliance"]
_QA_SIGNALS = {
 7: f'=IF(Dashboard!$B${_dbrow("Live accounts without 2FA")}=0,"Pass",'
    f'"FAIL — "&Dashboard!$B${_dbrow("Live accounts without 2FA")}&" live accounts have no 2FA")'
    if any(str(db8.cell(_r, 1).value or "") == "Live accounts without 2FA"
           for _r in range(1, 130)) else None,
 8: '=IF(COUNTIF(\'Platform Setup\'!$H$4:$H$136,"*password*")+'
    'COUNTIF(\'Platform Setup\'!$H$4:$H$136,"*pwd*")=0,"Pass",'
    '"FAIL — a password may be typed into Platform Setup column H")',
 13: '=IF(COUNTIF(\'Content Calendar\'!$L$4:$L$403,"*")=0,"No published URLs yet",'
     '"Manual check — read the last 5 published URLs")',
 15: '=IF(COUNTA(\'UTM Builder\'!$A$4:$A$103)=0,"No links built yet",'
     '"Pass — "&COUNTA(\'UTM Builder\'!$A$4:$A$103)&" tagged links built")',
}
_QA_SIGNALS = {k: v for k, v in _QA_SIGNALS.items() if v}
_QA_SIGNALS[16] = ('=IF(COUNTA(\'Platform Setup\'!$F$4:$F$136)=0,"No profiles live yet",'
                   '"Manual check — read the live profile bios")')
_QA_SIGNALS[18] = ('=IF(COUNTA($E$4:$E$18)=15,"Pass — every check has an owner",'
                   '"FAIL — "&(15-COUNTA($E$4:$E$18))&" checks have no named owner")')
for _r, _f in _QA_SIGNALS.items():
    if _qa7.cell(_r, 9).value in (None, ""):
        _qa7.cell(_r, 9).value = _f
        _qa7.cell(_r, 9).font = BODY
        _qa7.cell(_r, 9).alignment = WRAP
_qa7.conditional_formatting.add(
    "I4:I18", _FR(formula=['LEFT($I4,4)="FAIL"'],
                  fill=PatternFill("solid", fgColor="F4CCCC")))
_qa7.conditional_formatting.add(
    "I4:I18", _FR(formula=['LEFT($I4,4)="Pass"'],
                  fill=PatternFill("solid", fgColor="C6EFCE")))
_qa7.column_dimensions["I"].width = 34

# --- the guidance layer's own contradictions, resolved
# 42 columns with no indication of which are needed when
_lo9 = wb["LinkedIn Outreach"]
_lo9.cell(2, 1).value = (
    "Log every lead the same day. Fill A-L when you research the lead; M-T when you send "
    "and hear back; Y-Z when you score it; AD onward only if it becomes a meeting or a "
    "sale; AO-AP (Objective and For (brand)) always. Column AN flags a duplicate the "
    "moment you paste a URL that is already in the sheet — check it before you write "
    "anything else.")
_lo9.cell(2, 1).alignment = WRAP
_lo9.row_dimensions[2].height = 30

# "never delete rows" and "delete the example row" were both true and read as
# a contradiction. There are no example rows any more, so the rule is simple.
for _r in range(1, wb["TEAM GUIDE"].max_row + 1):
    _v = wb["TEAM GUIDE"].cell(_r, 1).value
    if isinstance(_v, str) and "Never delete rows" in _v:
        wb["TEAM GUIDE"].cell(_r, 1).value = _v.replace(
            "Never delete rows yourself",
            "Never delete rows — the logs ship empty, so there is nothing to clear out "
            "before you start")

# filtering: the file told people not to filter, then told them four times to
# filter. Under the new protection settings filtering genuinely works, so the
# rule is now the useful one.
_sh10 = wb["START HERE"]
for _r in range(1, _sh10.max_row + 1):
    _v = _sh10.cell(_r, 1).value
    if isinstance(_v, str) and "Nobody sorts or filters" in _v:
        _sh10.cell(_r, 1).value = (
            "FILTERING IS ALLOWED and works on every table — the sheets are locked "
            "against changing formulas, not against looking. In a file several people "
            "have open at once, use View → Sheet View → New first so your filter is "
            "yours alone; then clear it when you are done. SORTING a shared log is "
            "still a bad idea: it moves other people's rows under them.")
        _sh10.cell(_r, 1).alignment = WRAP
    if isinstance(_v, str) and _v.strip().startswith("FOR THE MANAGER"):
        # the manager's list named four sheets and left out the three that
        # actually answer the Monday questions — and the Dashboard's own
        # subtitle then redirects to one of them. This is the real route.
        _sh10.cell(_r, 1).value = ("FOR THE MANAGER  ->  Weekly Pulse, Objective "
                                   "Performance, Team Scorecard, Dashboard, Summary")
        _sh10.cell(_r, 2).value = (
            "Read-only pages — nobody types in them. Monday, in this order: WEEKLY PULSE "
            "(what moved last week, green or red) → OBJECTIVE PERFORMANCE (which campaigns "
            "earned their hours: compare Value rank with Share of minutes) → TEAM SCORECARD "
            "column Q (the one-sentence verdict per person) → DASHBOARD section 9 DATA "
            "HEALTH (every number there must read zero before you quote anything to anyone). "
            "SUMMARY is the one page to forward upwards.")
        _sh10.cell(_r, 2).alignment = WRAP
        _sh10.row_dimensions[_r].height = 56
    if isinstance(_v, str) and "PCI World or Certuvo" in _v and "one certification" in _v:
        _sh10.cell(_r, 1).value = _v.replace(
            "PCI World or Certuvo", "PCI World, Certuvo, or All / shared when the work "
            "genuinely serves everything")
    if isinstance(_v, str) and "Lost? The MAP tab" in _v:
        _sh10.cell(_r, 1).value = _v.replace(
            "Lost? The MAP tab",
            "A word you do not know? The GLOSSARY tab. Lost? The MAP tab")

# the character counter counts the TEMPLATE, not the message that gets sent
_mb10 = wb["Message Bank"]
_mb10.cell(2, 1).value = (
    "Column D counts the TEMPLATE, including the [brackets]. After you replace "
    "[Company] and [Name] with the real ones your message gets longer — re-count before "
    "you send. The Dashboard flags anything that went out over the limit.")
_mb10.cell(2, 1).alignment = WRAP
_mb10.row_dimensions[2].height = 30

# the per-platform cadences are per ACTIVE platform, and only one of the two
# sheets that list them said so
_hg10 = wb["How-To Guides"]
_hgr = _hg10.max_row + 2
_hg10.cell(_hgr, 1).value = (
    "These cadences are PER ACTIVE PLATFORM — nobody runs all of them at once, and the "
    "hours in this sheet do not add up to one person's week. The manager assigns "
    "platforms by Value rank (Platform Setup column P) against the hours the team "
    "actually has; only what is assigned gets logged in DAILY ENTRY.")
_hg10.cell(_hgr, 1).font = GREYNOTE
_hg10.cell(_hgr, 1).alignment = WRAP
_hg10.merge_cells(start_row=_hgr, start_column=1, end_row=_hgr, end_column=8)
_hg10.row_dimensions[_hgr].height = 34

_tg7 = wb["TEAM GUIDE"]
_tr = _tg7.max_row + 2
section(_tg7, _tr, 5, "WORKED EXAMPLES  —  what a good row looks like on each log")
_tr += 1
_tg7.cell(_tr, 1).value = ("These used to sit on the first row of each log as an example to "
                           "delete. Typing over one instead of deleting it lost the work "
                           "from every report, so the logs now start empty and the examples "
                           "live here, where nothing can be typed over them.")
_tg7.cell(_tr, 1).font = GREYNOTE
_tg7.cell(_tr, 1).alignment = WRAP
_tg7.merge_cells(start_row=_tr, start_column=1, end_row=_tr, end_column=5)
_tg7.row_dimensions[_tr].height = 28
_tr += 1
for _sn, _txt in _EX_TEXT:
    _tg7.cell(_tr, 1).value = _sn
    _tg7.cell(_tr, 1).font = BODYB
    _tg7.cell(_tr, 1).border = BOX
    _tg7.cell(_tr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    _tg7.merge_cells(start_row=_tr, start_column=2, end_row=_tr, end_column=5)
    _tg7.cell(_tr, 2).value = _txt
    _tg7.cell(_tr, 2).font = BODY
    _tg7.cell(_tr, 2).border = BOX
    _tg7.cell(_tr, 2).alignment = WRAP
    if _tr % 2 == 0:
        for _c in range(1, 6):
            _tg7.cell(_tr, _c).fill = ZEBRA
    _tg7.row_dimensions[_tr].height = max(30, 12 * (len(_txt) // 88 + 1) + 6)
    _tr += 1
_tg7.print_area = f"A1:E{_tr}"

# every log now starts empty, so the "delete the example first" instruction is
# gone from the sheet headers too
for _sn in _EX_ROWS:
    _wsx = wb[_sn]
    _h2 = _wsx.cell(2, 1)
    if isinstance(_h2.value, str):
        _h2.value = re.sub(r"\s*(The example rows?[^.]*\.|Remove the example row[^.]*\.)",
                           " Start typing on row 4 — the sheet ships empty, and a worked "
                           "example for this log is on TEAM GUIDE.", _h2.value)

print(f"example-row trap closed: {_shifted} formulas re-anchored, "
      f"{len(_EX_TEXT)} worked examples moved to TEAM GUIDE")

# ------------------------------------ 7b. Independent-review usability pass
# An outside review of the finished file drove this block. Three of its
# findings made the workbook unusable rather than merely imperfect: a locked
# sheet blocks AutoFilter and column resizing unless you explicitly permit
# them (so the Article Bank's thousands of rows could not be filtered and
# nobody could widen a clipped column); three columns of dedup rules on Lists
# had never been given a width or a header band at all; and several AutoFilter
# ranges stopped short of the real table, hiding whole columns from the filter.
from openpyxl.utils import get_column_letter as _GL

_BANDRGB = "FF" + BAND


def _band_row(ws):
    """The row carrying this sheet's header band, or None."""
    for r in range(1, 9):
        n = 0
        for c in range(1, min(ws.max_column, 40) + 1):
            cell = ws.cell(r, c)
            if (cell.value and cell.fill and cell.fill.patternType == "solid"
                    and str(cell.fill.fgColor.rgb) == _BANDRGB):
                n += 1
        if n >= 3:
            return r
    return None


def _last_data_row(ws, hrow, cols):
    """Last row of the contiguous table under the header band."""
    last, blanks = hrow, 0
    for r in range(hrow + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in cols):
            last, blanks = r, 0
        else:
            blanks += 1
            if blanks >= 3:
                break
    return last


# --- Lists W/X/Y: appended late, so they never got the band or a width
_ls_fix = wb["Lists"]
for _c in range(1, 26):
    _hc = _ls_fix.cell(3, _c)
    if _hc.value and not (_hc.fill and _hc.fill.patternType == "solid"
                          and str(_hc.fill.fgColor.rgb) == _BANDRGB):
        _hc.font = H2
        _hc.fill = BANDFILL
        _hc.alignment = Alignment(wrap_text=True, vertical="center")
        _hc.border = BOX

# --- every headed column gets a width that fits what is actually in it
_MAXW = {"Lists": 34, "Article Bank": 62, "PLATFORM GUIDE": 56}
for ws in wb.worksheets:
    hrow = _band_row(ws)
    if hrow is None:
        continue
    hcols = [c for c in range(1, ws.max_column + 1) if ws.cell(hrow, c).value]
    if not hcols:
        continue
    cap = _MAXW.get(ws.title, 40)
    for c in hcols:
        L = _GL(c)
        cur = ws.column_dimensions[L].width if L in ws.column_dimensions else None
        if cur:
            continue                      # a deliberate width always wins
        longest = len(str(ws.cell(hrow, c).value))
        for r in range(hrow + 1, min(ws.max_row, hrow + 400) + 1):
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.startswith("=")):
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[L].width = max(10.0, min(float(cap), longest + 2.5))

# --- a dropdown column must be wide enough to show its longest legal value
def _dv_values(ws, dv):
    """The legal values behind one list validation, resolved through ranges."""
    f = str(dv.formula1 or "")
    if f.startswith('"') and f.endswith('"'):
        return [x for x in f[1:-1].split(",") if x]
    m = re.match(r"^='?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$", f)
    if not m:
        return []
    sname, c1, r1, c2, r2 = m.group(1), m.group(2), int(m.group(3)), m.group(4), int(m.group(5))
    if sname not in wb.sheetnames:
        return []
    src = wb[sname]
    from openpyxl.utils import column_index_from_string as _CI
    out = []
    for r in range(r1, r2 + 1):
        for c in range(_CI(c1), _CI(c2) + 1):
            v = src.cell(r, c).value
            if v is not None and not (isinstance(v, str) and v.startswith("=")):
                out.append(str(v))
    return out


import re as _re_dv
for ws in wb.worksheets:
    for dv in list(ws.data_validations.dataValidation):
        if dv.type != "list":
            continue
        vals = _dv_values(ws, dv)
        if not vals:
            continue
        need = min(38.0, max(len(v) for v in vals) + 3.0)
        for rng in str(dv.sqref).split():
            letters = sorted({_re_dv.sub(r"\d+", "", part)
                              for part in rng.split(":") if part})
            for L in letters:
                if not L:
                    continue
                cur = ws.column_dimensions[L].width if L in ws.column_dimensions else 8.43
                if (cur or 8.43) < need:
                    ws.column_dimensions[L].width = need

# --- AutoFilter must span the whole table, not a stale sub-range
for ws in wb.worksheets:
    if not ws.auto_filter.ref:
        continue
    hrow = _band_row(ws) or 3
    hcols = [c for c in range(1, ws.max_column + 1) if ws.cell(hrow, c).value]
    if not hcols:
        continue
    last = _last_data_row(ws, hrow, hcols)
    ws.auto_filter.ref = f"A{hrow}:{_GL(max(hcols))}{max(last, hrow + 1)}"

# --- DAILY ENTRY: 1,200 rows of solid yellow drowned the signal. Yellow now
# means "this one is required"; the optional columns are plain with the same
# zebra the rest of the workbook uses.
_de7 = wb["DAILY ENTRY"]
_DE_REQUIRED = {1, 2, 3, 4, 5, 7}        # date, name, platform, kind, what, minutes
for _r in range(7, 1204):
    for _c in range(1, 11):
        _cell = _de7.cell(_r, _c)
        _cell.fill = (YELLOW if _c in _DE_REQUIRED
                      else (ZEBRA if _r % 2 == 0 else PatternFill()))
_de7.cell(3, 6).value = "How many\n(optional)"
_de7.cell(3, 8).value = "Link / proof\n(optional)"
_de7.cell(3, 9).value = "Result\n(optional)"
_de7.cell(3, 10).value = "Notes or blocker\n(optional)"
_de7.cell(3, 3).comment = None
_deleg = _de7.cell(2, 1)
if isinstance(_deleg.value, str) and "yellow" in _deleg.value.lower():
    _deleg.value = _deleg.value.replace(
        "yellow ones", "yellow ones — the six yellow columns are required, "
                       "the rest are optional")
# (§7a removed the example rows entirely, so there is no delete instruction
# left to fit — the grid starts empty at row 4)
for _r in (4, 5, 6):
    _de7.row_dimensions[_r].height = 15

# --- one date format for the whole workbook. Mixed dd-mmm-yyyy / yyyy-mm-dd
# columns (LinkedIn Outreach carried both) make a team in five countries
# second-guess every date. dd-mmm-yyyy cannot be misread either way round.
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            nf = str(cell.number_format or "")
            if "yy" in nf.lower() and "h" not in nf.lower():
                cell.number_format = "dd\\-mmm\\-yyyy"

# --- not one data validation in the workbook carried a tooltip. Every
# ambiguity a first-day user hits — "how many of what?", "which objective?",
# "why won't it take my name?" — is answered somewhere in the guides and
# nowhere at the cell where the decision is made. Excel shows an input message
# the moment the cell is selected; that is where the answer belongs.
_PROMPTS = {
 ("DAILY ENTRY", "B"): ("Your name",
   "Pick your name from the list. Not there? Your manager adds it on START HERE "
   "section 6 — you cannot log anything until they do."),
 ("DAILY ENTRY", "C"): ("Which platform",
   "One row per platform. If the same piece of work went to three platforms, "
   "that is three rows."),
 ("DAILY ENTRY", "D"): ("What kind of work",
   "Pick the closest match. This drives every activity total on the Dashboard, "
   "so guessing here shows up there."),
 ("DAILY ENTRY", "F"): ("How many — count of the thing in column D",
   "One post = 1. Ten leads researched = 10. Email or WhatsApp = messages "
   "delivered. Never combine platforms in one row. Leave blank if a count makes "
   "no sense for this activity."),
 ("DAILY ENTRY", "G"): ("Minutes spent",
   "Real minutes on this one activity. 0 to 960. This is how the workbook works "
   "out what each campaign costs in time."),
 ("DAILY ENTRY", "I"): ("Result",
   "How it ended. 'Blocked' puts the row in front of the manager on Monday."),
 ("DAILY ENTRY", "M"): ("Objective — which campaign this is for",
   "WHY you did it. If the work sells one specific certification, pick that "
   "Certification Sales line; if it builds general traffic or authority, pick "
   "Content & SEO Growth or Authority & Entity Building. Required on every row."),
 ("DAILY ENTRY", "N"): ("For (brand) — who the work is for",
   "WHICH property benefits: the institute, one certification, PCI World, "
   "Certuvo, or All / shared when it genuinely serves everything. Required on "
   "every row."),
 ("DAILY ENTRY", "A"): ("Date",
   "The day the work actually happened, typed as a date. A date typed as text "
   "lands in the running totals and in no week."),
 ("LinkedIn Outreach", "B"): ("Owner",
   "Whoever researched or contacted this lead. Must match the roster exactly."),
 ("LinkedIn Outreach", "S"): ("Outcome",
   "Where this lead stands now. Meetings are counted from the Meeting date "
   "column (AG), not from here — so log the date as well as the outcome."),
 ("Content Calendar", "F"): ("Objective",
   "Which campaign this piece serves. Same list as DAILY ENTRY."),
 ("Content Calendar", "J"): ("Status",
   "'Published' also needs a Published date in column K, or the piece counts on "
   "the Dashboard and vanishes from every weekly view."),
 ("Content Scheduler", "D"): ("Cadence",
   "How often this schedule posts. Planned posts are computed from this, the "
   "window and posts-per-cycle."),
 ("Partnership Pipeline", "N"): ("Stage",
   "A deal only counts as revenue once a real contract-signed date is in column "
   "W — a typed note there is not a signature."),
}
for (_sn, _col), (_title, _msg) in _PROMPTS.items():
    if _sn not in wb.sheetnames:
        continue
    for _dv in wb[_sn].data_validations.dataValidation:
        _cols = {re.sub(r"\d+", "", p) for part in str(_dv.sqref).split()
                 for p in part.split(":")}
        if _col in _cols:
            _dv.promptTitle = _title
            _dv.prompt = _msg
            _dv.showInputMessage = True
# and a plain-language default everywhere else, so no dropdown is silent
for ws in wb.worksheets:
    for _dv in ws.data_validations.dataValidation:
        if _dv.type == "list" and not _dv.prompt:
            _dv.promptTitle = "Pick from the list"
            _dv.prompt = ("Free-typed values fall out of every report — use the "
                          "dropdown. If the value you need is missing, ask the "
                          "manager to add it to the Lists sheet.")
            _dv.showInputMessage = True

# --- the week starts on Monday in the formula and on SUNDAY for a Gulf team.
# Every Sunday of output was landing in last week's column. The anchor is now
# a setting.
_sh9 = wb["START HERE"]
_sh9.cell(22, 1).value = "Our week starts on (1 = Sunday, 2 = Monday)"
_sh9.cell(22, 1).font = BODY
_sh9.cell(22, 2).value = 2
_sh9.cell(22, 2).fill = YELLOW
_sh9.cell(22, 2).font = BODYB
_sh9.cell(22, 2).protection = Protection(locked=False)
_sh9.cell(22, 2).border = BOX
_sh9.merge_cells(start_row=22, start_column=3, end_row=22, end_column=6)
_sh9.cell(22, 3).value = ("Set 1 for a Sunday-Thursday working week (Saudi, Qatar, Kuwait, "
                          "Bahrain) so Sunday's output lands in the right week. Weekly Pulse "
                          "and Weekly Review both follow this.")
_sh9.cell(22, 3).font = GREYNOTE
_sh9.cell(22, 3).alignment = WRAP
_sh9.row_dimensions[22].height = 26
_sh9.cell(22, 2).border = BOX                # match the other setting rows
_wkdv = DataValidation(type="whole", operator="between", formula1="1", formula2="2",
                       allow_blank=False)
_wkdv.error = "1 = Sunday, 2 = Monday"
_wkdv.promptTitle = "Which day does your week start on?"
_wkdv.prompt = "1 = Sunday (Gulf working week). 2 = Monday (UK / Europe / India)."
_wkdv.showInputMessage = True
_wkdv.showErrorMessage = True
_sh9.add_data_validation(_wkdv)
_wkdv.add("B22:B22")
wp.cell(3, 2).value = "=TODAY()-MOD(TODAY()-IF(N('START HERE'!$B$22)=0,2,'START HERE'!$B$22),7)"
wp.cell(3, 3).value = ('=IF(N(\'START HERE\'!$B$22)=1,"Sunday-Thursday week '
                       '(set on START HERE)","Monday-Sunday week (set on START HERE)")')
wp.cell(3, 3).font = GREYNOTE

# --- every fill colour written as six hex digits comes back out of the file
# as "00RRGGBB" — alpha zero. Excel ignores the alpha channel on a solid fill;
# LibreOffice honours it and paints nothing. That silently stripped the yellow
# "you type here" signal from thousands of input cells on every sheet this
# stage built, while the cells inherited from earlier stages kept it. Both
# engines agree once the alpha byte is explicit.
_fixed_fills = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            fl = cell.fill
            if not (fl and fl.patternType == "solid"):
                continue
            changed = False
            for attr in ("fgColor", "bgColor"):
                col = getattr(fl, attr, None)
                rgb = getattr(col, "rgb", None) if col is not None else None
                if isinstance(rgb, str) and len(rgb) == 8 and rgb[:2] == "00":
                    changed = True
            if changed:
                _fg = fl.fgColor.rgb
                cell.fill = PatternFill("solid", fgColor="FF" + _fg[2:])
                _fixed_fills += 1
print(f"fill alpha normalised on {_fixed_fills} cells")

# --- a locked sheet blocks formatting, filtering and sorting by default:
# permit the three that cannot damage a formula.
for ws in wb.worksheets:
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.autoFilter = False
    ws.protection.sort = False

# Formula lock: the six remaining open sheets get protected too (with their
# genuine input cells unlocked first), then EVERY sheet takes the owner's
# password so nobody can silently change a formula, and the workbook
# structure is locked against sheet deletion/renaming.
_sh_unlock = (["B18", "C18", "B19", "B20", "B69"]
              + [f"B{r}" for r in range(25, 33)]
              + [f"B{r}" for r in range(55, 65)]
              + [f"B{r}" for r in range(104, 114)])
for _co in _sh_unlock:
    _sh8[_co].protection = Protection(locked=False)
_sh8.protection.sheet = True
_mb9 = wb["Message Bank"]
for _r in range(4, 21):
    for _c in range(1, 10):
        _mb9.cell(_r, _c).protection = Protection(locked=False)
_mb9.protection.sheet = True
for _wsn in ("LinkedIn Playbook", "How-To Guides", "Benchmarks"):
    wb[_wsn].protection.sheet = True
_ls9 = wb["Lists"]
# Lists is the dropdown master for the whole workbook. Renaming one seeded
# value silently desynchronises every row already logged from it, so the
# seeded values are locked and only the empty extension rows stay open.
for _c in range(1, 26):
    _seed_end = 3
    for _r in range(4, 251):
        if _ls9.cell(_r, _c).value not in (None, ""):
            _seed_end = _r
    for _r in range(4, 251):
        _ls9.cell(_r, _c).protection = Protection(locked=(_r <= _seed_end))
_ls9.protection.sheet = True

# Audit finding: the blanket unlock passes on Message Bank and Lists also
# unlocked their formula cells, so a stray paste could destroy a character
# count or the roster selector. Re-lock every formula EXCEPT those inside the
# documented example rows (those must stay unlocked so the row can be deleted).
# (the logs no longer ship example rows, so every formula in the workbook is
# locked — there is nothing left that a user is meant to delete)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.protection = Protection(locked=True)

from openpyxl.workbook.protection import WorkbookProtection
# The sheet-protection password is a guard rail against accidental formula
# edits, not a security control — but it is the owner's, so it is passed in at
# build time and never committed. Export PCI_XLSX_PASSWORD before building.
_PWD = os.environ.get("PCI_XLSX_PASSWORD")
if not _PWD:
    raise SystemExit("set PCI_XLSX_PASSWORD before building (see growth-os/README.md)")
for ws in wb.worksheets:
    if ws.protection.sheet:
        ws.protection.password = _PWD
wb.security = WorkbookProtection(workbookPassword=_PWD, lockStructure=True)

wb.save(OUT)
print("v8 stage B saved:", OUT, f"({NP} platforms in register)")
