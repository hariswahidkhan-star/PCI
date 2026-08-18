#!/usr/bin/env python3
"""Excel-integrity checks: the things that trigger a repair prompt or silently
drop features when the file is opened in real Excel.
"""
import re
import sys
import zipfile
from collections import defaultdict
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import range_boundaries

SRC = sys.argv[1] if len(sys.argv) > 1 else "PCI_AI_Growth_OS_FINAL.xlsx"
issues = []

# 1) every part must be well-formed XML
with zipfile.ZipFile(SRC) as z:
    names = z.namelist()
    for n in names:
        if n.endswith(".xml") or n.endswith(".rels"):
            try:
                ET.fromstring(z.read(n))
            except ET.ParseError as e:
                issues.append(("BLOCKER", n, f"malformed XML: {e}"))
    required = ["[Content_Types].xml", "xl/workbook.xml", "xl/styles.xml"]
    for r in required:
        if r not in names:
            issues.append(("BLOCKER", r, "missing required part"))
print(f"parts checked: {len(names)}")

wb = openpyxl.load_workbook(SRC)

# 2) defined names / print areas must resolve
for ws in wb.worksheets:
    pa = ws.print_area
    if pa:
        for ref in (pa if isinstance(pa, list) else [pa]):
            if ws.title.replace("'", "''") not in str(ref) and "!" in str(ref):
                issues.append(("MAJOR", ws.title, f"print area points elsewhere: {ref}"))

# 3) merged ranges must not overlap (Excel drops or repairs)
for ws in wb.worksheets:
    boxes = []
    for m in ws.merged_cells.ranges:
        b = range_boundaries(str(m))
        for prev, prng in boxes:
            if not (b[2] < prev[0] or b[0] > prev[2] or b[3] < prev[1] or b[1] > prev[3]):
                issues.append(("MAJOR", ws.title, f"merged ranges overlap: {m} and {prng}"))
        boxes.append((b, str(m)))

# 4) conditional formatting ranges must be valid and non-empty
cf_count = 0
for ws in wb.worksheets:
    for rng in ws.conditional_formatting:
        cf_count += 1
        try:
            for part in str(rng.sqref).split():
                range_boundaries(part)
        except Exception as e:
            issues.append(("MAJOR", ws.title, f"bad CF range {rng.sqref}: {e}"))
        for rule in rng.rules:
            if rule.type == "expression" and not rule.formula:
                issues.append(("MAJOR", ws.title, f"CF rule with no formula at {rng.sqref}"))

# 5) data validations: valid ranges, no duplicate coverage with conflicting rules
dv_count = 0
for ws in wb.worksheets:
    cover = defaultdict(list)
    for dv in ws.data_validations.dataValidation:
        dv_count += 1
        for part in str(dv.sqref).split():
            try:
                b = range_boundaries(part)
            except Exception as e:
                issues.append(("MAJOR", ws.title, f"bad DV range {part}: {e}"))
                continue
            cover[(dv.type, part)].append(dv.formula1)
        if dv.type == "list" and not dv.formula1:
            issues.append(("MAJOR", ws.title, f"list DV with no source at {dv.sqref}"))
    for (typ, part), formulas in cover.items():
        if len(set(map(str, formulas))) > 1:
            issues.append(("MAJOR", ws.title,
                           f"{len(formulas)} conflicting {typ} DVs on the same range {part}"))

# 6) hyperlinks must have a target or location
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            h = c.hyperlink
            if h is not None and not (h.target or h.location):
                issues.append(("MAJOR", ws.title, f"hyperlink with no destination at {c.coordinate}"))

# 7) sheet names: Excel limits and illegal characters
for name in wb.sheetnames:
    if len(name) > 31:
        issues.append(("BLOCKER", name, "sheet name over 31 characters"))
    if re.search(r"[\[\]\*\?/\\:]", name):
        issues.append(("BLOCKER", name, "illegal character in sheet name"))

# 8) formulas: unbalanced parentheses / quotes would break on open
bad_formula = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                if v.count("(") != v.count(")") or v.count('"') % 2:
                    bad_formula += 1
                    if bad_formula <= 5:
                        issues.append(("BLOCKER", f"{ws.title}!{c.coordinate}",
                                       "unbalanced parentheses/quotes in formula"))
                if len(v) > 8192:
                    issues.append(("MAJOR", f"{ws.title}!{c.coordinate}", "formula over 8192 chars"))

# 9) cell character limit (32,767) — the Article Bank prompts are long
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and len(c.value) > 32767:
                issues.append(("BLOCKER", f"{ws.title}!{c.coordinate}", "cell text over 32,767 chars"))

print(f"CF rules: {cf_count} | DVs: {dv_count} | formula defects: {bad_formula}")
print("\n=========== EXCEL INTEGRITY ===========")
if not issues:
    print("CLEAN — no repair-prompt triggers found")
for sev in ("BLOCKER", "MAJOR"):
    for s, w, m in issues:
        if s == sev:
            print(f"[{s}] {w}: {m}")
print(f"TOTAL: {len(issues)}")
