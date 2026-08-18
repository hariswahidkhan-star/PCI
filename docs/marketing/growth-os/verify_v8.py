#!/usr/bin/env python3
"""End-to-end test of V7: simulate a three-person team, recalculate, assert.

Proves the thing the user asked about: that the workbook genuinely works for a
whole team, not one employee — plus the blocker fixes (UTM, duplicates, targets).
"""
import datetime as dt
import shutil
import subprocess
import sys

import openpyxl

SRC = "PCI_AI_Growth_OS_V9.xlsx"
TST = "v8_team_test.xlsx"
RECALC = "/root/.claude/skills/synced/xlsx/scripts/recalc.py"

shutil.copy(SRC, TST)
wb = openpyxl.load_workbook(TST)

TEAM = ["Aisha Khan", "Bilal Ahmed", "Clara Torres"]
sh = wb["START HERE"]
for i, name in enumerate(TEAM):
    sh.cell(55 + i, 2).value = name          # roster slots 1-3

today = dt.date.today()   # real date so the Weekly Pulse windows contain the data
de = wb["DAILY ENTRY"]
# example rows left in place: openpyxl's delete_rows shifts cells without adjusting
# formulas (Excel adjusts both), so the honest simulation writes below them — which
# is exactly what the counted ranges are designed for. All rows dated today so
# every entry falls inside the current Monday-anchored pulse week.
row = 7
for d in range(3):
    for i, name in enumerate(TEAM):
        de.cell(row, 1).value = today
        de.cell(row, 2).value = name
        de.cell(row, 3).value = "LinkedIn Sales Navigator"
        de.cell(row, 4).value = "Lead researched (Sales Navigator)"
        de.cell(row, 5).value = "Built lead list"
        de.cell(row, 6).value = 10 + i
        de.cell(row, 7).value = 30
        de.cell(row, 9).value = "Done"
        de.cell(row, 13).value = "Honorary Certification Outreach"
        de.cell(row, 14).value = "PCI AI - Institute (umbrella)"
        row += 1
        de.cell(row, 1).value = today
        de.cell(row, 2).value = name
        de.cell(row, 3).value = "LinkedIn Company Page"
        de.cell(row, 4).value = "Post / content published"
        de.cell(row, 5).value = "Published post"
        de.cell(row, 6).value = 1
        de.cell(row, 7).value = 20
        de.cell(row, 9).value = "Done"
        de.cell(row, 13).value = "Content & SEO Growth"
        de.cell(row, 14).value = "PCL-AI certification"
        row += 1
# one email campaign (new direct-channel activity type)
de.cell(row, 1).value = today
de.cell(row, 2).value = TEAM[0]
de.cell(row, 3).value = "Email Marketing (ESP)"
de.cell(row, 4).value = "Email campaign sent"
de.cell(row, 5).value = "Monthly newsletter to the full list"
de.cell(row, 6).value = 500
de.cell(row, 7).value = 30
de.cell(row, 9).value = "Done"
de.cell(row, 13).value = "General Brand Awareness"
de.cell(row, 14).value = "All / shared"
row += 1

lo = wb["LinkedIn Outreach"]
r = 5
for i, name in enumerate(TEAM):
    for k in range(4):
        lo.cell(r, 1).value = today
        lo.cell(r, 2).value = name
        lo.cell(r, 3).value = f"Lead {i}-{k}"
        lo.cell(r, 4).value = f"https://www.linkedin.com/in/lead-{i}-{k}/"
        lo.cell(r, 5).value = "Acme Projects"
        lo.cell(r, 8).value = "PMO Leader"
        lo.cell(r, 10).value = 10
        lo.cell(r, 13).value = "Yes"                       # connection sent
        lo.cell(r, 14).value = "Yes" if k < 3 else "No"    # accepted
        lo.cell(r, 18).value = "Yes" if k < 2 else "No"    # message sent
        lo.cell(r, 31).value = today                        # message sent date
        if k == 0:
            lo.cell(r, 19).value = "Meeting Booked"
            lo.cell(r, 33).value = today      # meeting date (AG) — feeds Weekly Pulse
        elif k == 1:
            lo.cell(r, 19).value = "Interested"
        lo.cell(r, 41).value = "Honorary Certification Outreach"
        lo.cell(r, 42).value = "PCI AI - Institute (umbrella)"
        if k == 0:
            lo.cell(r, 37).value = today      # purchase date (AK)
            lo.cell(r, 38).value = 450        # revenue on the converted-track lead
        r += 1
# duplicate URL on purpose: two employees log the same lead
lo.cell(r, 1).value = today
lo.cell(r, 2).value = TEAM[2]
lo.cell(r, 3).value = "Lead 0-0 again"
lo.cell(r, 4).value = "https://www.linkedin.com/in/lead-0-0/"

cc_sheet = wb["Content Calendar"]
cc_sheet.cell(5, 1).value = today
cc_sheet.cell(5, 2).value = TEAM[0]
cc_sheet.cell(5, 3).value = "LinkedIn Company Page"
cc_sheet.cell(5, 4).value = "Carousel"
cc_sheet.cell(5, 5).value = "5 EVM formulas explained"
cc_sheet.cell(5, 6).value = "Content & SEO Growth"
cc_sheet.cell(5, 10).value = "Published"
cc_sheet.cell(5, 11).value = today               # published date
cc_sheet.cell(5, 19).value = "PCI AI - Institute (umbrella)"

cs = wb["Content Scheduler"]
cs.cell(5, 1).value = "LinkedIn Company Page"
cs.cell(5, 2).value = "PCI AI - Institute (umbrella)"
cs.cell(5, 3).value = "Content & SEO Growth"
cs.cell(5, 4).value = "3x Weekly"
cs.cell(5, 5).value = 1
cs.cell(5, 6).value = "Carousel"
cs.cell(5, 7).value = "Native scheduler"
cs.cell(5, 8).value = today - dt.timedelta(days=7)
cs.cell(5, 9).value = today + dt.timedelta(days=20)   # 28-day window
cs.cell(5, 13).value = TEAM[0]
cs.cell(5, 14).value = "Active"

lb = wb["Link Building"]
lb.cell(5, 1).value = today
lb.cell(5, 2).value = TEAM[0]
lb.cell(5, 3).value = "Unlinked mention reclaimed"
lb.cell(5, 4).value = "projecttimes.com"
lb.cell(5, 6).value = "https://projectcontrolsinstitute.org/"
lb.cell(5, 9).value = "Earned - live"
lb.cell(5, 10).value = "Dofollow"
lb.cell(5, 11).value = today                # went live this week
lb.cell(5, 12).value = "Authority & Entity Building"
lb.cell(5, 13).value = "PCI AI - Institute (umbrella)"

jp = wb["Job Postings"]
jp.cell(5, 1).value = today
jp.cell(5, 2).value = TEAM[1]
jp.cell(5, 3).value = "Marketing Executive (Remote)"
jp.cell(5, 4).value = "Marketing"
jp.cell(5, 5).value = "PCI AI - Institute (umbrella)"
jp.cell(5, 6).value = "LinkedIn Company Page"
jp.cell(5, 7).value = "Remote / Global"
jp.cell(5, 9).value = "Open"
jp.cell(5, 10).value = 12

utm = wb["UTM Builder"]
utm.cell(5, 1).value = "https://projectcontrolsinstitute.org/pcl-ai.html"
utm.cell(5, 2).value = "linkedin"
utm.cell(5, 3).value = "social"
utm.cell(5, 4).value = "pcl-aug26"

wb.save(TST)
res = subprocess.run(["python3", RECALC, TST, "900"], capture_output=True, text=True)
print("recalc:", res.stdout.strip()[:300])

wb = openpyxl.load_workbook(TST, data_only=True)
fails = []
def check(label, actual, expect):
    ok = actual == expect
    print(("PASS" if ok else "FAIL"), label, "->", repr(actual), "expected", repr(expect))
    if not ok:
        fails.append(label)

ts = wb["Team Scorecard"]
names = [ts.cell(rr, 1).value for rr in (4, 5, 6)]
check("Team Scorecard lists all three", names, TEAM)
check("Aisha leads researched (3 days x 10)", ts.cell(4, 3).value, 30)
check("Bilal leads researched (3 days x 11)", ts.cell(5, 3).value, 33)
check("Clara connections sent", ts.cell(6, 4).value, 4)
check("Aisha accepted", ts.cell(4, 5).value, 3)
check("Aisha meetings booked", ts.cell(4, 10).value, 1)

wdw = wb["Who Did What"]
hdr = [wdw.cell(3, c).value for c in (2, 3, 4)]
check("Who Did What columns carry all three", hdr, TEAM)
sn_row = next(rr for rr in range(4, 76) if wdw.cell(rr, 1).value == "LinkedIn Sales Navigator")
check("Sales Nav entries split per person",
      [wdw.cell(sn_row, c).value for c in (2, 3, 4)], [3, 3, 3])

es = wb["Employee Score"]
check("Employee Score row for Clara exists", es.cell(6, 1).value, TEAM[2])

db = wb["Dashboard"]
check("Dashboard leads researched actual", db.cell(42, 2).value, 99)   # (10+11+12) x 3 days
lk = wb["LinkedIn Outreach"]
dupes = [lk.cell(rr, 40).value for rr in range(5, 18)]
check("Duplicate flagged at least twice",
      sum(1 for d in dupes if d and "DUPLICATE" in str(d)) >= 2, True)
u5 = wb["UTM Builder"].cell(5, 6).value
check("UTM link uses ? and utm_campaign",
      isinstance(u5, str) and "?utm_source=linkedin" in u5 and "utm_campaign=pcl-aug26" in u5, True)
exp = db.cell(42, 3).value
check("Expected leads = 30/day x days x 3 people", exp, 30 * db.cell(41, 2).value * 3)

op = wb["Objective Performance"]
check("Objective: honorary activities (9 research sessions)", op.cell(4, 2).value, 9)
check("Objective: honorary minutes (9 x 30)", op.cell(4, 3).value, 270)
check("Objective: content minutes (9 x 20)", op.cell(9, 3).value, 180)
check("Objective: honorary leads logged", op.cell(4, 6).value, 12)
check("Objective: honorary meetings booked", op.cell(4, 7).value, 3)
check("Objective: honorary revenue (3 x 450)", op.cell(4, 8).value, 1350)
check("Objective: untagged lead flagged (the duplicate row)", op.cell(15, 6).value, 1)
check("Objective: TOTAL minutes (450 + 30 email)", op.cell(16, 3).value, 480)
# Rows are found by their LABEL, never by a fixed number — the same discipline
# the Dashboard tiles now use. A row inserted above a block must never be able
# to make this suite pass while reading the wrong line.
def oprow_prefix(prefix, after=1):
    for rr in range(after, 60):
        if str(op.cell(rr, 1).value or "").strip().startswith(prefix):
            return rr
    raise AssertionError(f"Objective Performance row not found: {prefix!r}*")


def oprow(label, after=1):
    for rr in range(after, 60):
        if str(op.cell(rr, 1).value or "").strip() == label:
            return rr
    raise AssertionError(f"Objective Performance row not found: {label!r}")

_bfirst = oprow("PCI AI - Institute (umbrella)", after=15)
check("Brand: institute (umbrella) minutes", op.cell(_bfirst, 3).value, 270)
check("Brand: PCL-AI minutes", op.cell(oprow("PCL-AI certification", _bfirst), 3).value, 180)
check("Brand: All/shared minutes (the email row)", op.cell(oprow("All / shared", _bfirst), 3).value, 30)
check("Brand: institute leads logged", op.cell(_bfirst, 6).value, 12)
check("Brand: institute revenue", op.cell(_bfirst, 8).value, 1350)
check("Brand: untagged lead flagged (the duplicate row)", op.cell(oprow_prefix("(no brand set", _bfirst), 6).value, 1)
_btot = oprow("TOTAL", oprow_prefix("(no brand set", _bfirst))
check("Brand: TOTAL minutes", op.cell(_btot, 3).value, 480)
_mfirst = oprow("Honorary Certification Outreach", _btot)
check("Objective matrix: Aisha honorary minutes", op.cell(_mfirst, 2).value, 90)
check("Objective matrix: Aisha content minutes",
      op.cell(oprow("Content & SEO Growth", _mfirst), 2).value, 60)
check("Objective matrix: Aisha total minutes (incl email 30)",
      op.cell(oprow("TOTAL", oprow_prefix("(no objective set", _mfirst)), 2).value, 180)

pu = wb["Weekly Pulse"]
check("Pulse: minutes this week", pu.cell(6, 2).value, 480)
check("Pulse: activities this week", pu.cell(7, 2).value, 19)
check("Pulse: leads researched this week", pu.cell(8, 2).value, 99)
check("Pulse: email recipients this week", pu.cell(13, 2).value, 500)
check("Pulse: job posts this week", pu.cell(15, 2).value, 1)
check("Pulse: backlinks gone live this week", pu.cell(16, 2).value, 1)
check("Pulse: meetings booked this week (meeting dates)", pu.cell(17, 2).value, 3)
check("Pulse: revenue this week (purchase dates)", pu.cell(18, 2).value, 1350)
check("Pulse: new leads this week", pu.cell(19, 2).value, 13)
check("Pulse: last week minutes is zero", pu.cell(6, 3).value, 0)
check("Pulse: Aisha minutes this week (person block)", pu.cell(23, 2).value, 180)

sched = wb["Content Scheduler"]
check("Scheduler: planned posts (28 days at 3x weekly)", sched.cell(5, 10).value, 12)
check("Scheduler: published in window (from Content Calendar)", sched.cell(5, 11).value, 1)
check("Scheduler: coverage computes", round(sched.cell(5, 12).value, 3), round(1 / 12, 3))
def tile(label):
    for rr in range(3, 34):
        if str(db.cell(rr, 6).value or "") == label:
            return db.cell(rr + 1, 6).value
    return "TILE NOT FOUND"

check("Dashboard tile: active schedules", tile("ACTIVE SCHEDULES"), 1)
check("Dashboard tile: revenue this week (not meetings!)", tile("REVENUE THIS WEEK"), 1350)
check("Dashboard tile: meetings this week (not backlinks!)", tile("MEETINGS THIS WEEK"), 3)
check("Dashboard tile: backlinks this week", tile("BACKLINKS THIS WEEK"), 1)
check("Dashboard tile: content this week", tile("CONTENT THIS WEEK"), 1)
check("Dashboard tile: minutes this week", tile("MINUTES THIS WEEK"), 480)

print("\n" + ("ALL PASS" if not fails else f"{len(fails)} FAILURES: {fails}"))
sys.exit(1 if fails else 0)
