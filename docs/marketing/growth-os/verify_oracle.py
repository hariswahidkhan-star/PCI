#!/usr/bin/env python3
"""Regression suite for the independent numerical audit's findings.

Every case below is the exact adversarial input that exposed a defect. The
assertions state the FIXED behaviour, so a future edit that reintroduces any
of them fails here rather than in a manager's hands.

Findings covered: 3 blockers (revenue definitions, rates past 100% rewarding
under-logging, meetings erased by their own success) and the majors —
partial-week average, scheduler multi-count, negative planned posts,
Objective Performance leaks, phantom dropdown rows, text dates, future dates,
Weekly Review's lost Sunday and pre-start blindness, person-days, the
double-weighted Employee Score component, text in the contract-signed cell,
and unattributed minutes.
"""
import datetime as dt
import shutil
import subprocess
import sys

import openpyxl

SRC = "PCI_AI_Growth_OS_V9.xlsx"
TST = "oracle_test.xlsx"
RECALC = "/root/.claude/skills/synced/xlsx/scripts/recalc.py"

shutil.copy(SRC, TST)
wb = openpyxl.load_workbook(TST)

TEAM = ["Aisha Khan", "Bilal Ahmed"]
sh = wb["START HERE"]
for i, n in enumerate(TEAM):
    sh.cell(55 + i, 2).value = n
today = dt.date.today()
monday = today - dt.timedelta(days=(today.weekday()))
LAST_MON = monday - dt.timedelta(days=7)
LAST_SUN = monday - dt.timedelta(days=1)
# the programme started LAST Monday, so Weekly Review row 4 is last week and
# row 5 is this one — that lets the Sunday-with-a-time case sit in the past
sh.cell(20, 2).value = LAST_MON

de = wb["DAILY ENTRY"]
lo = wb["LinkedIn Outreach"]
cc = wb["Content Calendar"]
pp = wb["Partnership Pipeline"]
cs = wb["Content Scheduler"]

# ---------------------------------------------------------------- DAILY ENTRY
r = 7


def d_entry(date, name, plat, kind, howmany, minutes, obj=None, brand=None):
    global r
    de.cell(r, 1).value = date
    de.cell(r, 2).value = name
    de.cell(r, 3).value = plat
    de.cell(r, 4).value = kind
    de.cell(r, 5).value = "test row"
    de.cell(r, 6).value = howmany
    de.cell(r, 7).value = minutes
    de.cell(r, 9).value = "Done"
    if obj:
        de.cell(r, 13).value = obj
    if brand:
        de.cell(r, 14).value = brand
    r += 1


# two clean days for both people (person-days = 4, calendar days = 2)
for day in (monday, monday + dt.timedelta(days=1)):
    for n in TEAM:
        d_entry(day, n, "LinkedIn Sales Navigator",
                "Lead researched (Sales Navigator)", 10, 60,
                "Honorary Certification Outreach", "PCI AI - Institute (umbrella)")
CLEAN_MINUTES = 4 * 60
CLEAN_LEADS = 4 * 10
CLEAN_PERSON_DAYS = 4

# finding 13: the last day of a week, with a time of day. Week end is a bare
# date, so "<=" meant Sunday 00:00 and every Sunday afternoon fell out of its
# own week — counted by Weekly Pulse, invisible to Weekly Review.
d_entry(dt.datetime.combine(LAST_SUN, dt.time(16, 30)), TEAM[0], "Medium",
        "Lead researched (Sales Navigator)", 7, 30,
        "Honorary Certification Outreach", "PCI AI - Institute (umbrella)")
SUNDAY_LEADS, SUNDAY_MIN = 7, 30

# finding 18/19: minutes logged under a name that is not on the roster
d_entry(monday, "aisha khan ", "Medium", "Post / content published", 1, 25,
        "Content & SEO Growth", "PCI AI - Institute (umbrella)")
UNATTRIBUTED = 25

# finding 7: minutes with NO name and NO objective — used to vanish entirely
de.cell(r, 1).value = monday
de.cell(r, 5).value = "orphan row, no name, no tag"
de.cell(r, 7).value = 66
r += 1
ORPHAN = 66

# finding 10: a date typed as TEXT
de.cell(r, 1).value = str(monday)                  # text, not a date
de.cell(r, 2).value = TEAM[0]
de.cell(r, 4).value = "Lead researched (Sales Navigator)"
de.cell(r, 6).value = 100
de.cell(r, 7).value = 120
r += 1
TEXT_MIN = 120

# finding 7 again: minutes with NO date at all — the reconciliation row is the
# only thing on the page that can reveal them
de.cell(r, 7).value = 44
de.cell(r, 5).value = "no date at all"
r += 1
NO_DATE = 44

# finding 11: a wrong-year typo
de.cell(r, 1).value = monday.replace(year=monday.year + 1)
de.cell(r, 2).value = TEAM[1]
de.cell(r, 4).value = "Lead researched (Sales Navigator)"
de.cell(r, 6).value = 999
de.cell(r, 7).value = 480
r += 1
FUTURE_MIN = 480

# finding 9: a stray dropdown click on an otherwise empty row, on both the
# DAILY ENTRY and the Outreach log. Objective/brand are dropdowns; pre-clicking
# one is ordinary behaviour and must never create a record.
de.cell(60, 13).value = "Honorary Certification Outreach"
de.cell(60, 14).value = "PCI AI - Institute (umbrella)"
lo.cell(80, 41).value = "Honorary Certification Outreach"
lo.cell(80, 42).value = "PCI AI - Institute (umbrella)"

# ------------------------------------------------------------ OUTREACH
lr = 5


def lead(name, sent, accepted, msg, outcome, meeting=None, revenue=None,
         purchase=None, owner=TEAM[0]):
    global lr
    lo.cell(lr, 1).value = monday
    lo.cell(lr, 2).value = owner
    lo.cell(lr, 3).value = name
    lo.cell(lr, 4).value = f"https://www.linkedin.com/in/{name.replace(' ', '-')}/"
    lo.cell(lr, 13).value = sent
    lo.cell(lr, 14).value = accepted
    lo.cell(lr, 18).value = msg
    lo.cell(lr, 19).value = outcome
    if meeting:
        lo.cell(lr, 33).value = meeting
    if revenue:
        lo.cell(lr, 38).value = revenue
        lo.cell(lr, 37).value = purchase or monday
    lo.cell(lr, 41).value = "Honorary Certification Outreach"
    lo.cell(lr, 42).value = "PCI AI - Institute (umbrella)"
    lr += 1


# 4 requests sent, 3 accepted, 3 messaged
lead("Clean One", "Yes", "Yes", "Yes", "Awaiting Reply")
lead("Clean Two", "Yes", "Yes", "Yes", "Interested")
# finding 3: a meeting that PROGRESSED — outcome no longer says "Meeting Booked"
lead("Progressed", "Yes", "Yes", "Yes", "Converted", meeting=monday,
     revenue=1200, purchase=monday)
lead("Quiet", "Yes", "No", "No", "No Response")
# finding 2: accepted without a logged request — the commonest logging slip
lead("Slip", "No", "Yes", "No", "Awaiting Reply")
# a meeting still at the meeting stage
lead("At Meeting", "Yes", "Yes", "Yes", "Meeting Booked", meeting=monday)

REQUESTS = 5          # Clean One, Clean Two, Progressed, Quiet, At Meeting
ACCEPTED = 5          # the four above minus Quiet, plus Slip
MEETINGS_BY_DATE = 2  # Progressed + At Meeting
TAGGED_LEADS = 6      # every real lead row carries the objective tag
CERT_REVENUE = 1200

# ------------------------------------------------------------ PARTNERSHIPS
pp.cell(5, 1).value = monday
pp.cell(5, 3).value = "Real Deal Ltd"
pp.cell(5, 22).value = 20000
pp.cell(5, 23).value = monday                      # a real signed date
pp.cell(5, 24).value = "Partnerships & PR"
pp.cell(5, 25).value = "PCI AI - Institute (umbrella)"
# finding 16: a typed note in the signed-date cell used to book the revenue
pp.cell(6, 1).value = monday
pp.cell(6, 3).value = "Not Signed Yet Ltd"
pp.cell(6, 22).value = 9000
pp.cell(6, 23).value = "signed - awaiting countersign"
pp.cell(6, 24).value = "Partnerships & PR"
pp.cell(6, 25).value = "PCI AI - Institute (umbrella)"
SIGNED_VALUE = 20000
OPEN_PIPELINE = 9000

# ------------------------------------------------------------ CONTENT
cr = 5
for i in range(5):                                  # 5 real published posts
    cc.cell(cr, 1).value = monday
    cc.cell(cr, 3).value = "LinkedIn Company Page"
    cc.cell(cr, 5).value = f"Post {i}"
    cc.cell(cr, 6).value = "Content & SEO Growth"
    cc.cell(cr, 10).value = "Published"
    cc.cell(cr, 11).value = monday
    cc.cell(cr, 19).value = "PCI AI - Institute (umbrella)"
    cr += 1
PUBLISHED = 5

# ------------------------------------------------------------ SCHEDULER
# finding 5: three schedules on the same platform used to count the same five
# posts twelve times. Only the one matching brand AND objective may claim them.
cs.cell(5, 1).value = "LinkedIn Company Page"
cs.cell(5, 2).value = "PCI AI - Institute (umbrella)"
cs.cell(5, 3).value = "Content & SEO Growth"
cs.cell(5, 4).value = "3x Weekly"
cs.cell(5, 5).value = 1
cs.cell(5, 8).value = monday
cs.cell(5, 9).value = monday + dt.timedelta(days=6)
cs.cell(6, 1).value = "LinkedIn Company Page"
cs.cell(6, 2).value = "PCL-AI certification"
cs.cell(6, 3).value = "Certification Sales - PCL-AI"
cs.cell(6, 4).value = "Weekly"
cs.cell(6, 5).value = 1
cs.cell(6, 8).value = monday
cs.cell(6, 9).value = monday + dt.timedelta(days=6)
# finding 6: end date before start date
cs.cell(7, 1).value = "Medium"
cs.cell(7, 4).value = "Weekly"
cs.cell(7, 5).value = 1
cs.cell(7, 8).value = monday + dt.timedelta(days=10)
cs.cell(7, 9).value = monday
# finding 21: a monthly cadence over exactly one year
cs.cell(8, 1).value = "Telegram Channel"
cs.cell(8, 4).value = "Monthly"
cs.cell(8, 5).value = 1
cs.cell(8, 8).value = dt.date(2027, 1, 1)
cs.cell(8, 9).value = dt.date(2027, 12, 31)

wb.save(TST)
res = subprocess.run([sys.executable, RECALC, TST, "1200"],
                     capture_output=True, text=True)
print("recalc:", res.stdout.strip()[:400])
if '"status": "success"' not in res.stdout:
    sys.exit("recalc failed")

wv = openpyxl.load_workbook(TST, data_only=True)
db, wp, op, wr, es, ts = (wv["Dashboard"], wv["Weekly Pulse"],
                          wv["Objective Performance"], wv["Weekly Review"],
                          wv["Employee Score"], wv["Team Scorecard"])
csv_ = wv["Content Scheduler"]

fails = []


def check(name, got, want, tol=0.001):
    ok = (abs(got - want) <= tol) if isinstance(want, (int, float)) and \
        isinstance(got, (int, float)) else got == want
    print(("PASS " if ok else "FAIL ") + name, "->", got,
          "" if ok else f"expected {want}")
    if not ok:
        fails.append(name)


def label_row(ws, text, col=1, lo_=1, hi=140):
    for rr in range(lo_, hi):
        if str(ws.cell(rr, col).value or "").strip() == text:
            return rr
    raise AssertionError(f"{ws.title}: row not found {text!r}")


def dbv(label):
    return db.cell(label_row(db, label), 2).value


def tile(label):
    for rr in range(3, 40):
        if str(db.cell(rr, 6).value or "") == label:
            return db.cell(rr + 1, 6).value
    raise AssertionError(f"tile not found {label!r}")


print("\n--- BLOCKER 1: one revenue definition across the tiles")
check("Enterprise value signed ignores a typed note",
      dbv("Enterprise value signed"), SIGNED_VALUE)
check("Open pipeline holds the unsigned deal",
      dbv("Enterprise pipeline value (not yet signed)"), OPEN_PIPELINE)
check("Total revenue = certifications + signed enterprise",
      dbv("TOTAL REVENUE RECORDED  (certifications + signed enterprise)"),
      CERT_REVENUE + SIGNED_VALUE)
check("REVENUE TO DATE tile reads the total, not one component",
      tile("REVENUE TO DATE"), CERT_REVENUE + SIGNED_VALUE)
check("revenue to date is never below revenue this week",
      tile("REVENUE TO DATE") >= tile("REVENUE THIS WEEK"), True)

print("\n--- BLOCKER 2: rates cannot exceed 100%, under-logging cannot pay")
check("acceptance rate stays within 100%", round(dbv("Acceptance rate"), 4),
      round(ACCEPTED / max(REQUESTS, ACCEPTED), 4))
check("acceptance rate <= 1", dbv("Acceptance rate") <= 1, True)
check("reply rate <= 1", dbv("Reply rate on messages sent") <= 1, True)
check("Team Scorecard acceptance % <= 1",
      all((ts.cell(rr, 6).value or 0) <= 1 for rr in range(4, 14)
          if isinstance(ts.cell(rr, 6).value, (int, float))), True)
check("the contradiction is reported, not hidden",
      dbv("Accepted without a logged connection request"), 1)

print("\n--- BLOCKER 3: a meeting survives its own success")
check("meetings to date counted by meeting date",
      dbv("Meetings booked to date (by meeting date)"), MEETINGS_BY_DATE)
check("cumulative meetings >= this week's meetings",
      tile("MEETINGS BOOKED") >= tile("MEETINGS THIS WEEK"), True)
check("stage count is labelled as a stage, not a total",
      db.cell(label_row(db, "Currently at meeting-booked stage"), 2).value, 1)

print("\n--- MAJOR 4: the 4-week average uses completed weeks only")
_e5 = wp.cell(5, 5).value
check("4-week avg header says completed weeks",
      "completed" in str(_e5).lower(), True)
# the window is the four COMPLETED weeks before this one, and the divisor is
# however many completed weeks the programme has actually run — so week 1 does
# not read as a quarter of itself. Only the pre-start Sunday row is in range.
check("4-week average covers completed weeks only, divided by weeks run",
      wp.cell(label_row(wp, "Minutes logged"), 5).value, SUNDAY_MIN)

print("\n--- MAJOR 5/6/21: scheduler counts its own posts, once")
check("schedule 1 claims only its own brand+objective posts",
      csv_.cell(5, 11).value, PUBLISHED)
check("schedule 2 (different brand) claims none", csv_.cell(6, 11).value, 0)
check("reversed dates produce no planned posts", csv_.cell(7, 10).value, None)
check("reversed dates produce no coverage", csv_.cell(7, 12).value, None)
check("monthly over a year plans 12, not 13", csv_.cell(8, 10).value, 12)

print("\n--- MAJOR 7/8/9: nothing leaks out of Objective Performance")
_otot = label_row(op, "TOTAL")
DATED_MINUTES = (CLEAN_MINUTES + SUNDAY_MIN + UNATTRIBUTED + ORPHAN
                 + TEXT_MIN + FUTURE_MIN)
check("TOTAL minutes include the nameless, untagged row",
      op.cell(_otot, 3).value, DATED_MINUTES)
check("the reconciliation row exposes the dateless row",
      op.cell(_otot + 1, 9).value, NO_DATE)
check("share of minutes is no longer tautologically 100%",
      round(op.cell(_otot, 9).value, 6) < 1.0, True)
check("a stray dropdown click on an empty row is not a lead",
      op.cell(4, 6).value, TAGGED_LEADS)
check("nor does it add minutes", op.cell(4, 3).value, CLEAN_MINUTES + SUNDAY_MIN)

print("\n--- MAJOR 10/11: bad dates are found, not silently absorbed")
check("text dates are reported", dbv("Rows with an unreadable (text) date"), 1)
check("future-dated rows are reported", dbv("Future-dated rows"), 1)
check("the dateless row is not mistaken for a text date",
      dbv("Rows with an unreadable (text) date"), 1)

print("\n--- MAJOR 12/13: Weekly Review sees the whole week and the whole record")
check("the last day of a week, at 16:30, is inside that week",
      wr.cell(4, 4).value, SUNDAY_LEADS)
check("this week's row holds this week's work", wr.cell(5, 4).value, CLEAN_LEADS)
check("a Before-programme-start row exists",
      label_row(wr, "Before programme start") > 0, True)
check("a TOTAL row exists", label_row(wr, "TOTAL (all weeks + before start)") > 0, True)

print("\n--- MAJOR 14: expected-to-date uses person-days")
_r42 = label_row(db, "Sales Navigator leads researched")
_dev = wv["DAILY ENTRY"]
_person_days = sum(_dev.cell(rr, 12).value or 0 for rr in range(7, 1007))
check("expected-to-date = daily target x person-days (not x headcount)",
      db.cell(_r42, 3).value, 30 * _person_days)
check("person-days are fewer than headcount x calendar days",
      _person_days < 2 * 400, True)

print("\n--- MAJOR 15: no two Employee Score components are the same number")
_esr = label_row(es, TEAM[0])
check("reply-rate and positive-reply components differ",
      es.cell(_esr, 4).value != es.cell(_esr, 5).value or
      es.cell(_esr, 4).value in (0, None), True)

print("\n--- MAJOR 18/19: per-person minutes reconcile to the team total")
_pt = None
for rr in range(6, 60):
    if str(wp.cell(rr, 1).value or "").startswith("TOTAL"):
        _pt = rr
        break
check("per-person block totals the team's minutes",
      wp.cell(_pt, 2).value, wp.cell(label_row(wp, "Minutes logged"), 2).value)
check("unattributed minutes are named",
      wp.cell(_pt - 1, 2).value, UNATTRIBUTED + ORPHAN)
check("unrecognised names are reported on the Dashboard",
      dbv("Minutes logged under a name not on the roster"), 1)

print("\n--- DATA HEALTH block is present and computing")
for _lab in ("Rows with an unreadable (text) date", "Future-dated rows",
             "Accepted without a logged connection request",
             "Minutes logged under a name not on the roster",
             "Content marked Published with no published date",
             "Published with a date but not marked published",
             "Signed deals with no deal value",
             "Duplicate leads flagged in the Outreach log",
             "Published content with no platform named"):
    check(f"health check present: {_lab}", isinstance(dbv(_lab), (int, float)), True)

print()
if fails:
    print(f"{len(fails)} FAILURES:", fails)
    sys.exit(1)
print("ALL PASS")
