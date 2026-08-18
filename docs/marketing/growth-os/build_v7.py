#!/usr/bin/env python3
"""Build PCI_AI_Growth_OS_V7.xlsx from the V6 original.

Applies the findings of the three audit agents: canonical platform LOVs, the
UTM/target/duplicate/follow-up blocker fixes, funnel-to-revenue columns,
sheet protection, operating-model documentation and the V7 upgrade-notes sheet.
Re-runnable: always starts from the untouched V6 copy.
"""
import copy as copymod
import shutil

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from canonical_lov import (ACTIVITY_TYPES, AREAS, BRANDS, CERTUVO_PLATFORMS,
                           CONTENT_TYPE, FREQUENCY, FUNNEL_STAGE, GEO, GEO_DEFAULT,
                           OBJECTIVES, LEAD_SEGMENT, ORG_TYPE, OUTCOME, PLATFORMS,
                           PLATFORM_VALUE_RANKS, RENAMES, SCORE_1_5)

SRC = "growth_os_v6.xlsx"
OUT = "PCI_AI_Growth_OS_V7.xlsx"

NP = len(PLATFORMS)                 # 71
SETUP_LAST = 3 + NP                 # Platform Setup data rows 4..74

YELLOW = PatternFill("solid", fgColor="FFF2CC")
HDRFONT = Font(name="Arial", bold=True, size=10)
NOTEFONT = Font(name="Arial", size=9, italic=True, color="666666")

wb = openpyxl.load_workbook(SRC)


def sweep_rename(ws, cols, first, last, mapping):
    """Replace platform strings in value cells (never formulas)."""
    changed = []
    for r in range(first, last + 1):
        for c in cols:
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, str) and not v.startswith("=") and v.strip() in mapping:
                cell.value = mapping[v.strip()]
                changed.append((ws.title, cell.coordinate, v))
    return changed


# ---------------------------------------------------------------- 1. Lists
ls = wb["Lists"]
def write_col(header_letter, values, note_row1=None):
    c = openpyxl.utils.column_index_from_string(header_letter)
    # clear old values well past previous extent
    for r in range(4, 120):
        ls.cell(r, c).value = None
    for i, v in enumerate(values):
        cell = ls.cell(4 + i, c)
        cell.value = v
        cell.font = Font(name="Arial", size=10)
    return f"Lists!${header_letter}$4:${header_letter}${3 + len(values)}"

R_PLAT   = write_col("J", [p[0] for p in PLATFORMS])
R_AREA   = write_col("E", AREAS)
R_OUT    = write_col("F", OUTCOME)
R_SEG    = write_col("G", LEAD_SEGMENT)
R_CTYPE  = write_col("H", CONTENT_TYPE)
R_ORG    = write_col("K", ORG_TYPE)
R_FREQ   = write_col("D", FREQUENCY)
R_SCORE  = write_col("T", SCORE_1_5)
R_STAGE  = write_col("S", FUNNEL_STAGE)
# Pipeline Type: align the three mismatched strings, keep as subset list
R_PIPE   = write_col("U", ["Employer / Enterprise", "University", "Professional Association",
                           "Influencer / Expert", "Industry Media", "Podcast", "Conference",
                           "Training Provider"])
# WR selector: guard the empty roster slots that showed as literal 0
for i in range(1, 11):
    ls.cell(4 + i, 18).value = (f"=IF('START HERE'!$B${54 + i}=\"\",\"\","
                                f"'START HERE'!$B${54 + i})")

R_OBJ = write_col("Y", OBJECTIVES)
ls.cell(3, 25).value = "Objective"
ls.cell(3, 25).font = HDRFONT
# Brand list becomes the full property list (institute / certifications /
# PCI World / Certuvo) so every effort states WHO it is for
R_BR = write_col("O", BRANDS)
ls.cell(3, 15).value = "Brand / property"
ls.cell(3, 15).font = HDRFONT
# activity types gain the direct channels + job posting
R_ACT = write_col("P", ACTIVITY_TYPES)

# Platform notes column: dedup rule beside each platform (new col W)
ls.cell(3, 23).value = "Platform logging rule (dedup)"
ls.cell(3, 23).font = HDRFONT
for i, (name, area, prio, new, note) in enumerate(PLATFORMS):
    ls.cell(4 + i, 23).value = note or ""
    ls.cell(4 + i, 23).font = NOTEFONT

# ------------------------------------------------- 2. Platform Setup rebuild
ps = wb["Platform Setup"]
old = {}
for r in range(4, 50):
    name = ps.cell(r, 2).value
    if name:
        old[str(name).strip()] = [ps.cell(r, c).value for c in range(1, 15)]

SETUP_STEPS = {
    "LinkedIn Live": "1) Verify Live access on the Company Page  2) Test a private stream with StreamYard  3) Schedule the first session as a LinkedIn Event",
    "Bluesky": "1) Create account, secure the @pci handle  2) Complete profile with site link  3) Watching brief - post only repurposed content",
    "Planning Planet": "1) Create a named personal account (no brand accounts)  2) Complete profile honestly, employer = PCI  3) Read forum rules; answer first, promote second",
    "PMI Community (ProjectManagement.com)": "1) Create personal account  2) Complete profile  3) Contribute answers/articles; disclose PCI affiliation",
    "AACE Communities": "1) Personal account  2) Disclose affiliation - peer body, engage respectfully  3) Answer technical questions only",
    "Project Management Stack Exchange": "1) Personal account  2) Earn 50 reputation before adding profile links  3) Answers must stand alone without PCI links",
    "GitHub": "1) Only if Certuvo open-sources tooling  2) Organisation account + README linking to Certuvo",
    "Credly / Digital Badges": "1) Open issuer account  2) Design badge templates for PCL-AI / PFL-AI / PML-AI  3) Wire issue-on-conferral into the credential workflow  4) Test one badge share to LinkedIn",
    "Course Aggregators (findcourses etc.)": "1) List PCL-AI on findcourses.co.uk and equivalents  2) Track referral traffic with UTM links",
    "CPD Directory Listings": "1) Identify CPD directories accepting certification bodies  2) List only what is factually accurate - no accreditation claims",
    "Wikipedia / Wikidata": "1) NEVER self-edit an article  2) Create a Wikidata item with sourced facts only  3) Revisit when independent press coverage exists",
    "Product Hunt": "1) Certuvo launch only  2) Prepare assets and a launch-day plan  3) One-off - do not maintain",
    "Zoom Webinars": "1) Confirm webinar licence  2) Brand the registration page  3) Template: registration -> reminder -> replay emails  4) UTM every link",
    "StreamYard": "1) Create account  2) Connect LinkedIn Live + YouTube  3) Test multistream before first live",
    "Podcast Guesting": "1) Build a target list of PM / construction / engineering podcasts  2) Use the M12 pitch template  3) Log every pitch in Community & PR",
    "YouTube Podcasts": "1) Enable podcast RSS in YouTube Studio once the show exists",
    "Partnership / PR Outreach": "1) This is the platform value for ALL partnership work  2) The target's type goes in the Org Type column  3) Log rows in Partnership Pipeline / Community & PR",
    "Journalist Requests (Qwoted / Featured)": "1) Create expert-source profiles on Qwoted and Featured  2) Answer only where PCI has real expertise  3) Log placements in Community & PR",
    "Press Release Distribution": "1) Only for credential/programme news  2) Draft through Message Bank approval  3) Publish on own newsroom page first (NewsArticle schema)  4) PRLog free account, submit with image + UTM link  5) openPR free (1 per 30 days)  6) Log placements in PR & Target Directory",
    "LinkedIn Ads": "1) Create Campaign Manager account on the Company Page  2) Install Insight Tag on the website  3) Start with event promotion + engaged-follower retargeting only  4) Set a monthly cap in Channel Costs",
    "Google Ads": "1) Create account, link GA4 + Search Console  2) Start with exact-match certification queries + brand terms  3) Conversion = application started  4) Set a monthly cap in Channel Costs",
    "Meta Ads": "1) Business Manager + pixel  2) Retargeting and Certuvo early-career audiences only  3) Set a monthly cap in Channel Costs",
    "Microsoft Ads": "1) Import the Google Ads campaigns  2) Desktop B2B overflow only",
    "Email Marketing (ESP)": "1) Choose the ESP (owned list)  2) Import consented contacts only - record lawful basis  3) Set up welcome + nurture sequences  4) Every send logs here, whatever the newsletter surface",
    "WhatsApp Business API / SMS": "1) Business verification  2) Opt-in only - exam, webinar and application reminders  3) Template approval before sending",
    "Affiliate / Referral Programme": "1) Define alumni referral terms  2) Unique codes per referrer  3) Track redemptions monthly",
    "Website / Blog": "1) Confirm CMS access + author bio template  2) Blog index page live  3) RSS feed on  4) GA4 + Search Console verified  5) Publish cadence per Publishing Plan",
    "Bing Webmaster Tools + IndexNow": "1) Verify the domain (2-click import from Search Console)  2) Submit the sitemap  3) Enable IndexNow pings on publish  —  ChatGPT retrieves from Bing's index",
    "AI Answer Engines (ChatGPT / Perplexity / AI Overviews)": "1) Write the 20-prompt list buyers actually ask  2) Run it across ChatGPT / Perplexity / AI Overviews monthly  3) Log which domains get cited  4) Target those exact pages for inclusion",
    "PM World Journal": "1) Read pmworldjournal.com/authors  2) Pitch a named-author monthly series to the editor  3) Every published paper links to the author's PCI bio page",
    "Project Controls Expo (UK / USA / AUS)": "1) Apply to speak: London (Nov), Washington DC (Oct), Melbourne (Nov)  2) Enter the Project Controls Expo Awards  3) Log all contacts in Partnership Pipeline",
    "Source of Sources (SOS)": "1) Subscribe a named PCI expert (free)  2) Answer 3 on-expertise queries weekly, 80-120 words, data-led  3) Log placements in Community & PR",
    "Help a B2B Writer": "1) Register spokespeople under Education/Business  2) Respond same day - B2B writers close fast",
    "ResponseSource (UK)": "1) Paid subscription scoped to Business / Education / Construction  2) Assign an owner to answer within 2 hours",
    "Project Times": "1) Submit one evergreen article via projecttimes.com/contribute  2) Byline links to the author bio page",
    "eLearning Industry": "1) Create the author profile via /post-here  2) Submit on AI-era credentialing  3) L&D / instructional-design angle, never HR",
    "Training Industry": "1) Submit a bylined article via /article-submit  2) Research criteria for the relevant Top Training Companies list",
    "Coursecheck": "1) Sign up at coursecheck.com/training-providers  2) Embed review capture at the end of each cohort  3) Never incentivise reviews",
    "Project Control Summit": "1) Apply to present a session  2) Sponsor only if a speaking slot is secured",
    "Google Publisher Center (News / Discover)": "1) Add the blog in Publisher Center  2) NewsArticle schema + named authors + dates  3) Consistent publishing cadence - inclusion is algorithmic now",
    "The Digital Project Manager": "1) Pitch an expert contribution  2) Pitch the podcast  3) Target their certification roundup articles for inclusion",
    "PBC Today (UK construction)": "1) Email a 700-word opinion piece on AI in project controls for UK infrastructure",
    "Construction Week ME + CBNME": "1) Build the 10-name Gulf construction journalist list  2) Offer expert commentary on giga-project controls talent",
    "Amazon KDP (authority book)": "1) Publish the handbook (paperback + Kindle)  2) Author Central page links to a free chapter + email capture",
    "ResearchGate / SSRN / Academia.edu": "1) Upload framework whitepapers as preprints under named authors  2) Link each to the canonical PDF on the PCI site",
    "BrightTALK": "1) Get channel pricing  2) Run one quarter of monthly webinars  3) Compare acquisition vs the Zoom-only baseline before renewing",
    "Sessionize + SpeakerHub": "1) Create speaker profiles for two PCI leaders  2) Apply to 3 open CFPs per quarter",
    "PodMatch + MatchMaker.fm": "1) Build one strong guest profile (AI + project controls)  2) Book 2 shows per month  3) Log pitches in Community & PR",
    "Skool": "1) Stand up the free exam-prep community  2) Weekly threads  3) Promote via personal LinkedIn profiles",
    "Digg (2026 relaunch)": "1) Claim / found a project-controls community  2) Seed 2 discussions weekly  3) Never link-drop",
    "Apple Business Connect": "1) Claim listings for any physical office / exam centre  2) Mirror the Google Business Profile categories exactly",
    "beehiiv Recommendations": "1) If the newsletter migrates, set 5 reciprocal recommendations with PM / construction newsletters  2) Test a small Boosts budget",
    "Eng-Tips Forums": "1) One named staff account  2) Answer cost / schedule questions occasionally  3) Profile link only",
    "Scribd + Issuu": "1) Upload syllabus guides and the annual report with branded CTAs  2) Never bulk-submit - that is link spam",
    "Flipboard / Surf": "1) Connect the blog RSS to a Flipboard magazine  2) Revisit in 6 months and check referrals",
    "GPT Store (custom GPT)": "1) Build one exam-prep GPT linking to Certuvo  2) Measure referral clicks before investing more",
    "Credential Engine Registry": "1) Open a free organization account (accounts@credentialengine.org)  2) Publish PCL-AI, PFL-AI and PML-AI in CTDL via manual entry or the Registry Assistant API  3) Confirm the credentialfinder.org listings appear  4) Keep credential facts in sync with the website",
    "Google Knowledge Panel & Entity Graph": "1) Pick the entity home page  2) EducationalOrganization JSON-LD with sameAs to every profile  3) Create/complete the Wikidata record (add ISNI, Crunchbase, LinkedIn IDs)  4) Enforce one canonical name/description everywhere  5) When the panel appears, claim it via 'Claim this knowledge panel'",
    "Education Schema Markup (Course List + Credential)": "1) Course List pattern on the certification catalogue page  2) Course + CourseInstance JSON-LD per certification page  3) EducationalOccupationalCredential per credential  4) Validate in the Rich Results Test  —  the retired Course Info format is NOT the goal",
    "CareerOneStop Certification Finder": "1) Email the certification details to info@careeronestop.org  2) Verify the listing and its API propagation  3) Re-verify after any credential change",
    "D&B D-U-N-S Number": "1) Apply free at dnb.com (about 30 business days)  2) One number per legal entity (PCI and Certuvo)  3) Record numbers in the vault, never in this file",
    "Tracxn": "1) Submit both entities at tracxn.com/listyourstartup (free, analyst-reviewed)  2) Check the published profile matches the canonical NAP",
    "AlternativeTo": "1) Submit Certuvo with honest category + alternatives  2) Ask genuine users to like/review  3) Never astroturf",
    "Udemy (funnel course)": "1) Publish a low-priced intro course (AI-driven project controls)  2) Links only where the Promotions Policy allows  3) Course closes with the Certuvo/PCI next step",
    "Dealroom": "1) Add the company free at dealroom.co/for-builders  2) Claim and complete the profile",
    "Magnitt": "1) Create/claim the company profile  2) Complete the Gulf market-entry facts  3) Keep funding data current",
    "F6S": "1) Free company profile  2) Watch relevant grant/accelerator programmes quarterly",
    "SaaSHub": "1) Submit Certuvo (free)  2) Maintain the alternatives listing honestly",
    "AI Tool Directories (TAAFT, Toolify, FutureTools)": "1) Only if Certuvo ships real AI features  2) TAAFT first (free submission)  3) Toolify/FutureTools second wave  4) Never pay $497-class listing fees before free directories prove qualified traffic",
    "ISNI": "1) Register via Bowker (~$5 one-time)  2) Add the ISNI to Wikidata and the Organization schema identifier",
    "PitchBook (passive)": "1) No self-serve route  2) Keep incorporation/funding news publicly findable  3) Respond to any PitchBook data request promptly",
    "Bayt.com (employer profile)": "1) Free employer registration  2) Complete the company profile  3) Seed credential-preferred language in partner job posts",
    "OpenCorporates (verify-only)": "1) Find the Delaware record at opencorporates.com/companies/us_de/  2) Verify the name matches the canonical NAP exactly  3) Done — no maintenance",
    "LEI (GLEIF)": "1) Optional: register via a GLEIF-accredited LOU (~€60-70/yr)  2) Renew annually or lapse visibly",
    "StartupBlink": "1) Add the startup to the free map  2) No further effort",
    "Startup Ranking": "1) Free listing (long queue)  2) Do not pay to fast-track",
    "Wellfound (AngelList)": "1) Create the Certuvo profile only when hiring  2) Post real roles",
    "Glassdoor / Indeed Employer Pages": "1) One-time free employer profiles  2) Align name/description with the canonical NAP",
    "Naukrigulf (employer)": "1) Employer zone registration  2) Light presence; review quarterly",
    "GulfTalent (employer)": "1) Only after Bayt shows Gulf demand",
    "LinkedIn Learning Instructor": "1) Founder/SME applies at learning.linkedin.com/instructors with a sample video  2) Warm intro via existing instructors if possible",
    "Google Scholar (named authors)": "1) After staff publish: personal profile with PCI affiliation + institutional email  2) Keep publication list current",
    "Zenodo (DOI for frameworks)": "1) Free account at zenodo.org  2) Create the 'PCI AI' community  3) Upload each framework/whitepaper PDF with named authors + abstract  4) Record the DOI on the source page and cite it in articles",
    "OER Commons + MERLOT": "1) Free member accounts (oercommons.org + merlot.org)  2) Publish one CC-licensed exam-prep primer per quarter via Open Author  3) MERLOT: CC BY-NC-SA licence required  4) Link back to the credential page as the source",
    "TrainingZone + HRZone (UK L&D)": "1) Contributor profile + directory entry  2) Pitch editor@hrzone.com against their 2026 editorial themes  3) One practical L&D-angle article per quarter  4) Byline links to the author page on PCI's site",
    "Snapchat": "1) Business account + Saudi-market profile  2) Ads Manager with KSA targeting (25-45 professionals)  3) Test one lead-gen campaign against LinkedIn CPA before scaling  4) Organic only as campaign support",
    "Careers Page + Google for Jobs": "1) Careers page on the main site, one page per open role  2) JobPosting JSON-LD on each role page (validate in Rich Results test)  3) Submit sitemap in Search Console  4) Remove the page the day a role closes",
    "Naukri.com (employer)": "1) Free employer registration  2) Institute/company page with brand assets  3) Post India-relevant roles (Rs 400-1,650 per post)  4) Mirror Gulf-relevant roles to Naukrigulf",
    "Jobberman (employer)": "1) Free employer registration (Nigeria/Ghana)  2) Company profile only — do not pay  3) Post only if a West Africa ambassador role opens",
}

for r in range(4, 80):          # clear old data area
    for c in range(1, 18):
        ps.cell(r, c).value = None
ps.cell(3, 15).value = "For (brand)"
ps.cell(3, 15)._style = copymod.copy(ps.cell(3, 14)._style)
ps.cell(3, 16).value = "Value rank (1 = most valuable)"
ps.cell(3, 16)._style = copymod.copy(ps.cell(3, 14)._style)
ps.cell(3, 17).value = "Strongest in (countries)"
ps.cell(3, 17)._style = copymod.copy(ps.cell(3, 14)._style)

for i, (name, area, prio, new, note) in enumerate(PLATFORMS):
    r = 4 + i
    prev = old.get(name)
    ps.cell(r, 1).value = i + 1
    ps.cell(r, 2).value = name
    ps.cell(r, 3).value = area
    ps.cell(r, 4).value = prio
    ps.cell(r, 16).value = PLATFORM_VALUE_RANKS[name]
    ps.cell(r, 17).value = GEO.get(name, GEO_DEFAULT)
    if name in CERTUVO_PLATFORMS:
        ps.cell(r, 15).value = "Certuvo (exam prep)"
    if prev:
        for c in range(5, 15):                    # E..N preserved
            ps.cell(r, c).value = prev[c - 1]
        if not prev[4]:
            ps.cell(r, 5).value = SETUP_STEPS.get(name, "")
    else:
        ps.cell(r, 5).value = SETUP_STEPS.get(name, "")
        ps.cell(r, 9).value = "No"                # 2FA
        ps.cell(r, 10).value = "Not Started"      # Status
        ps.cell(r, 11).value = 0                  # Profile %

# ------------------------------- 3. Platform Progress + Who Did What extension
# G/I/K are regenerated from the V6 row-4 originals (row-token substitution) so the
# array idioms LibreOffice already evaluates are preserved exactly, then G gains a
# no-activity guard so a fresh file doesn't show "Days since: 46251".
import re
_v6 = openpyxl.load_workbook(SRC)
_p6 = _v6["Platform Progress"]
V6_G, V6_I, V6_K = (_p6.cell(4, c).value for c in (7, 9, 11))

def rerow(formula, r):
    return re.sub(r'(\$?[A-Z]{1,2}\$?)4(?![0-9])', lambda m: f"{m.group(1)}{r}", formula)

prog = wb["Platform Progress"]
for i in range(NP):
    r = 4 + i
    s = 4 + i                                     # matching Platform Setup row
    prog.cell(r, 1).value = f"='Platform Setup'!$B${s}"
    prog.cell(r, 2).value = f"='Platform Setup'!$C${s}"
    prog.cell(r, 3).value = f"='Platform Setup'!$J${s}"
    prog.cell(r, 4).value = f"=COUNTIF('DAILY ENTRY'!$C$7:$C$1203,$A{r})"
    prog.cell(r, 5).value = (f"=SUMIF('DAILY ENTRY'!$C$7:$C$1203,$A{r},"
                             f"'DAILY ENTRY'!$F$7:$F$1203)")
    prog.cell(r, 6).value = (f"=SUMIF('DAILY ENTRY'!$C$7:$C$1203,$A{r},"
                             f"'DAILY ENTRY'!$G$7:$G$1203)")
    # last-worked guard: no activity -> blank (kills the 46251-day noise)
    prog.cell(r, 7).value = f'=IF(D{r}=0,"",{rerow(V6_G, r)[1:]})'
    prog.cell(r, 8).value = f"=IF(G{r}=\"\",\"\",TODAY()-G{r})"
    prog.cell(r, 9).value = f'=IF(D{r}=0,"",{rerow(V6_I, r)[1:]})'
    prog.cell(r, 10).value = (f"=COUNTIFS('DAILY ENTRY'!$C$7:$C$1203,$A{r},"
                              f"'DAILY ENTRY'!$I$7:$I$1203,\"Blocked\")")
    prog.cell(r, 11).value = rerow(V6_K, r)

wdw = wb["Who Did What"]
for i in range(NP):
    r = 4 + i
    s = 4 + i
    wdw.cell(r, 1).value = f"='Platform Setup'!$B${s}"
    for slot in range(10):
        col = 2 + slot
        cl = get_column_letter(col)
        wdw.cell(r, col).value = (f"=IF({cl}$3=\"(empty slot)\",\"\","
                                  f"COUNTIFS('DAILY ENTRY'!$C$7:$C$1203,$A{r},"
                                  f"'DAILY ENTRY'!$B$7:$B$1203,{cl}$3))")
    wdw.cell(r, 12).value = f"=COUNTIF('DAILY ENTRY'!$C$7:$C$1203,$A{r})"

# The V6 table styles stop at row 49; our extension must carry them to row 103 or
# the sheets read as two files glued together. Copy styles by row parity (zebra) and
# extend every conditional-format range that ended at the old extent.
def extend_table_styles(ws, ncols, first_new=50, template_even=48, template_odd=49):
    tmpl = {0: [ws.cell(template_even, c)._style for c in range(1, ncols + 1)],
            1: [ws.cell(template_odd, c)._style for c in range(1, ncols + 1)]}
    h = ws.row_dimensions[template_odd].height or 18.75
    for r in range(first_new, SETUP_LAST + 1):
        ws.row_dimensions[r].height = h
        for c in range(1, ncols + 1):
            ws.cell(r, c)._style = copymod.copy(tmpl[r % 2][c - 1])
    fresh = []
    for rng in ws.conditional_formatting:
        sq = re.sub(r"([A-Z]{1,2})49\b", lambda m: f"{m.group(1)}{SETUP_LAST}", str(rng.sqref))
        fresh.append((sq, list(rng.rules)))
    ws.conditional_formatting = type(ws.conditional_formatting)()
    for sq, rules in fresh:
        for rule in rules:
            ws.conditional_formatting.add(sq, rule)

# the new O (brand) and P (value rank) columns take the same row styling as N,
# so the style templates at rows 48/49 exist before the extension runs
for _r in range(4, 50):
    ps.cell(_r, 15)._style = copymod.copy(ps.cell(_r, 14)._style)
    ps.cell(_r, 16)._style = copymod.copy(ps.cell(_r, 14)._style)
    ps.cell(_r, 17)._style = copymod.copy(ps.cell(_r, 14)._style)
extend_table_styles(ps, 17)
extend_table_styles(prog, 11)
extend_table_styles(wdw, 12)
prog.column_dimensions["A"].width = 34
wdw.column_dimensions["A"].width = 34
ps.column_dimensions["O"].width = 20
ps.column_dimensions["P"].width = 12
ps.column_dimensions["Q"].width = 26

# Platform Progress mirrors the value rank so the Monday review sees
# "high-value platform, no activity" at a glance
prog.cell(3, 12).value = "Value rank"
prog.cell(3, 12)._style = copymod.copy(prog.cell(3, 11)._style)
for _r in range(4, SETUP_LAST + 1):
    prog.cell(_r, 12).value = f"=IF($A{_r}=\"\",\"\",'Platform Setup'!$P${_r})"
    prog.cell(_r, 12)._style = copymod.copy(prog.cell(_r, 4)._style)
prog.column_dimensions["L"].width = 11

# top-10 value platforms glow gold on Platform Setup so nobody buries them
GOLDFILL = PatternFill("solid", fgColor="FFE699")
for _colrng in (f"B4:B{SETUP_LAST}", f"P4:P{SETUP_LAST}"):
    ps.conditional_formatting.add(
        _colrng, FormulaRule(formula=[f'AND($P4<>"",$P4<=10)'], fill=GOLDFILL))

# ----------------------------------------- 4. Rename sweep across value cells
log = []
log += sweep_rename(wb["Master Tasks"], [3], 4, 66, RENAMES)
log += sweep_rename(wb["Publishing Plan"], [2], 4, 13, RENAMES)
log += sweep_rename(wb["DAILY ENTRY"], [3], 4, 6, RENAMES)
log += sweep_rename(wb["Experiments"], [4], 4, 6, RENAMES)
log += sweep_rename(wb["Content Calendar"], [3], 4, 6, RENAMES)
log += sweep_rename(wb["Community & PR"], [3], 4, 6, RENAMES)
# frequency + segment + org-type string alignment in pre-typed rows
extra = {"Biweekly": "Fortnightly",
         "Project Controls Lead": "Project Controls Manager",
         "Enterprise / Employer": "Employer / Enterprise"}
log += sweep_rename(wb["Master Tasks"], [7], 4, 66, extra)
log += sweep_rename(wb["LinkedIn Outreach"], [8], 4, 6, extra)
log += sweep_rename(wb["Partnership Pipeline"], [4], 4, 6, extra)
# Master Tasks area column: realign to canonical platform areas
area_fix = {"LinkedIn Newsletter": "LinkedIn", "LinkedIn Articles": "LinkedIn",
            "LinkedIn Groups": "LinkedIn"}
mt = wb["Master Tasks"]
plat_area = {p[0]: p[1] for p in PLATFORMS}
for r in range(4, 67):
    plat = mt.cell(r, 3).value
    if isinstance(plat, str) and plat in plat_area:
        mt.cell(r, 2).value = plat_area[plat]

print("renames applied:", len(log))

# ============================== STAGE 2 ==============================
lo = wb["LinkedIn Outreach"]
pp = wb["Partnership Pipeline"]
db = wb["Dashboard"]
de = wb["DAILY ENTRY"]

# ---------------------------------------------------- 5. UTM Builder blocker
# SEARCH("?") treats ? as a wildcard, so every link joined with "&" and GA4
# never parsed a parameter. FIND takes no wildcards. Also un-nests campaign/
# content from inside the medium's SUBSTITUTE (a latent bracket bug).
utm = wb["UTM Builder"]
for r in range(4, 104):
    utm.cell(r, 6).value = (
        f'=IF(OR($A{r}="",$B{r}="",$C{r}=""),"",$A{r}'
        f'&IF(ISNUMBER(FIND("?",$A{r})),"&","?")'
        f'&"utm_source="&LOWER(SUBSTITUTE(TRIM($B{r})," ","-"))'
        f'&"&utm_medium="&LOWER(SUBSTITUTE(TRIM($C{r})," ","-"))'
        f'&IF($D{r}="","","&utm_campaign="&LOWER(SUBSTITUTE(TRIM($D{r})," ","-")))'
        f'&IF($E{r}="","","&utm_content="&LOWER(SUBSTITUTE(TRIM($E{r})," ","-"))))')
utm.cell(4, 7).value = "LinkedIn outreach messages  (EXAMPLE ROW — delete)"

# --------------------------- 6. LinkedIn Outreach: funnel-to-revenue columns
NEWCOLS = [  # (col, header, kind)  kind: entry | auto
    (30, "Accepted date", "entry"),          # AD
    (31, "Message sent date", "entry"),      # AE
    (32, "Follow-up 2 due (auto)", "auto"),  # AF
    (33, "Meeting date", "entry"),           # AG
    (34, "Handed to (PCI closer)", "entry"), # AH
    (35, "Application date", "entry"),       # AI
    (36, "Certification", "entry"),          # AJ
    (37, "Purchase date", "entry"),          # AK
    (38, "Revenue", "entry"),                # AL
    (39, "PCI order / application ref", "entry"),  # AM
    (40, "Duplicate? (auto)", "auto"),       # AN
]
hdr_style = lo.cell(3, 29)._style
for col, header, kind in NEWCOLS:
    c = lo.cell(3, col)
    c.value = header
    c._style = copymod.copy(hdr_style)
ARIAL9 = Font(name="Arial", size=9)
for r in range(4, 1204):
    lo.cell(r, 32).value = f'=IF($AE{r}="","",$AE{r}+10)'
    lo.cell(r, 40).value = (f'=IF($D{r}="","",IF(COUNTIF($D$4:$D$1203,$D{r})>1,'
                            f'"DUPLICATE — check earlier row",""))')
    lo.cell(r, 32).font = ARIAL9
    lo.cell(r, 40).font = ARIAL9
    # follow-up 1 due: from the message-sent date when present, else legacy
    lo.cell(r, 21).value = (f'=IF($AE{r}<>"",$AE{r}+4,'
                            f'IF(OR(R{r}<>"Yes",A{r}=""),"",A{r}+4))')
    # funnel stage: adds a real won stage and renames closed-lost
    lo.cell(r, 29).value = (
        f'=IF($C{r}="","",IF($S{r}="Converted","8 Won / Certified",'
        f'IF($S{r}="Application Started","7 Certification / Enterprise opportunity",'
        f'IF($S{r}="Meeting Booked","6 Meeting / Application",'
        f'IF(OR($S{r}="Interested",$S{r}="Info Requested"),"5 Interested",'
        f'IF(OR($S{r}="Declined",$S{r}="Not Relevant",$S{r}="Do Not Contact / Unsubscribed"),'
        f'"9 Closed - Lost",'
        f'IF($R{r}="Yes","4 Contacted",IF($N{r}="Yes","2 Engaged",'
        f'IF($M{r}="Yes","1 Awareness",'
        f'IF($AA{r}>=60,"3 Qualified","1 Awareness"))))))))))')
lo.conditional_formatting.add(
    "D4:D1203",
    FormulaRule(formula=['AND($D4<>"",COUNTIF($D$4:$D$1203,$D4)>1)'],
                fill=PatternFill("solid", fgColor="F4CCCC")))

# Partnership Pipeline: deal economics + duplicate defence
pp.cell(3, 22).value = "Deal value (number)"
pp.cell(3, 23).value = "Contract signed date"
for c in (22, 23):
    pp.cell(3, c)._style = copymod.copy(pp.cell(3, 21)._style)
pp.conditional_formatting.add(
    "C4:C403",
    FormulaRule(formula=['AND($C4<>"",COUNTIF($C$4:$C$403,$C4)>1)'],
                fill=PatternFill("solid", fgColor="F4CCCC")))

# --------------------------------------------------------- 7. Dashboard fixes
# every formula anywhere that still reads the old 46-row Platform Setup extent
_ps_re = re.compile(r"('Platform Setup'!\$[A-Z]{1,2}\$4:\$[A-Z]{1,2}\$)49")
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and _c.value.startswith("=") \
                    and "'Platform Setup'" in _c.value:
                _c.value = _ps_re.sub(lambda m: f"{m.group(1)}{SETUP_LAST}", _c.value)
# the platform target must track the estate, not a hardcoded 46
db.cell(17, 3).value = f"=COUNTA('Platform Setup'!$B$4:$B${SETUP_LAST})"
# expected-vs-actual: multiply by active headcount, not one person
HEADS = 'MAX(1,COUNTIF(\'START HERE\'!$B$55:$B$64,"?*"))'
for r in range(42, 50):
    db.cell(r, 3).value = f"='START HERE'!$B${r - 17}*$B$41*{HEADS}"
# follow-ups due: blank and No Response outcomes are also genuinely due
db.cell(37, 2).value = (
    '=COUNTIFS(\'LinkedIn Outreach\'!$U$5:$U$1203,"<="&TODAY(),'
    "'LinkedIn Outreach'!$U$5:$U$1203,\">0\",'LinkedIn Outreach'!$S$5:$S$1203,\"Awaiting Reply\")"
    '+COUNTIFS(\'LinkedIn Outreach\'!$U$5:$U$1203,"<="&TODAY(),'
    "'LinkedIn Outreach'!$U$5:$U$1203,\">0\",'LinkedIn Outreach'!$S$5:$S$1203,\"\")"
    '+COUNTIFS(\'LinkedIn Outreach\'!$U$5:$U$1203,"<="&TODAY(),'
    "'LinkedIn Outreach'!$U$5:$U$1203,\">0\",'LinkedIn Outreach'!$S$5:$S$1203,\"No Response\")")

# area block 8 -> 11 areas, then funnel (9 stages) and a revenue section
area_row_style = [copymod.copy(db.cell(65, c)._style) for c in range(1, 6)]
funnel_style = [copymod.copy(db.cell(78, c)._style) for c in range(1, 3)]
head_style = copymod.copy(db.cell(63, 1)._style)
note_style = copymod.copy(db.cell(74, 1)._style)
for rng in [r for r in list(db.merged_cells.ranges) if r.min_row >= 63]:
    db.unmerge_cells(str(rng))
for r in range(63, 106):       # clear the tail before rewriting
    for c in range(1, 8):
        db.cell(r, c).value = None
db.cell(63, 1).value = "6.  BY AREA  —  where the work stands"
db.cell(63, 1)._style = head_style
hdr = ["Area", "Tasks", "Complete", "Completion", "Platforms live in this area"]
for c, h in enumerate(hdr, 1):
    db.cell(64, c).value = h
for i, area in enumerate(AREAS):
    r = 65 + i
    for c in range(1, 6):
        db.cell(r, c)._style = copymod.copy(area_row_style[c - 1])
    db.cell(r, 1).value = area
    db.cell(r, 2).value = f"=COUNTIF('Master Tasks'!$B$4:$B$66,$A{r})"
    db.cell(r, 3).value = (f"=COUNTIFS('Master Tasks'!$B$4:$B$66,$A{r},"
                           f"'Master Tasks'!$J$4:$J$66,\"Complete\")")
    db.cell(r, 4).value = f"=IFERROR(C{r}/B{r},0)"
    db.cell(r, 5).value = (f"=COUNTIFS('Platform Setup'!$C$4:$C${SETUP_LAST},$A{r},"
                           f"'Platform Setup'!$J$4:$J${SETUP_LAST},\"Complete\")")
NOTE_ROW = 65 + len(AREAS) + 1
FUN_HDR = NOTE_ROW + 2
FUN_COLS = FUN_HDR + 1
FUN_FIRST = FUN_HDR + 2
FUN_NOTE = FUN_FIRST + len(FUNNEL_STAGE)
REV_HDR = FUN_NOTE + 2
REV_FIRST = REV_HDR + 1
db.cell(NOTE_ROW, 1).value = ("MANAGER RULE: activity only counts when it is logged with "
                              "evidence the same day.")
db.cell(NOTE_ROW, 1)._style = note_style
db.cell(FUN_HDR, 1).value = "7.  FUNNEL  —  where every contact currently sits"
db.cell(FUN_HDR, 1)._style = head_style
db.cell(FUN_COLS, 1).value = "Stage"
db.cell(FUN_COLS, 2).value = "LinkedIn leads"
for i, stage in enumerate(FUNNEL_STAGE):
    r = FUN_FIRST + i
    db.cell(r, 1)._style = copymod.copy(funnel_style[0])
    db.cell(r, 2)._style = copymod.copy(funnel_style[1])
    db.cell(r, 1).value = stage
    db.cell(r, 2).value = f"=COUNTIF('LinkedIn Outreach'!$AC$5:$AC$1203,$A{r})"
db.cell(FUN_NOTE, 1).value = "Stages are derived automatically from what has been logged."
db.cell(FUN_NOTE, 1)._style = note_style
db.cell(REV_HDR, 1).value = "8.  REVENUE & CHANNEL COST  (from Outreach, Partnerships, Channel Costs)"
db.cell(REV_HDR, 1)._style = head_style
rev = [
    ("Certification revenue recorded", "=SUM('LinkedIn Outreach'!$AL$5:$AL$1203)"),
    ("Certifications sold (rows with revenue)", '=COUNTIF(\'LinkedIn Outreach\'!$AL$5:$AL$1203,">0")'),
    ("Converted rows missing a PCI order ref", "=COUNTIFS('LinkedIn Outreach'!$S$5:$S$1203,\"Converted\",'LinkedIn Outreach'!$AM$5:$AM$1203,\"\")"),
    ("Enterprise pipeline value (open)", "=SUMIFS('Partnership Pipeline'!$V$5:$V$403,'Partnership Pipeline'!$W$5:$W$403,\"\")"),
    ("Enterprise value signed", "=SUMIFS('Partnership Pipeline'!$V$5:$V$403,'Partnership Pipeline'!$W$5:$W$403,\"<>\")"),
    ("Monthly channel cost (Channel Costs)", "=SUM('Channel Costs'!$B$4:$B$23)"),
    ("Meetings booked to date", "=COUNTIF('LinkedIn Outreach'!$S$5:$S$1203,\"Meeting Booked\")+COUNTIF('Community & PR'!$N$5:$N$403,\"Meeting Booked\")"),
]
for i, (label, formula) in enumerate(rev):
    r = REV_FIRST + i
    db.cell(r, 1).value = label
    db.cell(r, 2).value = formula
    db.cell(r, 1)._style = copymod.copy(funnel_style[0])
    db.cell(r, 2)._style = copymod.copy(funnel_style[1])
REV_NOTE = REV_FIRST + len(rev)
db.cell(REV_NOTE, 1).value = ("Revenue figures reconcile against PCI platform order references — "
                              "a Converted row without a PCI ref is unverified.")
db.cell(REV_NOTE, 1)._style = note_style

# structural polish on the rewritten tail (Judge 3): merged band headers like
# sections 1-5, no stray fills, uniform row heights, notes never input-yellow
NAVYBAND = PatternFill("solid", fgColor="1F3864")
NONEFILL = PatternFill(fill_type=None)
GREYNOTEF = Font(name="Arial", size=8, italic=True, color="833C00")
for hr in (63, FUN_HDR, REV_HDR):
    db.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=5)
    hc = db.cell(hr, 1)
    hc.fill = NAVYBAND
    hc.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hc.alignment = Alignment(vertical="center")
    db.row_dimensions[hr].height = 21.75
    for c in range(2, 6):
        db.cell(hr, c).fill = NONEFILL
for nr in (NOTE_ROW, FUN_NOTE, REV_NOTE):
    db.cell(nr, 1).fill = NONEFILL
    db.cell(nr, 1).font = GREYNOTEF
    for c in range(2, 6):
        db.cell(nr, c).fill = NONEFILL
for rr in (list(range(65, 65 + len(AREAS))) + list(range(FUN_FIRST, FUN_FIRST + len(FUNNEL_STAGE)))
           + list(range(REV_FIRST, REV_FIRST + 7))):
    db.row_dimensions[rr].height = 18.75
for c, h in enumerate(["Area", "Tasks", "Complete", "Completion",
                       "Platforms live in this area"], 1):
    hc = db.cell(64, c)
    hc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hc.fill = PatternFill("solid", fgColor="2E5C9A")
from openpyxl.formatting.rule import ColorScaleRule
db.conditional_formatting.add(
    f"D65:D{64 + len(AREAS)}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                              mid_type="num", mid_value=0.5, mid_color="FFEB84",
                              end_type="num", end_value=1, end_color="63BE7B"))
# print sanity: manual breaks between sections so no chart or table is sliced
from openpyxl.worksheet.pagebreak import Break
for br in (38, 62, REV_HDR - 1):
    db.row_breaks.append(Break(id=br))

# ------------------------------------------ 8. Employee Score gate blanking
es = wb["Employee Score"]
for r in range(4, 14):
    es.cell(r, 10).value = (f'=IF($A{r}="","",IF($K{r}<>"Enough data","",'
                            f'IF(I{r}>=80,"A - excellent",IF(I{r}>=65,"B - solid",'
                            f'IF(I{r}>=50,"C - needs coaching",IF(I{r}>=30,"D - at risk",'
                            f'"E - not working"))))))')
    old_l = es.cell(r, 12).value
    if isinstance(old_l, str) and old_l.startswith('=IF($A'):
        es.cell(r, 12).value = old_l.replace(
            f'=IF($A{r}="","",',
            f'=IF($A{r}="","",IF($K{r}<>"Enough data","Log activity first — no fair score yet",',
            1) + ")"

# --------------------------------------- 9. Message Bank approval workflow
mb = wb["Message Bank"]
for c, h in ((7, "Status"), (8, "Approved by"), (9, "Approved date")):
    mb.cell(3, c).value = h
    mb.cell(3, c)._style = copymod.copy(mb.cell(3, 6)._style)
for r in range(4, 21):
    code = mb.cell(r, 1).value
    if isinstance(code, str) and code.strip().startswith("M") and code.strip()[1:].isdigit():
        mb.cell(r, 7).value = "Approved"
        mb.cell(r, 8).fill = YELLOW
        mb.cell(r, 9).fill = YELLOW
mb.cell(4, 6).value = ("House rule: keep connection notes under 200 characters (LinkedIn "
                       "allows 300; shorter converts better).")
mb.cell(2, 7).value = ("Manager initials + date make an approval real. New messages enter "
                       "as Draft; only the manager sets Approved; employees copy only "
                       "Approved rows.")
mb.cell(2, 7).font = NOTEFONT

# ------------------------------------------------- 10. QA live auto-signal
qa = wb["QA & Compliance"]
qa.cell(3, 9).value = "Auto signal (live)"
qa.cell(3, 9)._style = copymod.copy(qa.cell(3, 8)._style)
for _r in range(4, 21):
    for _c in (7, 8, 9):
        mb.cell(_r, _c).alignment = Alignment(vertical="top")
        mb.cell(_r, _c).font = Font(name="Arial", size=9)
_howto = wb["How-To Guides"]
for _r in range(4, 18):
    _howto.cell(_r, 8).alignment = Alignment(vertical="top")
qa.cell(5, 9).font = Font(name="Arial", size=9)
qa.cell(7, 9).font = Font(name="Arial", size=9)
qa.cell(5, 9).value = ('=IF(COUNTIF(\'LinkedIn Outreach\'!$Q$5:$Q$1203,">300")=0,'
                       '"Pass","FAIL — over-limit messages exist")')
qa.cell(7, 9).value = ('=IF(COUNTIFS(\'LinkedIn Outreach\'!$S$5:$S$1203,"Declined",'
                       "'LinkedIn Outreach'!$V$5:$V$1203,\">0\")=0,"
                       '"No declined lead shows follow-ups — verify manually",'
                       '"CHECK — declined leads carry follow-up counts")')

# ---------------------------------------------------- 11. Channel Costs sheet
cc = wb.create_sheet("Channel Costs")
cc["A1"] = "CHANNEL COSTS  —  what each channel costs per month, so cost-per-meeting is computable"
cc["A1"].font = Font(bold=True, size=12)
cc["A2"] = "Fill the yellow cells. The Dashboard revenue block reads B4:B23."
cc["A2"].font = NOTEFONT
for c, h in enumerate(["Channel / tool", "Monthly cost", "Owner", "Notes"], 1):
    cc.cell(3, c).value = h
    cc.cell(3, c).font = HDRFONT
seed = ["LinkedIn Sales Navigator seats", "LinkedIn Ads", "Google Ads", "Meta Ads",
        "Email Marketing (ESP)", "Zoom Webinars licence", "Website hosting / WordPress",
        "Design tools (Canva etc.)", "Credly / badge platform", "Other"]
for i, s in enumerate(seed):
    cc.cell(4 + i, 1).value = s
    for c in (2, 3, 4):
        cc.cell(4 + i, c).fill = YELLOW
cc.column_dimensions["A"].width = 36
cc.column_dimensions["B"].width = 14
cc.column_dimensions["C"].width = 16
cc.column_dimensions["D"].width = 40
cc.merge_cells("A1:D1")
cc["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
cc["A1"].fill = PatternFill("solid", fgColor="1F3864")
cc["A1"].alignment = Alignment(vertical="center")
cc.row_dimensions[1].height = 26
cc.merge_cells("A2:D2")
cc["A2"].font = NOTEFONT
for _c in range(1, 5):
    hc = cc.cell(3, _c)
    hc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hc.fill = PatternFill("solid", fgColor="2E5C9A")
for _r in range(4, 14):
    for _c in range(1, 5):
        cc.cell(_r, _c).font = Font(name="Arial", size=9)

# ============================== STAGE 3 ==============================

# ------------------- 12. Weekly Review: selector-independent all-team trend
wr = wb["Weekly Review"]
for c, h in ((20, "All-team: connections sent"), (21, "All-team: accepted"),
             (22, "All-team: positive replies")):
    wr.cell(3, c).value = h
    wr.cell(3, c)._style = copymod.copy(wr.cell(3, 5)._style)
for r in range(4, 30):
    base = (f"'LinkedIn Outreach'!$A$5:$A$1203,\">=\"&$B{r},"
            f"'LinkedIn Outreach'!$A$5:$A$1203,\"<=\"&$C{r}")
    wr.cell(r, 20).value = f"=IF($B{r}=\"\",\"\",COUNTIFS({base},'LinkedIn Outreach'!$M$5:$M$1203,\"Yes\"))"
    wr.cell(r, 21).value = f"=IF($B{r}=\"\",\"\",COUNTIFS({base},'LinkedIn Outreach'!$N$5:$N$1203,\"Yes\"))"
    wr.cell(r, 22).value = (f"=IF($B{r}=\"\",\"\",COUNTIFS({base},'LinkedIn Outreach'!$S$5:$S$1203,\"Interested\")"
                            f"+COUNTIFS({base},'LinkedIn Outreach'!$S$5:$S$1203,\"Info Requested\")"
                            f"+COUNTIFS({base},'LinkedIn Outreach'!$S$5:$S$1203,\"Meeting Booked\")"
                            f"+COUNTIFS({base},'LinkedIn Outreach'!$S$5:$S$1203,\"Converted\"))")
wr.cell(2, 20).value = "These three columns ignore the person selector — use them for team trend charts."
wr.cell(2, 20).font = NOTEFONT
for _c in (20, 21, 22):
    wr.column_dimensions[get_column_letter(_c)].width = 14
    hc = wr.cell(3, _c)
    hc.font = Font(name="Arial", size=9, bold=True)
    hc.alignment = Alignment(wrap_text=True, vertical="center")
    for _r in range(4, 30):
        wr.cell(_r, _c).font = Font(name="Arial", size=9)
if isinstance(wr.cell(2, 1).value, str) and "Show results for" in str(wr.cell(2, 1).value):
    wr.cell(2, 1).value = "Show:"

# ----------------------- 12b. Objective columns: the campaign dimension
# Every log row states WHICH effort it serves, so performance is testable by
# category (honorary outreach vs authority building vs sales ...), not only
# by platform. Feeds the Objective Performance sheet.
# Content Calendar already has an Objective column at F — it becomes the
# dropdown (no second column, no duplication). The other logs get new columns.
# Header style comes from a REAL banded header (never the hidden helper) and
# body cells inherit each row's input styling so borders/fonts/example-grey
# all match the table (Judge-3 blocker).
# (sheet, obj_col, brand_col, last_row, header_src, body_src, example_last)
TAG_SPEC = [(de, 13, 14, 1006, 10, 5, 6),
            (wb["Community & PR"], 20, 21, 403, 19, 12, 4),
            (lo, 41, 42, 1203, 20, 16, 4),
            (pp, 24, 25, 403, 21, 16, 4)]
for ws_, ocol, bcol, last, hsrc, bsrc, exlast in TAG_SPEC:
    for col, name in ((ocol, "Objective"), (bcol, "For (brand)")):
        hdr = ws_.cell(3, col)
        hdr._style = copymod.copy(ws_.cell(3, hsrc)._style)
        hdr.value = name
        for r in range(4, last + 1):
            cell = ws_.cell(r, col)
            cell._style = copymod.copy(ws_.cell(r, bsrc)._style)
            if r > exlast:
                cell.fill = YELLOW
    ws_.column_dimensions[get_column_letter(ocol)].width = 30
    ws_.column_dimensions[get_column_letter(bcol)].width = 26
# DAILY ENTRY title bands must span the new last column
for _rng in [str(r) for r in list(de.merged_cells.ranges)]:
    if _rng in ("A1:J1", "A2:J2"):
        de.unmerge_cells(_rng)
        de.merge_cells(_rng.replace("J", "N"))
for _rng in [str(r) for r in list(wb["Community & PR"].merged_cells.ranges)]:
    if _rng in ("A1:S1", "A2:S2"):
        wb["Community & PR"].unmerge_cells(_rng)
        wb["Community & PR"].merge_cells(_rng.replace("S", "U"))
for _rng in [str(r) for r in list(pp.merged_cells.ranges)]:
    if _rng in ("A1:W1", "A2:W2"):
        pp.unmerge_cells(_rng)
        pp.merge_cells(_rng.replace("W", "Y"))

# the V7-appended right-hand input columns get first-class styling too
# (Judge-3 finding 8): LO AD..AN and PP V/W inherit row styling, inputs go
# yellow, money and date formats applied.
for r in range(4, 1204):
    _src = lo.cell(r, 16)._style
    for c in range(30, 41):
        cell = lo.cell(r, c)
        cell._style = copymod.copy(_src)
        if r > 4 and c not in (32, 40):          # AF + AN are formulas
            cell.fill = YELLOW
        if c in (30, 31, 33, 35, 37):
            cell.number_format = "yyyy-mm-dd"
        elif c == 38:
            cell.number_format = '"$"#,##0'
for r in range(4, 404):
    _src = pp.cell(r, 16)._style
    for c in (22, 23):
        cell = pp.cell(r, c)
        cell._style = copymod.copy(_src)
        if r > 4:
            cell.fill = YELLOW
    pp.cell(r, 22).number_format = '"$"#,##0'
    pp.cell(r, 23).number_format = "yyyy-mm-dd"
mt.cell(3, 15).value = "Objective"
mt.cell(3, 15)._style = copymod.copy(mt.cell(3, 14)._style)
# the Content Calendar example row states its objective in canonical terms
wb["Content Calendar"].cell(4, 6).value = "Authority & Entity Building"

# ----------------------- 12c. For (brand): Content Calendar + Master Tasks
# (log-tab brand columns are created and styled by TAG_SPEC above)
mt.cell(3, 16).value = "For (brand)"
mt.cell(3, 16)._style = copymod.copy(mt.cell(3, 15)._style)
wb["Content Calendar"].cell(3, 19).value = "For (brand)"
wb["Content Calendar"].cell(4, 19).value = "PCI AI - Institute (umbrella)"
# example rows demonstrate the mandatory tags (Judge-2 finding 20) and use
# the seeded roster name so no example trips the tripwire (finding 21)
for _r in (4, 5):
    de.cell(_r, 13).value = "Honorary Certification Outreach"
    de.cell(_r, 14).value = "PCI AI - Institute (umbrella)"
de.cell(6, 13).value = "Content & SEO Growth"
de.cell(6, 14).value = "PCI AI - Institute (umbrella)"
lo.cell(4, 41).value = "Honorary Certification Outreach"
lo.cell(4, 42).value = "PCI AI - Institute (umbrella)"
pp.cell(4, 24).value = "Partnerships & PR"
pp.cell(4, 25).value = "PCI AI - Institute (umbrella)"
wb["Community & PR"].cell(4, 20).value = "Community Presence"
wb["Community & PR"].cell(4, 21).value = "PCI AI - Institute (umbrella)"
for _ws in (lo, pp, wb["Community & PR"], wb["Content Calendar"]):
    if str(_ws.cell(4, 2).value).strip() in ("[Your name]", "[Name]"):
        _ws.cell(4, 2).value = "Employee 1"

# --------------------------------------- 13. Data validations, rebuilt wholesale
ROSTER = "'START HERE'!$B$55:$B$64"
def rebuild_dv(ws, spec):
    ws.data_validations.dataValidation = []
    for formula1, sqref in spec:
        f1 = formula1 if formula1.startswith(('"', "'", "Lists")) else f'"{formula1}"'
        dv = DataValidation(type="list", formula1=f1, allow_blank=True)
        dv.error = "Pick a value from the dropdown — free-typed values fall out of every report."
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        for part in sqref.split():
            dv.add(part)

P = f"Lists!$J$4:$J${3 + NP}"
BR = f"Lists!$O$4:$O${3 + len(BRANDS)}"
rebuild_dv(wb["START HERE"], [(BR, "B69"),
                              ("Lists!$C$4:$C$6", "B104:B113")])
OBJ = f"Lists!$Y$4:$Y${3 + len(OBJECTIVES)}"
rebuild_dv(de, [(ROSTER, "B4:B1006"), (P, "C4:C1006"),
                (f"Lists!$P$4:$P${3 + len(ACTIVITY_TYPES)}", "D4:D1006"),
                ("Lists!$Q$4:$Q$11", "I4:I1006"),
                (OBJ, "M4:M1006"), (BR, "N4:N1006")])
rebuild_dv(lo, [(ROSTER, "B4:B1203"), ("Lists!$G$4:$G$20", "H4:H1203"),
                ("Lists!$C$4:$C$6", "M4:N1203 R4:R1203"),
                ("Lists!$M$4:$M$19", "O4:O1203"), ("Lists!$F$4:$F$14", "S4:S1203"),
                ("Lists!$T$4:$T$8", "Y4:Z1203"), (ROSTER, "AH4:AH1203"),
                ("PCL-AI,PFL-AI,PML-AI", "AJ4:AJ1203"), (OBJ, "AO4:AO1203"),
                (BR, "AP4:AP1203")])
rebuild_dv(pp, [(ROSTER, "B4:B403"), ("Lists!$U$4:$U$11", "D4:D403"),
                ("Lists!$T$4:$T$8", "J4:K403"), ("Lists!$S$4:$S$12", "N4:N403"),
                ("Lists!$A$4:$A$9", "T4:T403"), (OBJ, "X4:X403"), (BR, "Y4:Y403")])
rebuild_dv(wr, [("Lists!$R$4:$R$14", "B2")])
rebuild_dv(mt, [(f"Lists!$E$4:$E${3 + len(AREAS)}", "B4:B66"), (P, "C4:C66"),
                ("Lists!$B$4:$B$7", "F4:F66"), ("Lists!$D$4:$D$10", "G4:G66"),
                (ROSTER, "I4:I66"), ("Lists!$A$4:$A$9", "J4:J66"), (OBJ, "O4:O66"),
                (BR, "P4:P66")])
rebuild_dv(ps, [(f"Lists!$E$4:$E${3 + len(AREAS)}", f"C4:C{SETUP_LAST}"),
                ("Lists!$B$4:$B$7", f"D4:D{SETUP_LAST}"),
                ("Lists!$C$4:$C$6", f"I4:I{SETUP_LAST}"),
                ("Lists!$A$4:$A$9", f"J4:J{SETUP_LAST}"),
                (ROSTER, f"M4:M{SETUP_LAST}"), (BR, f"O4:O{SETUP_LAST}")])
# Publishing Plan B is a curated locked ranking — a dropdown there is a dead
# control (Judge 1 finding 6), so B gets no DV
rebuild_dv(wb["Publishing Plan"], [("Lists!$D$4:$D$10", "G4:G13"),
                                   (ROSTER, "H4:H13"), ("Lists!$A$4:$A$9", "I4:I13")])
rebuild_dv(wb["SEO Clusters"], [(ROSTER, "F4:F103"), ("Lists!$I$4:$I$10", "G4:G103"),
                                ("Lists!$C$4:$C$6", "I4:I103 K4:K103")])
rebuild_dv(wb["Content Calendar"], [(ROSTER, "B4:B403"), (P, "C4:C403"),
                                    ("Lists!$H$4:$H$21", "D4:D403"),
                                    ("Lists!$I$4:$I$10", "J4:J403"),
                                    (BR, "S4:S403"), (OBJ, "F4:F403")])
rebuild_dv(wb["Community & PR"], [(ROSTER, "B4:B403"), (P, "C4:C403"),
                                  ("Lists!$K$4:$K$16", "E4:E403"),
                                  ("Lists!$L$4:$L$16", "J4:J403"),
                                  ("Lists!$M$4:$M$19", "K4:K403"),
                                  ("Lists!$A$4:$A$9", "M4:M403"),
                                  ("Lists!$F$4:$F$14", "N4:N403"), (OBJ, "T4:T403"),
                                  (BR, "U4:U403")])
rebuild_dv(wb["Experiments"], [(ROSTER, "C4:C53"), (P, "D4:D53"),
                               ("Lists!$V$4:$V$9", "Q4:Q53"),
                               ("Lists!$C$4:$C$6", "S4:S53")])
rebuild_dv(qa, [(ROSTER, "E4:E18"),
                ("Pass,Fail,Not checked yet,Not applicable", "G4:G18")])
rebuild_dv(wb["Daily Log"], [(ROSTER, "B4:B403")])
rebuild_dv(mb, [("Draft,Approved,Retired", "G4:G20")])
rebuild_dv(cc, [(ROSTER, "C4:C13")])

# ------------------------------------------------ 14. Example-row honesty
de.cell(2, 1).value = ("One row for each thing you do, logged the same day. The example rows "
                       "below the headers must be REMOVED before live use: right-click the row "
                       "numbers -> Delete row (deleting shifts references correctly; clearing "
                       "the cells and typing over them makes those rows invisible to reports).")
lo.cell(2, 1).value = ("Log every lead the same day. Remove the example row before live use: "
                       "right-click the row number -> Delete row (never just clear and reuse it). "
                       "Duplicates: before logging a lead, Ctrl+F their URL first — if found, stop.")
ex_s = wb["Experiments"]
ex_s.cell(4, 5).value = "EXAMPLE ROW — delete before use: " + str(ex_s.cell(4, 5).value or "")
ex_s.cell(4, 18).value = "EXAMPLE — delete this row"

# ------------------------------------- 15. START HERE: sharing rules + day 1
sh = wb["START HERE"]
sh.cell(19, 2).fill = YELLOW
sh.cell(19, 3).value = "Required — the manager runs the Monday review and signs QA."
sh.cell(19, 3).font = NOTEFONT
sec = Font(bold=True, size=11)
sh.cell(93, 1).value = "9.  HOW THIS FILE IS SHARED  —  one file, whole team. Breaking these rules corrupts everyone's numbers."
sh.cell(93, 1).font = sec
rules = [
 "1.  The file lives in ONE place — a SharePoint / OneDrive folder. Everyone opens it from that link with AutoSave on. No local copies, no email attachments, ever.",
 "2.  Nobody sorts or filters a shared log sheet. To sort for yourself: View -> Sheet View -> New (it sorts only your view).",
 "3.  To claim a row on a log sheet, click the first fully empty row and type the Date first.",
 "4.  Never paste whole rows. Type, or paste values only, into yellow columns.",
 "5.  Calculated pages and columns are protected. If you cannot type somewhere, that is deliberate — nothing is broken.",
 "6.  After the example-row purge, rows are never deleted. Corrections are edits; disputes resolve through SharePoint version history.",
 "7.  One named person owns this workbook (the manager in cell B19): they run the Monday review and the Friday QA pass, and only they edit the roster, targets, Lists, Message Bank and Master Task owners.",
 "8.  Capacity: the outreach log holds ~1,200 leads, the daily log ~1,000 rows and the other logs ~400 rows. The owner archives a dated copy monthly. At sustained full-team volume, the lead log belongs in a CRM — this file then tracks activity, not the pipeline.",
]
for i, t in enumerate(rules):
    sh.cell(94 + i, 1).value = t
    sh.merge_cells(start_row=94 + i, start_column=1, end_row=94 + i, end_column=6)
    sh.cell(94 + i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sh.row_dimensions[94 + i].height = 26
sh.cell(103, 1).value = "10.  DAY-1 CHECKLIST  —  every new team member, ticked by the manager"
sh.cell(103, 1).font = sec
day1 = [
 "Read START HERE fully, including the Golden Rules and section 9",
 "Added to the team roster above by the manager — name matches exactly everywhere",
 "Password-manager access granted (passwords never go in this file)",
 "LinkedIn profile optimised per LinkedIn Playbook step A",
 "Message Bank read — copy only rows with Status = Approved",
 "Daily targets (section 3) understood and agreed",
 "First 5 logged leads reviewed by the manager BEFORE any request is sent",
 "First 3 sent messages reviewed by the manager",
 "Knows the duplicate rule: Ctrl+F the lead's URL before logging",
 "Assigned their Master Tasks rows and QA checks where relevant",
]
sh.cell(103, 2).value = "Done (manager initials + date)"
sh.cell(103, 2).font = Font(name="Arial", size=8, bold=True, color="1F3864")
_thin = Side(style="thin", color="BFBFBF")
for i, t in enumerate(day1):
    sh.cell(104 + i, 1).value = t
    sh.cell(104 + i, 2).fill = YELLOW
    sh.cell(104 + i, 2).border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _c in range(3, 7):
        sh.cell(104 + i, _c).fill = PatternFill(fill_type=None)

# ---------------------------------------------- 16. V7 UPGRADE NOTES sheet
un = wb.create_sheet("V7 UPGRADE NOTES", 1)
un["A1"] = "GROWTH OS V7  —  WHAT WAS TESTED, WHAT WAS MISSING, WHAT CHANGED"
un["A1"].font = Font(bold=True, size=13)
un["A2"] = ("Three independent audit passes were run end to end: every formula, every platform list, and the "
            "whole system as a product a team runs daily. Every finding below is classified as a PLATFORM issue "
            "(a channel or capability the workbook did not cover) or a PROCESS issue (a workflow or governance "
            "gap no spreadsheet can fix by itself).")
un["A2"].font = NOTEFONT
un["A2"].alignment = Alignment(wrap_text=True)
headers = ["#", "Finding", "Type", "Severity", "What V7 does", "What the team must still do"]
for c, h in enumerate(headers, 1):
    un.cell(4, c).value = h
    un.cell(4, c).font = HDRFONT
    un.cell(4, c).fill = PatternFill("solid", fgColor="D9E2F3")
NOTES = [
 ("Why the file showed one employee", "Process", "—",
  "The team roster (START HERE rows 55-64) has ten slots and only slot 1 was ever filled, with the placeholder 'Employee 1'. Every scorecard, dropdown and column is driven from those cells — the design was always multi-employee; the roster was never populated.",
  "Manager replaces 'Employee 1' with a real name and fills a slot per team member. Names must match exactly everywhere."),
 ("UTM links were malformed", "Platform", "Blocker",
  "The link formula used SEARCH('?'), which treats ? as a wildcard, so every URL was joined with '&' instead of '?' — GA4 could never read a single campaign parameter. Fixed with FIND; campaign/content segments un-nested.",
  "Regenerate and re-paste any UTM link created from V6."),
 ("Team targets compared to one person", "Platform", "Blocker",
  "Dashboard 'Expected to date' never multiplied by headcount, so five people were measured against one person's target. Now multiplied by the count of filled roster slots.",
  "Keep the roster accurate — an unfilled slot lowers the expectation."),
 ("No duplicate-lead defence", "Platform", "Blocker",
  "Nothing stopped two employees logging or re-contacting the same person — including someone who had declined. V7 adds a live Duplicate? column and red highlighting on the lead URL and organisation columns.",
  "Rule: Ctrl+F the lead's URL before logging. A red cell means stop and check the earlier row."),
 ("Platform lists disagreed with each other", "Platform", "Blocker",
  "3 platforms were selectable but invisible to every rollup; 'Quora / Quora Spaces' in the Publishing Plan could never match a count; generic 'LinkedIn' in examples/tasks counted nowhere. All lists now come from one canonical 71-platform list; every stray string remapped.",
  "Log only with dropdown values. The dedup rule for each platform is beside it on the Lists tab."),
 ("25 platforms were missing", "Platform", "Major",
  "No paid channels, no email marketing, no webinar delivery, no digital badges (Credly), no profession-specific communities (Planning Planet, PMI), no podcast guesting, no partnership/PR platform value. All added with areas, priorities and setup steps.",
  "Work the new Critical/High rows in Platform Setup: Zoom Webinars, Email Marketing (ESP), Credly, Partnership / PR Outreach, LinkedIn Live, Planning Planet, PMI Community."),
 ("Funnel stopped before money", "Platform", "Major",
  "Tracking ended at 'Application Started'. V7 adds per-lead columns: meeting date, PCI closer, application date, certification, purchase date, revenue and the PCI order reference, plus deal value / contract date on partnerships and a Dashboard revenue block.",
  "Rule: a lead may be marked Converted only when its PCI order/application reference is in the row. The PCI platform stays the ledger; this file reconciles to it."),
 ("Follow-ups fired from the wrong date", "Platform", "Major",
  "Follow-up due = date logged + 4, not date messaged — due before the first message went out. Now driven by the new 'Message sent date' (+4 and +10), and the Dashboard due-count includes blank/no-response outcomes.",
  "Fill 'Message sent date' when the first message goes out."),
 ("Formula sheets were unprotected", "Process", "Blocker",
  "Any employee could type over the Dashboard or a score formula with no warning. All computed sheets and formula columns are now protected (no password); entry columns stay open.",
  "If you cannot type somewhere, that is the protection working — do not remove it."),
 ("No operating model for a shared file", "Process", "Blocker",
  "Nothing said where the file lives or how ten people use it at once. START HERE section 9 now carries the eight sharing rules (one location, AutoSave, no sorting shared logs, claim rows by date, never paste whole rows, no deletions, one owner, monthly archive).",
  "Manager adopts the rules verbatim and names the owner in cell B19."),
 ("No onboarding checklist", "Process", "Major",
  "Rules existed as prose; nothing was tickable. START HERE section 10 is a ten-point Day-1 checklist with a manager sign-off column.",
  "Run it for every joiner; first 5 leads and first 3 messages get manager review before sending."),
 ("Message approval was undefined", "Process", "Major",
  "The rule said 'approved messages only' but nothing recorded approval. Message Bank now has Status / Approved by / Approved date; M1-M15 are grandfathered as Approved pending the manager's initials.",
  "Manager initials and dates every row; new messages enter as Draft."),
 ("Master Tasks had 0 owners, 0 due dates", "Process", "Blocker",
  "63 tasks, none assigned. V7 cannot assign people — that is a management act.",
  "Manager assigns Owner and Due Date to all 63 rows in one sitting. An unowned task is the manager's task."),
 ("QA checklist was unowned", "Process", "Major",
  "All 15 compliance checks had no owner and no results. Two checks now compute themselves live (character limits; declined-lead follow-ups); the rest need owners.",
  "Assign an owner per QA row on day 1; weekly result before the Monday review."),
 ("Employee Score misled on empty data", "Platform", "Minor",
  "A person with no activity showed grade 'E - not working'. Grades and advice now blank until the activity gate passes.",
  "Score is meaningful only after ~20 connections of history."),
 ("Weekly Review trend depended on the person selector", "Platform", "Major",
  "The Dashboard trend charts read selector-filtered cells, so the 'team trend' was whoever was last selected. Three selector-independent all-team columns added for trend reporting.",
  "Chart team trend from the All-team columns (T:V); use the selector for 1:1s only."),
 ("Fresh-file noise", "Platform", "Minor",
  "'Days since last worked: 46251' on every platform before any logging. Guarded — blank until real activity exists.", ""),
 ("Capacity ceilings", "Platform", "Major",
  "The outreach log holds ~1,200 leads (a 5-person team at target fills it in ~8 working days); logs hold ~400 rows; week list ends at 26. Validation ranges now match counted ranges so nothing silently drops out inside those limits.",
  "Monthly archive by the owner. At sustained full-team volume, move the lead log to a CRM and keep this file for activity and content tracking."),
 ("Channel costs were untracked", "Platform", "Minor",
  "Hours were logged but money was not — cost per meeting was uncomputable. New Channel Costs sheet feeds the Dashboard.",
  "Owner fills monthly costs; review cost-per-meeting monthly."),
 ("Two-brand ambiguity (Certuvo)", "Process", "Minor",
  "Only the Content Calendar has a Brand column. Until the other logs carry one, mixed-brand outreach would corrupt every rate.",
  "Rule: this file is PCI-only. Certuvo work goes in its own copy."),
 ("Workbook recalculates slowly", "Platform", "Minor",
  "The array formulas on Platform Progress and the Daily Log make full recalculation slow in some spreadsheet engines. Guards added in V7 reduce it; expect a pause on open with large data.",
  "Prefer Excel desktop/web; avoid opening in unrelated tools."),
]
r = 5
for i, (title, typ, sev, what, todo) in enumerate(NOTES, 1):
    un.cell(r, 1).value = i
    un.cell(r, 2).value = title
    un.cell(r, 3).value = typ
    un.cell(r, 4).value = sev
    un.cell(r, 5).value = what
    un.cell(r, 6).value = todo
    for c in range(1, 7):
        un.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        un.cell(r, c).font = Font(size=9, bold=(c in (2, 3)))
    r += 1
for col, w in (("A", 4), ("B", 30), ("C", 10), ("D", 9), ("E", 64), ("F", 46)):
    un.column_dimensions[col].width = w
un.sheet_properties.tabColor = "C13329"

# --------------------------------------------------------- 17. Protection
def protect(ws, unlock=(), max_row=None):
    for rng in unlock:
        area = ws[rng]
        if not isinstance(area, tuple):
            area = ((area,),)
        for row in area:
            for cell in row:
                cell.protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False   # focus users onto entry cells
    ws.protection.selectUnlockedCells = False

for name in ("Dashboard", "Summary", "Team Scorecard", "Employee Score",
             "Platform Progress", "Who Did What"):
    protect(wb[name])
protect(de, ["A4:J1006", "M4:N1006"])
protect(lo, ["A4:J1203", "L4:P1203", "R4:T1203", "V4:Z1203",
             "AD4:AE1203", "AG4:AM1203", "AO4:AP1203"])
protect(pp, ["A4:K403", "N4:Y403"])
protect(wb["Content Calendar"], ["A4:S403"])
protect(wb["Community & PR"], ["A4:U403"])
protect(ex_s, ["B4:J53", "L4:M53", "Q4:S53"])
protect(utm, ["A4:E103", "G4:H103"])
protect(wb["SEO Clusters"], ["A4:O103"])
protect(wb["Daily Log"], ["A4:B403"])
protect(wr, ["B2", "N4:Q29"])
protect(mt, ["A4:P66"])
protect(ps, [f"A4:Q{SETUP_LAST}"])
protect(wb["Publishing Plan"], ["G4:J13", "L4:L13"])
protect(qa, ["E4:H18"])
protect(cc, ["B4:D13"])

# ============================== STAGE 4 (formula-agent findings) ==============================

# F2/F3 minimum fix: example rows stop looking like entry rows — grey them out.
GREY = PatternFill("solid", fgColor="E7E6E6")
for ws_, rows, ncol in ((de, (4, 5, 6), 14), (lo, (4,), 42),
                        (pp, (4,), 25), (wb["Content Calendar"], (4,), 19),
                        (wb["Community & PR"], (4,), 21)):
    for rr in rows:
        for c in range(1, ncol + 1):
            ws_.cell(rr, c).fill = GREY
pp.cell(2, 1).value = ("One row per organisation. The grey row below the headers is an "
                       "EXAMPLE — right-click its row number and Delete row before live use. "
                       "Your data starts on the first white row.")

# F5: date and whole-number validation on the columns that silently killed SUMIFS
def add_dv(ws, kind, sqref, operator="greaterThan", f1="36526"):
    dv = DataValidation(type=kind, operator=operator, formula1=f1, allow_blank=True)
    dv.error = ("Enter a real date (not text)" if kind == "date"
                else "Enter a whole number")
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    for part in sqref.split():
        dv.add(part)

add_dv(de, "date", "A4:A1006")
add_dv(de, "whole", "F4:G1006", "greaterThanOrEqual", "0")
add_dv(lo, "date", "A4:A1203 AD4:AE1203 AG4:AG1203 AI4:AI1203 AK4:AK1203")
add_dv(lo, "whole", "J4:J1203 V4:V1203", "greaterThanOrEqual", "0")
add_dv(pp, "date", "A4:A403 Q4:Q403 S4:S403 W4:W403")
add_dv(wb["Content Calendar"], "date", "A4:A403 K4:K403")
add_dv(wb["Content Calendar"], "whole", "M4:P403", "greaterThanOrEqual", "0")
add_dv(wb["Community & PR"], "date", "A4:A403 O4:P403")
add_dv(wb["Daily Log"], "date", "A4:A403")

# F12: typo tripwires — red highlight for any name/platform that will not count
RED = PatternFill("solid", fgColor="F4CCCC")
def tripwire(ws, col, first, last, list_range):
    ws.conditional_formatting.add(
        f"{col}{first}:{col}{last}",
        FormulaRule(formula=[f'AND(${col}{first}<>"",COUNTIF({list_range},${col}{first})=0)'],
                    fill=RED))
for ws_, last in ((de, 1006), (lo, 1203), (pp, 403), (wb["Content Calendar"], 403),
                  (wb["Community & PR"], 403), (wb["Daily Log"], 403)):
    tripwire(ws_, "B", 4, last, "'START HERE'!$B$55:$B$64")
tripwire(de, "C", 4, 1006, f"Lists!$J$4:$J${3 + NP}")
tripwire(wb["Content Calendar"], "C", 4, 403, f"Lists!$J$4:$J${3 + NP}")
tripwire(wb["Community & PR"], "C", 4, 403, f"Lists!$J$4:$J${3 + NP}")

# Judge-2 corrections: the LinkedIn Playbook contradicted the operating model
lp = wb["LinkedIn Playbook"]
for _r in range(4, 19):
    for _c in range(1, 6):
        v = lp.cell(_r, _c).value
        if not isinstance(v, str):
            continue
        v2 = v.replace("Sort the sheet by Follow-up Date to see who is due",
                       "Check Dashboard 'Follow-ups due today or earlier', or sort privately via View -> Sheet View")
        v2 = v2.replace("Daily Log tab, column K", "DAILY ENTRY tab (Activity type: Engagement (commented on someone else))")
        v2 = v2.replace("the Daily Log tab", "the DAILY ENTRY tab")
        v2 = v2.replace("Daily Log tab", "DAILY ENTRY tab")
        v2 = v2.replace("LinkedIn Outreach, column M", "LinkedIn Outreach, column L (Personal line used)")
        v2 = v2.replace("LinkedIn Outreach, column N", "LinkedIn Outreach, column M (Connection sent?)")
        v2 = v2.replace("under 25%", "under 20%")
        if v2 != v:
            lp.cell(_r, _c).value = v2
hg = wb["How-To Guides"]
for _r in range(4, 18):
    v = hg.cell(_r, 5).value
    if isinstance(v, str) and "above 25%" in v:
        hg.cell(_r, 5).value = v.replace("above 25%", "above 30%")
bm = wb["Benchmarks"]
for _r in range(4, 19):
    v = bm.cell(_r, 4).value
    if isinstance(v, str) and "Above 0.10%" in v:
        bm.cell(_r, 4).value = v.replace("Above 0.10%", "Above 0.20%")
v = sh.cell(8, 2).value
if isinstance(v, str) and "Daily Summary tab" in v:
    sh.cell(8, 2).value = v.replace("Daily Summary tab", "Daily Log tab")
# Certuvo clarifier + verified-date column on Lists
_a2 = ps.cell(2, 1)
if isinstance(_a2.value, str) and "Certuvo" not in _a2.value:
    _a2.value = (_a2.value.rstrip() + "  Certuvo-flagged rows are set up from here and "
                 "logged in this file with For (brand) = Certuvo (exam prep).")
if isinstance(_a2.value, str) and "Value rank" not in _a2.value:
    _a2.value = (_a2.value.rstrip() + "  Value rank (column P): 1 = most valuable — "
                 "priority tier first, curated order within the tier. Work the lowest "
                 "numbers first; the gold rows are the top 10.")
ls.cell(3, 24).value = "Verified"
ls.cell(3, 24).font = HDRFONT
for _i in range(NP):
    ls.cell(4 + _i, 24).value = "Aug 2026 - re-verify 6-monthly"
    ls.cell(4 + _i, 24).font = NOTEFONT

# F14: the Daily Log example keyed a non-person and permanently showed "below plan"
dl = wb["Daily Log"]
dl.cell(4, 1).value = None
dl.cell(4, 2).value = None

# F1 mitigation: the distinct-day count was a 1.4M-operation array per cell.
# Two hidden running-count flags on DAILY ENTRY turn it into a plain SUM/SUMIFS:
#   K = 1 on the first row of each date (team-wide), L = 1 on the first row of
#   each (date, person) pair. The running COUNTIF short-circuits on empty rows.
de.cell(3, 11).value = "First entry of day (auto — leave)"
de.cell(3, 11).font = NOTEFONT
de.cell(3, 12).value = "First of day for this person (auto — leave)"
de.cell(3, 12).font = NOTEFONT
for r in range(7, 1007):
    de.cell(r, 11).value = f'=IF($A{r}="",0,IF(COUNTIF($A$7:$A{r},$A{r})=1,1,0))'
    de.cell(r, 12).value = (f'=IF($A{r}="",0,IF(COUNTIFS($A$7:$A{r},$A{r},'
                            f'$B$7:$B{r},$B{r})=1,1,0))')
    de.cell(r, 11).font = Font(name="Arial", size=9)
    de.cell(r, 12).font = Font(name="Arial", size=9)
de.column_dimensions["K"].hidden = True
de.column_dimensions["L"].hidden = True
db.cell(41, 2).value = "=SUM('DAILY ENTRY'!$K$7:$K$1203)"
sm = wb["Summary"]
for rr in range(1, 39):
    v = sm.cell(rr, 2).value
    if isinstance(v, str) and "SUMPRODUCT" in v and "COUNTIFS" in v and "$A$7" in v:
        sm.cell(rr, 2).value = "=SUM('DAILY ENTRY'!$K$7:$K$1203)"
ts_ = wb["Team Scorecard"]
for rr in range(4, 14):
    v = ts_.cell(rr, 2).value
    if isinstance(v, str) and "SUMPRODUCT" in v:
        ts_.cell(rr, 2).value = (f'=IF($A{rr}="","",SUMIFS(\'DAILY ENTRY\'!$L$7:$L$1203,'
                                 f"'DAILY ENTRY'!$B$7:$B$1203,$A{rr}))")

# protection for the two hidden helper columns
for r in range(7, 1007):
    de.cell(r, 11).protection = Protection(locked=True)
    de.cell(r, 12).protection = Protection(locked=True)

# F2 belt-and-braces: extend the daily-entry DV/tripwire down to 1203 was done in
# stage 3; also say where data starts, explicitly.
de.cell(2, 1).value = (str(de.cell(2, 1).value) +
                       "  YOUR DATA STARTS ON ROW 7 — the grey rows above it are examples.")

# UPGRADE NOTES additions from the formula audit
extra_notes = [
 ("Dropdowns did not actually enforce anything", "Platform", "Blocker",
  "Every one of V6's 54 validation rules had error alerts switched off — Excel showed the list but silently accepted anything typed, and mistyped names then vanished from every report. All rules rebuilt with hard errors on, plus red tripwire highlighting on any name or platform that will not count.",
  "If a cell turns red, fix the spelling — red means invisible to every report."),
 ("Example rows sat inside a dead zone", "Platform", "Major",
  "Reports read from row 7 (DAILY ENTRY) / row 5 (logs); the example rows above that carried live yellow cells and dropdowns, so typing over them (instead of deleting them) put real work where no report could see it. Examples are now grey, the notes say exactly where data starts, and the instruction is explicit: delete the rows, never clear them.",
  "Manager deletes all grey example rows in one pass on day 1 and confirms the Dashboard reads zero."),
 ("Date and number columns accepted text", "Platform", "Major",
  "Dates and counts had no validation, and a text date silently drops the row from weekly buckets. Date and whole-number validation added on every date/quantity column.", ""),
 ("Hidden helper columns added", "Platform", "Minor",
  "DAILY ENTRY columns K and L are hidden calculation flags (first entry per day, and per day per person) that drive the days-logged counts efficiently.",
  "Do not unhide, edit or delete DAILY ENTRY columns K and L."),
]
r = un.max_row + 1
for i, (title, typ, sev, what, todo) in enumerate(extra_notes, len(NOTES) + 1):
    un.cell(r, 1).value = i
    un.cell(r, 2).value = title
    un.cell(r, 3).value = typ
    un.cell(r, 4).value = sev
    un.cell(r, 5).value = what
    un.cell(r, 6).value = todo
    for c in range(1, 7):
        un.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        un.cell(r, c).font = Font(size=9, bold=(c in (2, 3)))
    r += 1

# F1: Daily Log carried 5,600 array formulas over 400 mostly-empty rows — the
# heaviest single block, on an explicitly optional sheet. 120 rows is a quarter
# at 10 people; the manager copies the last row down if ever needed.
dl2 = wb["Daily Log"]
for r in range(124, 404):
    for c in range(1, 17):
        dl2.cell(r, c).value = None
dl2.cell(2, 1).value = ("You do not have to use this page. It holds 120 formula rows — enough "
                        "for one quarter at 10 people; the manager copies the last row down if "
                        "more are needed. All targets and scores read DAILY ENTRY, not this sheet.")

# DAILY ENTRY's true capacity is row 1006 (its formatted extent); every consumer
# range aligns to it so no half-alive zone can exist.
_de_re = re.compile(r"('DAILY ENTRY'!\$[A-Z]{1,2}\$7:\$[A-Z]{1,2}\$)1203")
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and _c.value.startswith("=") \
                    and "'DAILY ENTRY'" in _c.value:
                _c.value = _de_re.sub(lambda m: f"{m.group(1)}1006", _c.value)

# Judge-1 #6: the delete-the-example-rows instruction must actually work under
# protection — fully unlock the example rows and allow row deletion.
for _ws, _rows, _ncols in ((de, (4, 5, 6), 12), (lo, (4,), 40), (pp, (4,), 23),
                           (wb["Content Calendar"], (4,), 19),
                           (wb["Community & PR"], (4,), 19)):
    for _r in _rows:
        for _cn in range(1, _ncols + 1):
            _ws.cell(_r, _cn).protection = Protection(locked=False)
    _ws.protection.deleteRows = False

# Judge-1 #7: money columns take numbers only
add_dv(lo, "decimal", "AL4:AL1203", "greaterThanOrEqual", "0")
add_dv(pp, "decimal", "V4:V403", "greaterThanOrEqual", "0")
add_dv(cc, "decimal", "B4:B13", "greaterThanOrEqual", "0")

# Judge-1 #11/#12: stale labels + programme start date as a real date
db.cell(39, 1).value = "4.  ACTIVITY VERSUS DAILY TARGET  (from DAILY ENTRY)"
for _r in range(1, 39):
    v = sm.cell(_r, 3).value
    if isinstance(v, str) and "Daily Log" in v:
        sm.cell(_r, 3).value = v.replace("Read the Daily Log Blocker column",
                                         "Read the DAILY ENTRY Notes/blocker column")
import datetime as _dt
if isinstance(sh.cell(20, 2).value, str):
    try:
        sh.cell(20, 2).value = _dt.datetime.strptime(sh.cell(20, 2).value.strip(), "%Y-%m-%d")
        sh.cell(20, 2).number_format = "yyyy-mm-dd"
    except ValueError:
        pass
add_dv(sh, "date", "B20")

# print/export sanity (Judge 3): bounded print areas so a 1,200-row grid does
# not print as ten empty yellow pages, and repeated headers everywhere
PRINT_AREAS = {"DAILY ENTRY": "A1:N70", "LinkedIn Outreach": "A1:AP65",
               "Partnership Pipeline": "A1:Y65", "Content Calendar": "A1:S65",
               "Community & PR": "A1:U65", "UTM Builder": "A1:H40",
               "Daily Log": "A1:P40", "Experiments": "A1:S30", "SEO Clusters": "A1:O45"}
for _n, _a in PRINT_AREAS.items():
    wb[_n].print_area = _a
for _n in ("UTM Builder", "Daily Log", "Experiments", "SEO Clusters"):
    wb[_n].print_title_rows = "1:3"
from openpyxl.worksheet.pagebreak import Break as _Br
sh.row_breaks.append(_Br(id=102))
# the outreach title bands must span the full 42-column table
for _rng in [str(r) for r in list(lo.merged_cells.ranges)]:
    if _rng in ("A1:X1", "A2:X2"):
        lo.unmerge_cells(_rng)
        lo.merge_cells(_rng.replace("X", "AP"))

wb.save(OUT)
print("stage 4 saved", OUT)
