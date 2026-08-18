#!/usr/bin/env python3
"""Static semantic analysis — the 'silent zero' hunt.

A SUMIFS/COUNTIFS whose criteria literal no longer exists in the list it is
matching against returns 0 forever, with no error. Nothing in a recalc catches
it. This parses every conditional aggregate, resolves which column it filters,
and checks the literal against that column's real vocabulary.
"""
import re
import sys
from collections import defaultdict

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "PCI_AI_Growth_OS_FINAL.xlsx"
wb = openpyxl.load_workbook(SRC)

# ---- build the real vocabulary of every data column that gets filtered on
DATA = {  # sheet -> (first data row, last data row)
    "DAILY ENTRY": (7, 1006), "LinkedIn Outreach": (5, 1203),
    "Content Calendar": (5, 403), "Community & PR": (5, 403),
    "Partnership Pipeline": (5, 403), "Job Postings": (5, 203),
    "Link Building": (5, 403), "Content Scheduler": (5, 103),
    "Master Tasks": (4, 66), "Platform Setup": (4, 136),
}
ls = wb["Lists"]
LOV = {}
for c in range(1, 26):
    vals = [ls.cell(r, c).value for r in range(4, 400)]
    vals = [v for v in vals if v not in (None, "")]
    if vals:
        LOV[openpyxl.utils.get_column_letter(c)] = set(map(str, vals))

# which Lists column supplies each data column, taken from the sheet's own DVs
supply = {}
for ws in wb.worksheets:
    if ws.title not in DATA:
        continue
    for dv in ws.data_validations.dataValidation:
        m = re.match(r"Lists!\$([A-Z])\$", str(dv.formula1 or ""))
        if not m:
            continue
        for rng in str(dv.sqref).split():
            col = re.match(r"([A-Z]{1,2})", rng)
            if col:
                supply[(ws.title, col.group(1))] = m.group(1)

# inline (quoted) validation vocabularies too
inline = {}
for ws in wb.worksheets:
    for dv in ws.data_validations.dataValidation:
        f1 = str(dv.formula1 or "")
        if f1.startswith('"') and f1.endswith('"'):
            vocab = set(f1.strip('"').split(","))
            for rng in str(dv.sqref).split():
                col = re.match(r"([A-Z]{1,2})", rng)
                if col:
                    inline[(ws.title, col.group(1))] = vocab

# ---- parse every conditional aggregate
CRIT = re.compile(
    r"'([^']+)'!\$([A-Z]{1,2})\$\d+:\$[A-Z]{1,2}\$\d+\s*,\s*\"([^\"<>=&]+)\"")
findings = []
checked = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            if not any(k in v for k in ("COUNTIFS", "SUMIFS", "COUNTIF(", "SUMIF(")):
                continue
            for sheet, col, literal in CRIT.findall(v):
                if sheet not in DATA:
                    continue
                checked += 1
                key = (sheet, col)
                vocab = None
                if key in supply:
                    vocab = LOV.get(supply[key])
                elif key in inline:
                    vocab = inline[key]
                if vocab and literal not in vocab:
                    findings.append(
                        (f"{ws.title}!{cell.coordinate}", sheet, col, literal,
                         sorted(vocab)[:4]))

print(f"conditional criteria checked: {checked}")
if findings:
    print(f"\n*** SILENT-ZERO RISKS: {len(findings)} ***")
    seen = set()
    for where, sheet, col, lit, sample in findings:
        k = (sheet, col, lit)
        if k in seen:
            continue
        seen.add(k)
        print(f"  [{where}] filters {sheet}!{col} on {lit!r} — not in that column's "
              f"list (e.g. {sample})")
else:
    print("no silent-zero criteria found")

# ---- duplicate-metric drift: same label computed differently in two places
labels = defaultdict(set)
for ws in wb.worksheets:
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if (isinstance(cell.value, str) and 3 < len(cell.value) < 60
                    and not cell.value.startswith("=")):
                nxt = ws.cell(cell.row, 2).value
                if isinstance(nxt, str) and nxt.startswith("="):
                    labels[cell.value.strip().lower()].add(
                        re.sub(r"\s+", "", nxt))
drift = {k: v for k, v in labels.items() if len(v) > 1}
print(f"\nmetrics with the same label but different formulas: {len(drift)}")
for k, v in list(drift.items())[:8]:
    print(f"  {k!r}: {len(v)} variants")

# ---- ranges that disagree with the declared data extent
END = re.compile(r"'([^']+)'!\$[A-Z]{1,2}\$(\d+):\$[A-Z]{1,2}\$(\d+)")
bad = defaultdict(int)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                for sheet, lo, hi in END.findall(v):
                    if sheet in DATA:
                        want_lo, want_hi = DATA[sheet]
                        if int(hi) != want_hi and int(hi) > 40:
                            bad[(sheet, int(lo), int(hi))] += 1
                        elif int(lo) not in (want_lo, want_lo - 1, 1, 2, 3, 4):
                            bad[(sheet, int(lo), int(hi))] += 1
print(f"\nranges disagreeing with declared extents: {len(bad)}")
for (s, lo, hi), n in sorted(bad.items(), key=lambda x: -x[1])[:10]:
    print(f"  {s} rows {lo}:{hi} used {n}x (declared {DATA[s]})")
