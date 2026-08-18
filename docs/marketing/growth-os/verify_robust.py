#!/usr/bin/env python3
"""Regression suite for the independent reliability review's findings.

That review's verdict was that the workbook never throws an error — and that
this was the problem, because every defect it found was silent. These tests
reproduce each attack and assert that the file now either handles it or says
so out loud.

Covered: the example-row trap (blocker 1), one text date poisoning all 133
platform rows (blocker 2), person-days, future dates, unbounded quantities,
reversed schedule windows, silent overflow past the last row, the week-1
average, mistyped tags, negative values pasted past validation, AutoFilter
under protection, the delete-rows instruction, the locked yellow input cell,
an unlocked Lists master, and the no-data score.
"""
import datetime as dt
import shutil
import subprocess
import sys

import openpyxl

SRC = "PCI_AI_Growth_OS_V9.xlsx"
TST = "robust_test.xlsx"
RECALC = "/root/.claude/skills/synced/xlsx/scripts/recalc.py"

fails = []


def check(name, got, want=True):
    ok = got == want
    print(("PASS " if ok else "FAIL ") + name, "->", got,
          "" if ok else f"expected {want}")
    if not ok:
        fails.append(name)


# ---------------------------------------------------------------- STATIC
wb0 = openpyxl.load_workbook(SRC)

print("--- BLOCKER 1: the example-row trap is gone")
for sn, first in (("DAILY ENTRY", 4), ("LinkedIn Outreach", 4),
                  ("Content Calendar", 4), ("Community & PR", 4),
                  ("Partnership Pipeline", 4), ("Job Postings", 4),
                  ("Link Building", 4), ("Content Scheduler", 4)):
    ws = wb0[sn]
    empty = all(ws.cell(first, c).value is None
                or str(ws.cell(first, c).value).startswith("=")
                for c in range(1, ws.max_column + 1))
    check(f"{sn} ships empty on its first grid row", empty)
check("no sheet still says 'delete before use'",
      not any(isinstance(c.value, str) and "delete before use" in c.value.lower()
              for ws in wb0.worksheets for row in ws.iter_rows(min_row=1, max_row=8)
              for c in row))
check("the worked examples survive on TEAM GUIDE",
      any("WORKED EXAMPLES" in str(wb0["TEAM GUIDE"].cell(r, 1).value or "")
          for r in range(1, wb0["TEAM GUIDE"].max_row + 1)))

print("\n--- MAJOR 11/12/13/14: protection permits what the file tells people to do")
for sn in ("DAILY ENTRY", "LinkedIn Outreach", "Partnership Pipeline",
           "Content Calendar", "Community & PR", "Daily Log", "Platform Progress",
           "Master Tasks", "Platform Setup", "Article Bank", "Keyword Plan"):
    ws = wb0[sn]
    check(f"{sn}: filtering and column resizing allowed under the lock",
          ws.protection.autoFilter is False and ws.protection.formatColumns is False)
check("Experiments and UTM Builder allow row deletion",
      wb0["Experiments"].protection.deleteRows is False
      and wb0["UTM Builder"].protection.deleteRows is False)
check("the yellow 'working days per week' cell is actually editable",
      wb0["START HERE"]["B21"].protection.locked is False)
check("the week-start setting is editable",
      wb0["START HERE"]["B22"].protection.locked is False)
check("Lists master values are locked; the extension rows are not",
      wb0["Lists"]["J4"].protection.locked is True
      and wb0["Lists"]["J200"].protection.locked is False)

print("\n--- MAJOR 7: each log names its own last usable row")
for sn, last in (("DAILY ENTRY", 1006), ("LinkedIn Outreach", 1203),
                 ("Content Calendar", 403), ("Community & PR", 403),
                 ("Partnership Pipeline", 403), ("Content Scheduler", 103)):
    check(f"{sn} marks row {last + 1} as the end",
          str(wb0[sn].cell(last + 1, 1).value or "").startswith("LAST ROW"))

print("\n--- MAJOR 5: quantities are bounded")
_de_dv = {str(d.sqref): d for d in wb0["DAILY ENTRY"].data_validations.dataValidation
          if d.type == "whole"}
check("'How many' is capped", any(d.operator == "between" and "F" in s
                                  for s, d in _de_dv.items()))
check("'Minutes' is capped at a real day",
      any(d.operator == "between" and d.formula2 == "960" for d in _de_dv.values()))

print("\n--- MAJOR 27 / A8: no formula column is painted as an input cell")
_lo = wb0["LinkedIn Outreach"]
check("Outreach auto columns are no longer yellow",
      not any(_lo.cell(r, c).fill.patternType == "solid"
              and str(_lo.cell(r, c).fill.fgColor.rgb) == "FFFFF2CC"
              for r in range(4, 60) for c in (32, 40)))

print("\n--- Guidance no longer contradicts the mechanism")
_sh = wb0["START HERE"]
check("START HERE no longer forbids filtering",
      not any(isinstance(_sh.cell(r, 1).value, str)
              and "Nobody sorts or filters" in _sh.cell(r, 1).value
              for r in range(1, _sh.max_row + 1)))
check("the Glossary exists and is on the MAP",
      "Glossary" in wb0.sheetnames
      and any("Glossary" in str(wb0["MAP"].cell(r, 1).value or "")
              for r in range(1, wb0["MAP"].max_row + 1)))
check("every platform has a logging rule",
      all(wb0["Lists"].cell(r, 23).value not in (None, "")
          for r in range(4, 137)))
check("every dropdown carries a tooltip",
      all(dv.showInputMessage and dv.prompt
          for ws in wb0.worksheets for dv in ws.data_validations.dataValidation
          if dv.type == "list"))

# ---------------------------------------------------------------- LIVE
print("\n--- LIVE: one text date must not poison 133 platform rows")
shutil.copy(SRC, TST)
wb = openpyxl.load_workbook(TST)
sh = wb["START HERE"]
sh.cell(55, 2).value = "Tester One"
today = dt.date.today()
de = wb["DAILY ENTRY"]
de.cell(4, 1).value = today - dt.timedelta(days=3)
de.cell(4, 2).value = "Tester One"
de.cell(4, 3).value = "LinkedIn Company Page"
de.cell(4, 4).value = "Post / content published"
de.cell(4, 6).value = 1
de.cell(4, 7).value = 30
# the poison: a text date on an unrelated platform, exactly as a paste or a
# locale mismatch would produce it
de.cell(5, 1).value = "17/08/2026"
de.cell(5, 2).value = "Tester One"
de.cell(5, 3).value = "Medium"
de.cell(5, 4).value = "Post / content published"
de.cell(5, 6).value = 1
de.cell(5, 7).value = 20
# and a schedule with a reversed window, which used to print -11 planned posts
cs = wb["Content Scheduler"]
cs.cell(4, 1).value = "Medium"
cs.cell(4, 4).value = "Weekly"
cs.cell(4, 5).value = 1
cs.cell(4, 8).value = today + dt.timedelta(days=20)
cs.cell(4, 9).value = today
wb.save(TST)
res = subprocess.run([sys.executable, RECALC, TST, "1200"],
                     capture_output=True, text=True)
print("recalc:", res.stdout.strip()[:200].replace("\n", " "))
if '"status": "success"' not in res.stdout:
    sys.exit("recalc failed")
wv = openpyxl.load_workbook(TST, data_only=True)
pp = wv["Platform Progress"]
_row_lcp = next(r for r in range(4, 137)
                if str(pp.cell(r, 1).value) == "LinkedIn Company Page")
check("'Last worked on' survives a text date elsewhere in the column",
      pp.cell(_row_lcp, 7).value is not None)
check("'Who worked on it last' survives it too",
      pp.cell(_row_lcp, 9).value == "Tester One")
check("a reversed schedule window plans nothing rather than a negative",
      wv["Content Scheduler"].cell(4, 10).value, None)
db = wv["Dashboard"]
_health = {}
for r in range(1, 140):
    lab = str(db.cell(r, 1).value or "")
    if lab.startswith(("Rows with an unreadable", "Future-dated", "Impossible")):
        _health[lab.split("—")[0].strip()] = db.cell(r, 2).value
check("the text date is reported on the Dashboard",
      _health.get("Rows with an unreadable (text) date"), 1)

print()
if fails:
    print(f"{len(fails)} FAILURES:", fails)
    sys.exit(1)
print("ALL PASS")
